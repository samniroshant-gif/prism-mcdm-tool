"""
PRISM — Sustainability MCDM Assessment Tool
Streamlit implementation

Levels:
  1. System definition (alternatives, categories, custom indicators, units, indicator values)
  1.5 Correlation check (within-category Spearman correlation review, informational)
  2. Indicator processing (MEREC normalisation/weights, N2 normalisation, category scores)
  3. Decision aggregation (Equal/Entropy/CRITIC weighting + RCW consolidation,
                            TOPSIS/VIKOR/ELECTRE-Score/MULTIMOORA/WASPAS, PSI compromise ranking,
                            category-combination rank-stability scatter)
  4. Optional validation: weighting-method sensitivity, benefit/cost indicator
     sensitivity, Monte Carlo Dirichlet uncertainty.

Run with:  streamlit run app.py
"""

import itertools
import json
import os
import base64
import secrets as pysecrets
import uuid
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib import font_manager as fm
from matplotlib.colors import LinearSegmentedColormap
import streamlit as st
import io
import plotly.graph_objects as go
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
from scipy.stats import spearmanr

try:
    from supabase import create_client
    SUPABASE_PKG_OK = True
except ImportError:
    SUPABASE_PKG_OK = False

st.set_page_config(
    page_title="PRISM | Sustainability MCDM Assessment",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Cloud-safe Times New Roman lookalike (Tinos) for matplotlib + CSS @font-face
_FONT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")
_TINO_FILES = [
    ("Tinos-Regular.ttf", "normal", 400),
    ("Tinos-Bold.ttf", "normal", 700),
    ("Tinos-Italic.ttf", "italic", 400),
    ("Tinos-BoldItalic.ttf", "italic", 700),
]
_FONT_FACE_CSS = []
for _fname, _style, _weight in _TINO_FILES:
    _fpath = os.path.join(_FONT_DIR, _fname)
    if os.path.isfile(_fpath):
        try:
            fm.fontManager.addfont(_fpath)
        except Exception:
            pass
        try:
            with open(_fpath, "rb") as _fh:
                _b64 = base64.b64encode(_fh.read()).decode("ascii")
            _FONT_FACE_CSS.append(
                f"@font-face{{font-family:'Tinos';font-style:{_style};font-weight:{_weight};"
                f"src:url(data:font/ttf;base64,{_b64}) format('truetype');"
                f"font-display:swap;}}"
            )
        except Exception:
            pass

_FONT_CSS = '"Inter", "Segoe UI", system-ui, -apple-system, sans-serif'
_FONT_FACES = "\n".join(_FONT_FACE_CSS)

# Brand palette — industry / corporate sustainability tooling
_BRAND_NAVY   = "#0D2B5E"
_BRAND_SLATE  = "#4A5568"
_BRAND_MUTED  = "#6B7A99"
_BRAND_TEXT   = "#1A202C"
_BRAND_BORDER = "#DCE3EF"
_BRAND_BG     = "#F7F9FC"
_BRAND_CARD   = "#FFFFFF"
_BRAND_ACCENT = "#1A3A6E"
_BRAND_TEAL   = "#0E7490"

_PSI_HELP = (
    "PSI trade-off (p): values near 0 favour rank stability across MCDM methods; "
    "values near 1 favour average performance. p = 0.50 balances both."
)

FIELD_HELP = {
    "n_alternatives": "Number of options to compare (2–10). Each alternative receives a name and rank.",
    "alternative_name": "Short label for this option — shown in charts, tables, and the Excel report.",
    "category": "Sustainability dimension to include. Select at least one category to proceed.",
    "use_custom": "Add criteria beyond PRISM's predefined indicator library for this assessment.",
    "custom_count": "How many extra indicators to define for this category (0–5).",
    "indicator_name": "Name of the custom criterion as it appears in tables and charts.",
    "indicator_unit": "Measurement unit (e.g. kg CO₂, £, %, dB).",
    "indicator_direction": "Benefit = higher is better; Cost = lower is better for ranking.",
    "indicator_enable": "Include or exclude this predefined indicator from the assessment.",
    "unit_select": "Measurement unit for this indicator in the decision matrix.",
    "custom_unit": "Enter a unit not available in the preset dropdown.",
    "indicator_value": "Raw measured value for this alternative on this indicator.",
    "l3_category": "Category score carried forward into Level 3 MCDM aggregation.",
    "weight_equal": "Treat every selected category as equally important.",
    "weight_entropy": "Weight categories by data dispersion — more variation means higher influence.",
    "weight_critic": "Weight by contrast and correlation — distinct, non-redundant categories score higher.",
    "mcdm_method": "Ranking technique applied to weighted category scores before PSI aggregation.",
    "psi_p": _PSI_HELP,
    "combo_p": "p value used when computing PSI ranks across every category subset combination.",
    "psi_combo_p": "p for the PSI-combo tie-breaker when multiple alternatives share rank 1.",
    "analytics_tool": "Optional post-results analysis — contribution, leave-one-out, or stakeholder simulation.",
    "rr_exclude": "Temporarily remove alternatives to test whether rankings reverse.",
    "alt_norm": "Alternative normalisation method to compare against the default N2 scaling.",
    "unc_pct": "Percentage perturbation applied to indicator values in uncertainty analysis.",
    "n_iter": "Number of Monte Carlo or bootstrap iterations — higher values give smoother estimates.",
    "mc_k": "Dirichlet concentration parameter — lower k means more weight uncertainty.",
    "mc_override": "Manually set the Dirichlet k instead of using the automatic estimate.",
    "ben_pct": "Percentage change applied to benefit-type indicators in sensitivity analysis.",
    "cost_pct": "Percentage change applied to cost-type indicators in sensitivity analysis.",
    "stakeholder_weight": "Relative importance assigned to this category in the stakeholder simulation.",
}

def tnr_label(text, *, color="#1A202C", size="12pt", weight="400", strike=False):
    """Render plain UI text in Times/Tinos (st.text uses a fixed sans font)."""
    deco = "text-decoration:line-through;" if strike else ""
    st.markdown(
        f"<div style='font-family:{_FONT_CSS};font-size:{size};font-weight:{weight};"
        f"color:{color};margin:4px 0;line-height:1.4;{deco}'>{text}</div>",
        unsafe_allow_html=True,
    )

st.markdown(f"""
<style>
{_FONT_FACES}
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
@import url('https://fonts.googleapis.com/css2?family=Tinos:ital,wght@0,400;0,700;1,400;1,700&display=swap');

/* ── Global Times / Tinos (no blanket span — preserves Material Icons) ── */
html, body, .stApp,
[data-testid="stAppViewContainer"],
[data-testid="stHeader"],
[data-testid="stToolbar"],
[data-testid="stSidebar"],
[data-testid="stMarkdownContainer"],
[data-testid="stWidgetLabel"],
[data-testid="stMetricValue"],
[data-testid="stMetricLabel"],
[data-testid="stMetricDelta"],
[data-testid="stCaptionContainer"],
[data-testid="stText"],
[data-testid="stText"] *,
.stMarkdown, .stText, .stCaption, .stAlert, .stTooltipContent,
.stCheckbox, .stRadio, .stSelectbox, .stMultiSelect,
.stNumberInput, .stTextInput, .stTextArea, .stSlider,
.stDownloadButton > button,
[data-baseweb="select"], [data-baseweb="input"],
[data-baseweb="checkbox"], [data-baseweb="radio"],
[data-baseweb="tag"], [data-baseweb="button"],
[data-testid="stDataFrame"] *,
[data-testid="stTable"] *,
[data-testid="stDataEditor"] *,
table, th, td, .stDataFrame, .stTable, .stDataEditor,
div, p, label, input, textarea, select, button, pre {{
    font-family: {_FONT_CSS} !important;
    font-size: 11pt !important;
}}

/* Force Streamlit fixed/monospace text widgets onto Tinos */
[data-testid="stText"],
[data-testid="stText"] pre,
.stText, .stText pre,
pre, code, .stCodeBlock, .stCode {{
    font-family: {_FONT_CSS} !important;
}}

/* ── Checkbox / radio / widget labels ── */
.stCheckbox label,
.stCheckbox p,
.stRadio label,
.stRadio p,
[data-testid="stWidgetLabel"],
[data-testid="stWidgetLabel"] p,
[data-testid="stMarkdownContainer"] p,
[data-testid="stCaptionContainer"],
.stCaption {{
    font-family: {_FONT_CSS} !important;
    font-size: 11pt !important;
}}

/* ── Headings — dark blue, bold ── */
h1, h2, h3, h4, h5, h6,
[data-testid="stMarkdownContainer"] h1,
[data-testid="stMarkdownContainer"] h2,
[data-testid="stMarkdownContainer"] h3,
[data-testid="stMarkdownContainer"] h4,
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 {{
    font-family: {_FONT_CSS} !important;
    font-weight: 700 !important;
    color: #0D2B5E !important;
}}

/* ── Sidebar professional styling ── */
[data-testid="stSidebar"] {{
    background: #F7F9FC !important;
    border-right: 1px solid #DCE3EF !important;
}}
[data-testid="stSidebar"] .stMarkdown p,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p {{
    font-size: 12pt !important;
    color: #4A5568 !important;
}}

/* ── Hide default Streamlit alert box colours — use neutral cards ── */
.stAlert {{
    border-radius: 4px !important;
    border-left: 3px solid #0D2B5E !important;
    background: #F7F9FC !important;
    color: #1A202C !important;
}}

/* ── Dataframe / table styling ── */
[data-testid="stDataFrame"] th,
[data-testid="stTable"] th {{
    background-color: #0D2B5E !important;
    color: #FFFFFF !important;
    font-weight: 700 !important;
    font-family: {_FONT_CSS} !important;
    font-size: 11pt !important;
}}
[data-testid="stDataFrame"] td,
[data-testid="stTable"] td,
[data-testid="stDataFrame"] *,
[data-testid="stTable"] *,
[data-testid="stDataEditor"] * {{
    font-family: {_FONT_CSS} !important;
    color: #1A202C !important;
}}

/* ── Button styling ── */
.stButton > button {{
    font-family: {_FONT_CSS} !important;
    font-size: 12pt !important;
    font-weight: 600 !important;
    border-radius: 4px !important;
}}
.stButton > button[kind="primary"] {{
    background-color: #0D2B5E !important;
    color: #FFFFFF !important;
    border: none !important;
}}
.stButton > button[kind="primary"]:hover {{
    background-color: #1A3A6E !important;
}}

/* ── Expander header label (not icon spans) ── */
[data-testid="stExpander"] summary p,
[data-testid="stExpander"] summary [data-testid="stMarkdownContainer"] {{
    font-family: {_FONT_CSS} !important;
    font-size: 12pt !important;
    font-weight: 600 !important;
    color: #0D2B5E !important;
}}

/* ── Material icons — sidebar, expanders, header buttons ── */
[data-testid="collapsedControl"] span,
[data-testid="stSidebarCollapseButton"] span,
[data-testid="stExpanderToggleIcon"],
[data-testid="stExpanderToggleIcon"] span,
[data-testid="stExpander"] summary > span,
[data-testid="stExpander"] summary span[class*="material"],
[data-testid="stExpander"] .material-icons,
[data-testid="stExpander"] [class*="material-symbols"],
span.material-icons,
span[class*="material-symbols"],
.material-icons,
[class*="material-symbols"] {{
    font-family: "Material Symbols Rounded", "Material Icons" !important;
    font-weight: normal !important;
    font-style: normal !important;
    letter-spacing: normal !important;
    text-transform: none !important;
    white-space: nowrap !important;
}}

/* ── Progress bar ── */
.stProgress > div > div {{
    background-color: #0D2B5E !important;
}}

/* ── Hide Streamlit chrome for a cleaner enterprise shell ── */
#MainMenu {{visibility: hidden;}}
footer {{visibility: hidden;}}
[data-testid="stToolbar"] {{visibility: hidden; height: 0;}}

/* ── Main content area ── */
[data-testid="stAppViewContainer"] > .main {{
    background: {_BRAND_BG} !important;
}}
.block-container {{
    padding-top: 2rem !important;
    padding-bottom: 3rem !important;
    max-width: 1180px !important;
}}

/* ── Metrics — dashboard-style KPI tiles ── */
[data-testid="stMetric"] {{
    background: {_BRAND_CARD} !important;
    border: 1px solid {_BRAND_BORDER} !important;
    border-radius: 8px !important;
    padding: 12px 16px !important;
    box-shadow: 0 1px 3px rgba(13,43,94,0.06) !important;
}}
[data-testid="stMetricLabel"] {{
    color: {_BRAND_MUTED} !important;
    font-size: 11pt !important;
    font-weight: 500 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.04em !important;
}}
[data-testid="stMetricValue"] {{
    color: {_BRAND_NAVY} !important;
    font-weight: 700 !important;
}}

/* ── Subheaders — section dividers ── */
[data-testid="stMarkdownContainer"] h3 {{
    border-bottom: 2px solid {_BRAND_BORDER} !important;
    padding-bottom: 6px !important;
    margin-top: 1.5rem !important;
}}

/* ── Sliders and inputs ── */
.stSlider [data-baseweb="slider"] div {{
    font-family: {_FONT_CSS} !important;
}}

/* ── Professional card utility (landing / info blocks) ── */
.prism-hero {{
    background: linear-gradient(135deg, {_BRAND_NAVY} 0%, #1A3A6E 100%);
    color: #FFFFFF;
    border-radius: 10px;
    padding: 36px 40px;
    margin-bottom: 28px;
    box-shadow: 0 4px 16px rgba(13,43,94,0.18);
}}
.prism-hero h1 {{
    color: #FFFFFF !important;
    font-size: 2rem !important;
    margin: 0 0 8px 0 !important;
    letter-spacing: -0.02em !important;
}}
.prism-hero p {{
    color: rgba(255,255,255,0.88) !important;
    font-size: 1.05rem !important;
    margin: 0 !important;
    line-height: 1.55 !important;
}}
.prism-card {{
    background: {_BRAND_CARD};
    border: 1px solid {_BRAND_BORDER};
    border-radius: 8px;
    padding: 20px 22px;
    height: 100%;
    box-shadow: 0 1px 4px rgba(13,43,94,0.05);
}}
.prism-card h4 {{
    color: {_BRAND_NAVY} !important;
    font-size: 13pt !important;
    margin: 0 0 8px 0 !important;
}}
.prism-card p {{
    color: {_BRAND_SLATE} !important;
    font-size: 11pt !important;
    line-height: 1.6 !important;
    margin: 0 !important;
}}
.prism-step-pill {{
    display: inline-block;
    background: {_BRAND_NAVY};
    color: #FFF;
    font-size: 10pt;
    font-weight: 600;
    padding: 4px 12px;
    border-radius: 20px;
    margin-bottom: 8px;
}}
.prism-workflow-step {{
    text-align: center;
    padding: 16px 8px;
}}
.prism-workflow-step .num {{
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 32px;
    height: 32px;
    background: {_BRAND_NAVY};
    color: #FFF;
    border-radius: 50%;
    font-weight: 700;
    font-size: 13pt;
    margin-bottom: 8px;
}}
.prism-workflow-step .lbl {{
    color: {_BRAND_NAVY};
    font-weight: 600;
    font-size: 11pt;
}}
.prism-workflow-step .sub {{
    color: {_BRAND_MUTED};
    font-size: 10pt;
    margin-top: 4px;
}}
.prism-sidebar-badge {{
    display: inline-block;
    background: #E8EEF7;
    color: {_BRAND_NAVY};
    font-size: 9pt;
    font-weight: 600;
    letter-spacing: 0.06em;
    text-transform: uppercase;
    padding: 3px 10px;
    border-radius: 4px;
    margin-top: 6px;
}}
.prism-about-panel {{
    background: {_BRAND_CARD};
    border: 1px solid {_BRAND_BORDER};
    border-left: 4px solid {_BRAND_NAVY};
    border-radius: 8px;
    padding: 28px 32px;
    margin-bottom: 8px;
    box-shadow: 0 2px 8px rgba(13,43,94,0.06);
}}
.prism-about-title {{
    color: {_BRAND_NAVY};
    font-family: {_FONT_CSS};
    font-size: 14pt;
    font-weight: 600;
    margin: 0 0 16px 0;
    letter-spacing: -0.01em;
}}
.prism-about-body {{
    color: {_BRAND_TEXT};
    font-family: {_FONT_CSS};
    font-size: 11pt;
    line-height: 1.8;
    margin: 0 0 14px 0;
}}
.prism-dimension-row {{
    margin: 4px 0 14px 0;
}}
.prism-dimension-pill {{
    display: inline-block;
    background: {_BRAND_BG};
    border: 1px solid {_BRAND_BORDER};
    color: {_BRAND_NAVY};
    font-family: {_FONT_CSS};
    font-size: 9.5pt;
    font-weight: 500;
    padding: 4px 11px;
    border-radius: 4px;
    margin: 3px 6px 3px 0;
}}
.prism-traffic-badge {{
    display: inline-block;
    font-family: {_FONT_CSS};
    font-size: 11pt;
    font-weight: 600;
    padding: 8px 18px;
    border-radius: 6px;
    margin-bottom: 12px;
}}
.prism-traffic-green {{
    background: #DCFCE7;
    color: #166534;
    border: 1px solid #86EFAC;
}}
.prism-traffic-amber {{
    background: #FEF3C7;
    color: #92400E;
    border: 1px solid #FCD34D;
}}
.prism-traffic-red {{
    background: #FEE2E2;
    color: #991B1B;
    border: 1px solid #FCA5A5;
}}
.prism-traffic-na {{
    background: #F3F4F6;
    color: #4B5563;
    border: 1px solid #D1D5DB;
}}
.prism-framework {{
    margin: 8px 0 24px 0;
}}
.prism-framework-heading {{
    color: {_BRAND_NAVY};
    font-family: {_FONT_CSS};
    font-size: 11pt;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    margin: 0 0 14px 0;
}}
.prism-framework-level {{
    border-radius: 8px;
    padding: 16px 18px;
    margin-bottom: 10px;
    border: 1px solid {_BRAND_BORDER};
}}
.prism-framework-level.level-01 {{ background: #E6F1FB; }}
.prism-framework-level.level-02 {{ background: #E1F5EE; }}
.prism-framework-level.level-03 {{ background: #FAEEDA; }}
.prism-framework-level-title {{
    color: {_BRAND_NAVY};
    font-family: {_FONT_CSS};
    font-size: 10pt;
    font-weight: 700;
    letter-spacing: 0.04em;
    text-transform: uppercase;
    margin-bottom: 12px;
}}
.prism-framework-row {{
    display: flex;
    flex-wrap: wrap;
    align-items: stretch;
    gap: 8px;
    justify-content: center;
}}
.prism-framework-box {{
    background: {_BRAND_CARD};
    border: 1px solid {_BRAND_BORDER};
    border-radius: 6px;
    padding: 10px 12px;
    flex: 1 1 140px;
    max-width: 220px;
    min-width: 120px;
    text-align: center;
    box-shadow: 0 1px 3px rgba(13,43,94,0.05);
}}
.prism-framework-box.wide {{
    flex: 2 1 260px;
    max-width: 340px;
}}
.prism-framework-box.dashed {{
    border-style: dashed;
    background: rgba(255,255,255,0.65);
}}
.prism-framework-box strong {{
    display: block;
    color: {_BRAND_NAVY};
    font-family: {_FONT_CSS};
    font-size: 10pt;
    font-weight: 600;
    line-height: 1.3;
    margin-bottom: 4px;
}}
.prism-framework-box span.sub {{
    display: block;
    color: {_BRAND_SLATE};
    font-family: {_FONT_CSS};
    font-size: 8.5pt;
    line-height: 1.45;
}}
.prism-framework-arrow {{
    display: flex;
    align-items: center;
    justify-content: center;
    color: {_BRAND_MUTED};
    font-size: 14pt;
    font-weight: 700;
    min-width: 20px;
}}
.prism-framework-flow {{
    display: flex;
    flex-wrap: wrap;
    align-items: center;
    justify-content: center;
    gap: 6px;
    margin-bottom: 10px;
}}
.prism-framework-side {{
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    justify-content: center;
    margin-top: 8px;
    padding-top: 10px;
    border-top: 1px dashed rgba(13,43,94,0.15);
}}
.prism-step-row {{
    display: flex;
    align-items: center;
    gap: 10px;
    margin: 3px 0;
    padding: 2px 0;
}}
.prism-step-row.current {{
    background: #E8EEF7;
    border-radius: 6px;
    border-left: 3px solid {_BRAND_NAVY};
    padding: 6px 10px;
    margin: 2px 0;
}}
.prism-step-check {{
    flex-shrink: 0;
    width: 18px;
    height: 18px;
    border-radius: 4px;
    border: 2px solid #CBD5E0;
    box-sizing: border-box;
    display: inline-flex;
    align-items: center;
    justify-content: center;
}}
.prism-step-check.done {{
    background: #22C55E;
    border-color: #16A34A;
    color: #FFFFFF;
    font-size: 13px;
    font-weight: 700;
    line-height: 1;
}}
.prism-step-check.current {{
    border-color: {_BRAND_NAVY};
    background: #FFFFFF;
    box-shadow: inset 0 0 0 2px {_BRAND_NAVY};
}}
.prism-step-check.pending {{
    background: #FFFFFF;
    border-color: #CBD5E0;
}}
.prism-step-label {{
    font-size: 10pt;
    line-height: 1.3;
    font-family: {_FONT_CSS};
}}
.prism-step-label.done {{
    color: #1A202C;
}}
.prism-step-label.current {{
    color: {_BRAND_NAVY};
    font-weight: 600;
}}
.prism-step-label.pending {{
    color: #9CA3AF;
}}
.prism-section-head {{
    border-bottom: 1px solid {_BRAND_BORDER};
    padding-bottom: 4px;
    margin-bottom: 6px !important;
}}

/* ── Branded progress indicator (sidebar) ── */
.prism-progress-wrap {{
    margin: 4px 0 12px 0;
}}
.prism-progress-label {{
    display: flex;
    justify-content: space-between;
    align-items: baseline;
    font-family: {_FONT_CSS};
    font-size: 10pt;
    font-weight: 600;
    color: {_BRAND_NAVY};
    margin-bottom: 6px;
    letter-spacing: 0.02em;
}}
.prism-progress-pct {{
    font-size: 13pt;
    font-weight: 700;
    color: {_BRAND_NAVY};
}}
.prism-progress-track {{
    height: 8px;
    background: {_BRAND_BORDER};
    border-radius: 999px;
    overflow: hidden;
}}
.prism-progress-fill {{
    height: 100%;
    background: linear-gradient(90deg, {_BRAND_NAVY} 0%, {_BRAND_ACCENT} 100%);
    border-radius: 999px;
    transition: width 0.3s ease;
}}
.prism-progress-caption {{
    font-family: {_FONT_CSS};
    font-size: 9pt;
    color: {_BRAND_MUTED};
    margin: 6px 0 0 0;
}}

/* ── Secondary buttons & download ── */
.stButton > button:not([kind="primary"]),
.stDownloadButton > button {{
    background: {_BRAND_CARD} !important;
    color: {_BRAND_NAVY} !important;
    border: 1px solid {_BRAND_BORDER} !important;
}}
.stButton > button:not([kind="primary"]):hover,
.stDownloadButton > button:hover {{
    background: #E8EEF7 !important;
    border-color: {_BRAND_NAVY} !important;
}}
[data-testid="stSidebar"] .stButton > button[kind="primary"] {{
    background: {_BRAND_NAVY} !important;
}}

/* ── Inputs, selects, radios — brand focus ring ── */
[data-baseweb="input"] input:focus,
[data-baseweb="select"]:focus-within,
[data-baseweb="textarea"]:focus {{
    border-color: {_BRAND_NAVY} !important;
    box-shadow: 0 0 0 1px {_BRAND_NAVY} !important;
}}
[data-baseweb="checkbox"] svg {{
    fill: {_BRAND_NAVY} !important;
}}
[data-baseweb="radio"] svg {{
    fill: {_BRAND_NAVY} !important;
}}

/* ── Tooltip help icons ── */
[data-testid="stTooltipIcon"],
[data-testid="stTooltipHoverTarget"] {{
    color: {_BRAND_MUTED} !important;
}}
[data-testid="stTooltipIcon"]:hover {{
    color: {_BRAND_NAVY} !important;
}}
.stTooltipContent {{
    background: {_BRAND_NAVY} !important;
    color: #FFFFFF !important;
    font-size: 10pt !important;
    line-height: 1.5 !important;
    border-radius: 6px !important;
    max-width: 320px !important;
}}

/* ── Tabs & expanders — brand accent ── */
[data-baseweb="tab"] {{
    color: {_BRAND_SLATE} !important;
}}
[data-baseweb="tab"][aria-selected="true"] {{
    color: {_BRAND_NAVY} !important;
    border-bottom-color: {_BRAND_NAVY} !important;
}}
[data-testid="stExpander"] {{
    border: 1px solid {_BRAND_BORDER} !important;
    border-radius: 8px !important;
    background: {_BRAND_CARD} !important;
}}

/* ── Scrollable tables on narrow screens ── */
[data-testid="stDataFrame"],
[data-testid="stDataEditor"],
[data-testid="stTable"] {{
    overflow-x: auto !important;
}}

/* ── Mobile-responsive layout ── */
@media (max-width: 768px) {{
    .block-container {{
        padding: 1rem 0.75rem 2rem !important;
        max-width: 100% !important;
    }}
    .prism-hero {{
        padding: 24px 18px !important;
        border-radius: 8px !important;
    }}
    .prism-hero h1 {{
        font-size: 1.5rem !important;
    }}
    .prism-about-panel {{
        padding: 18px 16px !important;
    }}
    .prism-framework-row {{
        flex-direction: column !important;
        align-items: stretch !important;
    }}
    .prism-framework-box,
    .prism-framework-box.wide {{
        flex: 1 1 100% !important;
        min-width: 100% !important;
        max-width: 100% !important;
    }}
    .prism-framework-arrow {{
        transform: rotate(90deg);
        margin: 4px auto !important;
    }}
    .prism-framework-side {{
        flex-direction: column !important;
    }}
    [data-testid="column"] {{
        flex: 1 1 100% !important;
        min-width: 0 !important;
        width: 100% !important;
    }}
    [data-testid="stHorizontalBlock"] {{
        flex-wrap: wrap !important;
        gap: 0.5rem !important;
    }}
    .stButton > button,
    .stDownloadButton > button {{
        width: 100% !important;
    }}
    [data-testid="stSidebar"] {{
        min-width: 280px !important;
    }}
    .prism-progress-pct {{
        font-size: 12pt !important;
    }}
}}

@media (max-width: 480px) {{
    .prism-dimension-pill {{
        display: block !important;
        margin: 4px 0 !important;
        text-align: center;
    }}
    .prism-card {{
        margin-bottom: 12px;
    }}
}}

/* ── Divider ── */
hr {{
    border-color: #DCE3EF !important;
}}
</style>
""", unsafe_allow_html=True)



MPL_STYLE = {
    "font.family": "sans-serif",
    "font.sans-serif": ["Inter", "Segoe UI", "Arial", "DejaVu Sans"],
    "axes.titlecolor": "#0D2B5E",
    "axes.labelcolor": "#0D2B5E",
    "xtick.color": "#0D2B5E",
    "ytick.color": "#0D2B5E",
    "axes.edgecolor": "#0D2B5E",
    "axes.spines.top": False,
    "axes.spines.right": False,
    "figure.facecolor": "white",
    "axes.facecolor": "white",
}

def apply_mpl_style():
    plt.rcParams.update(MPL_STYLE)

apply_mpl_style()

def mpl_show(fig):
    """Display matplotlib figure in Streamlit and close it."""
    apply_mpl_style()
    st.pyplot(fig, use_container_width=True)
    plt.close(fig)


CAT_SHORT = {
    "env": "E",
    "eco": "C",
    "soc": "S",
    "qua": "Q",
    "pro": "P",
}

def proc_short_labels(names):
    """Generate short labels for alternatives to avoid x-axis overlap."""
    shorts = []
    used = set()
    for name in names:
        words = name.strip().split()
        # Try first letters of each word
        abbr = "".join(w[0].upper() for w in words if w)
        if abbr in used or len(abbr) == 0:
            abbr = name[:4].strip()
        if abbr in used:
            abbr = name[:6].strip()
        used.add(abbr)
        shorts.append(abbr)
    return shorts


PROC_COLORS = ["#2563EB", "#16A34A", "#EA580C", "#9333EA", "#0891B2",
               "#CA8A04", "#DB2777", "#4F46E5", "#65A30D", "#DC2626"]

CATEGORY_ORDER = ["env", "eco", "soc", "qua", "pro"]

CATS = {
    "env": {
        "label": "Environmental", "color": "#0F6E56", "bg": "#E1F5EE",
        "indicators": ["Cumulative energy demand", "CO₂ emissions", "Water consumption"],
        "default_units": ["MJ", "kg CO₂-eq", "L"],
        "unit_options": [
            ["MJ", "GJ", "kWh", "MJ/kg"],
            ["kg CO₂-eq", "t CO₂-eq", "g CO₂-eq"],
            ["L", "m³", "mL", "kg"],
        ],
        "benefit": [False, False, False],
    },
    "eco": {
        "label": "Economic", "color": "#185FA5", "bg": "#E6F1FB",
        "indicators": ["Material cost", "Machine cost", "Labour cost", "Consumables cost"],
        "default_units": ["GBP/part", "GBP/part", "GBP/part", "GBP/part"],
        "unit_options": [
            ["GBP/part", "USD/part", "EUR/part", "GBP/kg", "USD/kg"],
            ["GBP/part", "USD/part", "EUR/part", "GBP/hr", "USD/hr"],
            ["GBP/part", "USD/part", "EUR/part", "GBP/hr", "USD/hr"],
            ["GBP/part", "USD/part", "EUR/part"],
        ],
        "benefit": [False, False, False, False],
    },
    "soc": {
        "label": "Social", "color": "#534AB7", "bg": "#EEEDFE",
        "indicators": ["Recordable injury rate", "Job satisfaction"],
        "default_units": ["per 100 workers", "GBP/year"],
        "unit_options": [
            ["per 100 workers", "per 200,000 h", "TRIR"],
            ["GBP/year", "USD/year", "EUR/year", "score (1-10)", "score (1-5)"],
        ],
        "benefit": [False, True],
    },
    "qua": {
        "label": "Quality", "color": "#854F0B", "bg": "#FAEEDA",
        "indicators": ["Tensile strength", "Yield strength", "% elongation"],
        "default_units": ["MPa", "MPa", "%"],
        "unit_options": [
            ["MPa", "GPa", "ksi", "N/mm²"],
            ["MPa", "GPa", "ksi", "N/mm²"],
            ["%", "ratio"],
        ],
        "benefit": [True, True, True],
    },
    "pro": {
        "label": "Productivity", "color": "#993C1D", "bg": "#FAECE7",
        "indicators": ["Total production time", "Material utilisation rate"],
        "default_units": ["h", "%"],
        "unit_options": [["h", "min", "days", "s"], ["%", "ratio", "g/g"]],
        "benefit": [False, True],
    },
}

CUSTOM_SENTINEL = "Custom..."


# ============================================================================
# CORE MATH
# ============================================================================

def merec_norm(vals, benefit):
    vals = np.asarray(vals, dtype=float)
    if benefit:
        mn = vals.min()
        return np.where(vals > 0, mn / vals, 0.0)
    else:
        mx = vals.max() or 1.0
        return vals / mx


def n2_norm(vals, benefit):
    vals = np.asarray(vals, dtype=float)
    if benefit:
        s = vals.sum() or 1.0
        return vals / s
    else:
        inv = np.where(vals > 0, 1.0 / vals, 0.0)
        s = inv.sum() or 1.0
        return inv / s


def merec_weights(norm_matrix):
    n_crit, n_alt = norm_matrix.shape
    safe = np.clip(norm_matrix, 1e-15, None)
    abs_ln = np.abs(np.log(safe))
    S = np.log(1 + abs_ln.sum(axis=0) / n_crit)
    E = np.zeros(n_crit)
    for j in range(n_crit):
        mask = np.ones(n_crit, dtype=bool)
        mask[j] = False
        s_prime = np.log(1 + abs_ln[mask, :].sum(axis=0) / n_crit)
        E[j] = np.sum(np.abs(s_prime - S))
    total = E.sum() or 1.0
    return E / total


def entropy_weights(mat):
    k, n = mat.shape
    E = np.zeros(k)
    for j in range(k):
        row = mat[j]
        s = row.sum() or 1.0
        p = row / s
        with np.errstate(divide="ignore", invalid="ignore"):
            term = np.where(p > 0, p * np.log(p), 0.0)
        e = -term.sum() / (np.log(n) or 1.0)
        E[j] = min(max(e, 0.0), 1.0)
    d = 1 - E
    total = d.sum() or 1.0
    return d / total


def critic_weights(mat):
    k, n = mat.shape
    rescaled = np.zeros_like(mat)
    for j in range(k):
        row = mat[j]
        mx, mn = row.max(), row.min()
        rng = (mx - mn) or 1.0
        rescaled[j] = (row - mn) / rng
    means = rescaled.mean(axis=1)
    stds = rescaled.std(axis=1, ddof=0)
    corr = np.ones((k, k))
    for a in range(k):
        for b in range(k):
            if a == b:
                continue
            da = rescaled[a] - means[a]
            db = rescaled[b] - means[b]
            num = np.sum(da * db)
            den = np.sqrt(np.sum(da ** 2) * np.sum(db ** 2)) or 1.0
            corr[a, b] = num / den
    C = np.array([stds[j] * np.sum(1 - corr[j, :]) for j in range(k)])
    total = C.sum() or 1.0
    return C / total


def rcw_consolidate(weight_sets):
    """
    Reciprocal Composite Weighting (RCW).

    Step 1 — Harmonic mean per category j:
        H_j = n / (1/w_j^Eq + 1/w_j^Ent + 1/w_j^Critic)

    Step 2 — Normalise:
        w_j^RCW = H_j / Σ_j H_j

    Reference: Novel contribution — see PRISM methodology chapter.
    """
    k = len(weight_sets[0])
    n = len(weight_sets)
    harmonics = np.zeros(k)
    for i in range(k):
        s_inv = sum(1.0 / max(ws[i], 1e-9) for ws in weight_sets)
        harmonics[i] = n / s_inv if s_inv > 0 else 0.0
    total = harmonics.sum() or 1.0
    return harmonics / total


def _harmonic_mean_weights(weight_sets):
    """Harmonic mean of a list of weight vectors. Returns unnormalised result."""
    k = len(weight_sets[0])
    h = np.zeros(k)
    n = len(weight_sets)
    for i in range(k):
        s_inv = sum(1.0 / max(ws[i], 1e-10) for ws in weight_sets)
        h[i] = n / s_inv if s_inv > 0 else 0.0
    return h


def hrcw_consolidate(w_equal, w_entropy, w_critic):
    """
    Hierarchical Reciprocal Composite Weighting (HRCW).

    Two-level harmonic consolidation grouped by epistemic basis:

      Group A — Prior-based (no data assumption):
          Equal weighting  →  w_prior = w_equal

      Group B — Posterior-based (data-informed):
          Step 1: Consolidate Entropy and CRITIC within Group B
                  w_posterior = Harmonic(w_entropy, w_critic)
          Step 2: Balance prior against posterior
                  w_HRCW = Harmonic(w_prior, w_posterior)
                  then normalise to sum = 1

    Reference: Novel contribution — see PRISM methodology chapter.
    """
    # Step 1 — within-group consolidation of posterior methods
    w_posterior_raw = _harmonic_mean_weights([w_entropy, w_critic])
    total_post = w_posterior_raw.sum()
    w_posterior = w_posterior_raw / total_post if total_post > 1e-10         else np.full(len(w_equal), 1.0 / len(w_equal))

    # Step 2 — cross-group consolidation: prior vs posterior
    w_hrcw_raw = _harmonic_mean_weights([w_equal, w_posterior])
    total_hrcw = w_hrcw_raw.sum()
    if total_hrcw < 1e-10:
        return np.full(len(w_equal), 1.0 / len(w_equal))
    return w_hrcw_raw / total_hrcw


def rank_with_ties(values, ascending, eps=1e-6):
    n = len(values)
    order = np.argsort(values if ascending else -values, kind="stable")
    ranks = np.zeros(n, dtype=int)
    cur_rank = 1
    for pos in range(n):
        if pos > 0:
            prev_val = values[order[pos - 1]]
            cur_val = values[order[pos]]
            if abs(cur_val - prev_val) > eps:
                cur_rank = pos + 1
        ranks[order[pos]] = cur_rank
    return ranks


def topsis(weighted_mat):
    ideal = weighted_mat.max(axis=1)
    anti = weighted_mat.min(axis=1)
    Dp = np.sqrt(((weighted_mat - ideal[:, None]) ** 2).sum(axis=0))
    Dm = np.sqrt(((weighted_mat - anti[:, None]) ** 2).sum(axis=0))
    C = np.where((Dp + Dm) > 0, Dm / (Dp + Dm), 0.0)
    return rank_with_ties(C, ascending=False)


def vikor(weighted_mat, weights, v=0.5):
    f_star = weighted_mat.max(axis=1)
    f_minus = weighted_mat.min(axis=1)
    denom = np.where((f_star - f_minus) != 0, f_star - f_minus, 1.0)
    S = np.sum(weights[:, None] * (f_star[:, None] - weighted_mat) / denom[:, None], axis=0)
    R = np.max(weights[:, None] * (f_star[:, None] - weighted_mat) / denom[:, None], axis=0)
    Sm, Sx = S.min(), S.max() or 1.0
    Rm, Rx = R.min(), R.max() or 1.0
    Q = v * (S - Sm) / ((Sx - Sm) or 1.0) + (1 - v) * (R - Rm) / ((Rx - Rm) or 1.0)
    return rank_with_ties(Q, ascending=True)


def electre_score(weighted_mat, weights, c_thresh=0.65, d_thresh=0.35, n_refs=5):
    """
    ELECTRE-Score (Figueira, Greco & Roy, 2022, EJOR 297:986-1005).
    Assigns a continuous score to each alternative by comparing it
    against uniformly-spaced reference profiles spanning [min, max]
    of each criterion. Produces meaningful discrimination even with
    few alternatives — unlike ELECTRE I which is binary.
    """
    n_crit, n_alt = weighted_mat.shape
    w = np.asarray(weights, dtype=float)
    total_w = w.sum() or 1.0
    w = w / total_w

    # Reference profiles: h profiles uniformly spanning [min, max] per criterion
    ref_scores = np.linspace(1.0, 0.0, n_refs)
    mat_max = weighted_mat.max(axis=1)
    mat_min = weighted_mat.min(axis=1)
    # ref_profiles shape: (n_crit, n_refs)
    ref_profiles = mat_min[:, None] + np.outer(
        mat_max - mat_min, ref_scores
    )

    # Ranges per criterion across all alternatives AND reference profiles
    all_vals = np.hstack([weighted_mat, ref_profiles])
    ranges = all_vals.max(axis=1) - all_vals.min(axis=1)
    ranges = np.where(ranges > 1e-10, ranges, 1.0)

    def _concordance(a_vec, b_vec):
        return np.sum(w[a_vec >= b_vec])

    def _discordance(a_vec, b_vec):
        return np.max(np.maximum(b_vec - a_vec, 0.0) / ranges)

    def _outranks(a_vec, b_vec):
        return (_concordance(a_vec, b_vec) >= c_thresh and
                _discordance(a_vec, b_vec) <= d_thresh)

    delta = 1.0 / (2.0 * n_refs)
    scores = np.zeros(n_alt)

    for i in range(n_alt):
        alt = weighted_mat[:, i]
        alt_outranks_ref = [_outranks(alt, ref_profiles[:, k]) for k in range(n_refs)]
        ref_outranks_alt = [_outranks(ref_profiles[:, k], alt) for k in range(n_refs)]

        # Upper: highest score among refs that alt outranks
        upper = max((ref_scores[k] for k in range(n_refs) if alt_outranks_ref[k]),
                    default=None)
        # Lower: lowest score among refs that outrank alt
        lower = min((ref_scores[k] for k in range(n_refs) if ref_outranks_alt[k]),
                    default=None)

        if upper is None and lower is None:
            scores[i] = 0.5          # incomparable — assign midpoint
        elif upper is None:
            scores[i] = lower - delta  # below all references
        elif lower is None:
            scores[i] = upper + delta  # above all references
        else:
            scores[i] = (upper + lower) / 2.0

    return rank_with_ties(scores, ascending=False)


def multimoora(weighted_mat):
    RS = weighted_mat.sum(axis=0)
    ref = weighted_mat.max(axis=1)
    RP = np.max(np.abs(ref[:, None] - weighted_mat), axis=0)
    FMF = np.prod(np.maximum(weighted_mat, 1e-9), axis=0)
    rs_rank = rank_with_ties(RS, ascending=False)
    rp_rank = rank_with_ties(RP, ascending=True)
    fmf_rank = rank_with_ties(FMF, ascending=False)
    combined = rs_rank + rp_rank + fmf_rank
    return rank_with_ties(combined.astype(float), ascending=True)


def waspas(weighted_mat, lam=0.5):
    # weighted_mat shape: (n_crit, n_alt) — already normalised x weights
    # WSM = sum of weighted normalised values (already in weighted_mat)
    WSM = weighted_mat.sum(axis=0)
    # WPM = product of (normalised_value ^ weight) per criterion
    # Since weighted_mat[j,i] = norm[j,i] * w[j], and we need norm[j,i]^w[j],
    # we use exp(w[j] * log(norm[j,i])) = exp(log(weighted_mat[j,i] / w[j] + 1e-12) * w[j])
    # Simpler: WPM = prod(weighted_mat^1) is wrong; use log space
    # Actually for WASPAS with pre-weighted matrix: WSM is correct already
    # WPM needs raw norm values — approximate via: norm = weighted_mat / weights (row-wise)
    # We pass weights separately via the lambda in MCDM_FUNCS
    Q = WSM  # fallback to WSM only if weights unavailable
    return rank_with_ties(Q, ascending=False)


def waspas_full(weighted_mat, weights, lam=0.5):
    """Correct WASPAS: recovers normalised matrix from weighted_mat and weights."""
    wm = np.asarray(weighted_mat, dtype=float)
    w = np.asarray(weights, dtype=float)
    # Recover normalised matrix: norm[j,i] = wm[j,i] / w[j]
    norm = np.where(w[:, None] > 1e-12, wm / w[:, None], 0.0)
    # WSM = sum of weighted normalised values (= wm.sum(axis=0))
    WSM = wm.sum(axis=0)
    # WPM = product of norm[j,i]^w[j]
    WPM = np.prod(np.maximum(norm, 1e-9) ** w[:, None], axis=0)
    Q = lam * WSM + (1 - lam) * WPM
    return rank_with_ties(Q, ascending=False)



MCDM_FUNCS = {
    "topsis": lambda wm, w: topsis(wm),
    "vikor": lambda wm, w: vikor(wm, w),
    "electre": lambda wm, w: electre_score(wm, w),
    "multimoora": lambda wm, w: multimoora(wm),
    "waspas": lambda wm, w: waspas_full(wm, w),
}
METHOD_LABELS = {
    "topsis": "TOPSIS", "vikor": "VIKOR", "electre": "ELECTRE-Score",
    "multimoora": "MULTIMOORA", "waspas": "WASPAS",
}
ALL_MCDM_KEYS = ["topsis", "vikor", "electre", "multimoora", "waspas"]
WEIGHT_COMBO_SETS = [
    (("equal",), "Equal"),
    (("entropy",), "Entropy"),
    (("critic",), "CRITIC"),
    (("equal", "entropy"), "RCW(Eq+En)"),
    (("equal", "critic"), "RCW(Eq+Cr)"),
    (("entropy", "critic"), "RCW(En+Cr)"),
]


def calc_psi(method_ranks, methods, p):
    n_alt = len(next(iter(method_ranks.values())))
    psi = np.zeros(n_alt)
    for i in range(n_alt):
        ranks = np.array([method_ranks[m][i] for m in methods], dtype=float)
        r_bar = ranks.mean()
        cv = ranks.std(ddof=0) / r_bar if (len(ranks) > 1 and r_bar > 0) else 0.0
        M = 1.0 / (r_bar or 1.0)
        A = 1.0 / (1 + cv)
        psi[i] = (M ** p) * (A ** (1 - p))
    return psi


def get_category_weights(mat, weight_methods):
    """
    Compute category weights using RCW (harmonic mean consolidation)
    across selected objective weighting methods.
    """
    k = mat.shape[0]
    sets = []
    if "equal"   in weight_methods: sets.append(np.full(k, 1.0 / k))
    if "entropy" in weight_methods: sets.append(entropy_weights(mat))
    if "critic"  in weight_methods: sets.append(critic_weights(mat))
    if not sets:
        return np.full(k, 1.0 / k)
    return rcw_consolidate(sets) if len(sets) > 1 else sets[0]


def run_mcdm_suite(weighted_mat, weights, methods):
    ranks = {}
    for m in methods:
        ranks[m] = MCDM_FUNCS[m](weighted_mat, weights)
    return ranks




def get_combinations(cats_list):
    order_index = {c: i for i, c in enumerate(CATEGORY_ORDER)}
    cats_sorted = sorted(cats_list, key=lambda c: order_index[c])
    combos = []
    for r in range(1, len(cats_sorted) + 1):
        combos.extend(itertools.combinations(cats_sorted, r))
    combos.sort(key=lambda c: (len(c), [order_index[x] for x in c]))
    return [list(c) for c in combos]


def _compute_combo_rank_grid(l3_cats, cat_scores, methods, combo_p=0.5):
    """Recompute category-combination rank grid (same logic as Step 12)."""
    combos = get_combinations(l3_cats)
    n_proc = st.session_state.n_proc
    rank_grid = np.zeros((n_proc, len(combos)), dtype=int)
    cat_initial = {c: CAT_SHORT.get(c, CATS[c]["label"][:1]) for c in l3_cats}
    combo_labels = ["+".join(cat_initial[c] for c in combo) for combo in combos]
    weight_methods = st.session_state.sel_weight_methods
    for ci, combo in enumerate(combos):
        if len(combo) == 1:
            rank_grid[:, ci] = rank_with_ties(cat_scores[combo[0]], ascending=False)
        else:
            sub_mat = np.array([cat_scores[c] for c in combo])
            sub_w = get_category_weights(sub_mat, weight_methods)
            sub_weighted = sub_mat * sub_w[:, None]
            sub_ranks = run_mcdm_suite(sub_weighted, sub_w, methods)
            psi_combo = calc_psi(sub_ranks, methods, combo_p)
            rank_grid[:, ci] = rank_with_ties(psi_combo, ascending=False)
    return rank_grid, combo_labels


def _resolve_overall_winner(names, method_ranks, methods):
    """Determine overall winner from PSI, PSI-combo, or single MCDM method."""
    ss = st.session_state
    combo_ranks = ss.get("last_psi_combo_ranks")
    if combo_ranks is not None:
        combo_ranks = np.array(combo_ranks, dtype=int)
        min_r = combo_ranks.min()
        leaders = np.where(combo_ranks == min_r)[0]
        if len(leaders) == 1:
            wi = int(leaders[0])
            combo_scores = ss.get("last_psi_combo_scores")
            psi_score = float(combo_scores[wi]) if combo_scores else None
            return {
                "winner_idx": wi,
                "winner_name": names[wi],
                "basis_label": "PSI-combo tie-breaker (p=0.50)",
                "is_tied": False,
                "tied_names": [],
                "psi_score": psi_score,
                "psi_ranks": combo_ranks,
                "psi_scores": np.array(combo_scores, dtype=float) if combo_scores else None,
            }

    if len(methods) > 1:
        psi_scores = ss.get("last_psi_scores")
        psi = np.array(psi_scores, dtype=float) if psi_scores is not None else calc_psi(method_ranks, methods, 0.5)
        psi_ranks = rank_with_ties(psi, ascending=False)
        min_r = psi_ranks.min()
        leaders = np.where(psi_ranks == min_r)[0]
        wi = int(leaders[0])
        return {
            "winner_idx": wi,
            "winner_name": names[wi],
            "basis_label": "PSI compromise rank (p=0.50)",
            "is_tied": len(leaders) > 1,
            "tied_names": [names[i] for i in leaders],
            "psi_score": float(psi[wi]),
            "psi_ranks": psi_ranks,
            "psi_scores": psi,
        }

    m = methods[0]
    ranks = method_ranks[m]
    min_r = ranks.min()
    leaders = np.where(ranks == min_r)[0]
    wi = int(leaders[0])
    return {
        "winner_idx": wi,
        "winner_name": names[wi],
        "basis_label": f"{METHOD_LABELS[m]} rank",
        "is_tied": len(leaders) > 1,
        "tied_names": [names[i] for i in leaders],
        "psi_score": None,
        "psi_ranks": None,
        "psi_scores": None,
    }


def _build_decision_support_tables(winner_info):
    """Build summary, strengths, weaknesses, and category profile tables."""
    names = st.session_state.proc_names
    n_proc = len(names)
    wi = winner_info["winner_idx"]
    method_ranks = st.session_state.last_method_ranks
    methods = list(st.session_state.sel_mcdm_methods) or list(method_ranks.keys())
    l3_cats = ordered_l3_cats()
    cat_scores = st.session_state.cat_scores

    summary_rows = [
        {"Field": "Recommended alternative", "Value": winner_info["winner_name"]},
        {"Field": "Decision basis", "Value": winner_info["basis_label"]},
    ]
    if winner_info.get("is_tied"):
        summary_rows.append({
            "Field": "Tie note",
            "Value": f"Tied with: {', '.join(winner_info.get('tied_names', []))}",
        })
    if winner_info.get("psi_score") is not None:
        summary_rows.append({
            "Field": "PSI score",
            "Value": round(winner_info["psi_score"], 4),
        })

    if winner_info.get("psi_scores") is not None:
        psi = winner_info["psi_scores"]
        for idx in np.argsort(-psi):
            if int(idx) != wi:
                runner_idx = int(idx)
                margin = float(psi[wi] - psi[runner_idx])
                summary_rows.append({
                    "Field": "Runner-up",
                    "Value": f"{names[runner_idx]} (PSI margin {margin:.4f})",
                })
                break
    elif len(methods) == 1:
        ranks = method_ranks[methods[0]]
        for pi in np.argsort(ranks):
            if int(pi) != wi:
                summary_rows.append({"Field": "Runner-up", "Value": names[int(pi)]})
                break

    strengths, weaknesses, profile_rows = [], [], []

    if len(methods) > 1 and winner_info.get("psi_ranks") is not None:
        pr = int(winner_info["psi_ranks"][wi])
        if pr == 1 and not winner_info.get("is_tied"):
            strengths.append({
                "Evidence": "PSI compromise rank",
                "Source": "Step 12",
                "Result": "Rank 1",
                "Interpretation": "Leading overall ranking",
            })
        elif winner_info.get("is_tied"):
            strengths.append({
                "Evidence": "PSI compromise rank",
                "Source": "Step 12",
                "Result": f"Tied rank 1 ({len(winner_info.get('tied_names', []))} alternatives)",
                "Interpretation": "Shared top position — review tie-breaker evidence",
            })
            weaknesses.append({
                "Area": "PSI compromise rank",
                "Source": "Step 12",
                "Result": "Tied at rank 1",
                "Impact": "No unique overall winner on headline PSI",
            })
        else:
            weaknesses.append({
                "Area": "PSI compromise rank",
                "Source": "Step 12",
                "Result": f"Rank {pr}",
                "Impact": "Not the top PSI-ranked alternative",
            })

    for ckey in l3_cats:
        scores = cat_scores[ckey]
        cat_ranks = rank_with_ties(scores, ascending=False)
        w_score = float(scores[wi])
        w_rank = int(cat_ranks[wi])
        best_idx = int(np.argmax(scores))
        best_score = float(scores[best_idx])
        label = CATS[ckey]["label"]
        profile_rows.append({
            "Category": label,
            "Winner score": round(w_score, 4),
            "Winner rank": w_rank,
            "Best alternative": names[best_idx],
            "Best score": round(best_score, 4),
        })
        if w_rank == 1 and best_idx == wi:
            strengths.append({
                "Evidence": f"{label} score",
                "Source": "Step 8",
                "Result": f"{w_score:.4f} (Rank 1 of {n_proc})",
                "Interpretation": f"Strongest on {label.lower()}",
            })
        elif w_rank > 1:
            weaknesses.append({
                "Area": f"{label} category",
                "Source": "Step 8",
                "Result": f"Rank {w_rank} ({w_score:.4f} vs best {best_score:.4f})",
                "Impact": f"Below best on {label.lower()}",
            })

    for m in methods:
        r = int(method_ranks[m][wi])
        if r == 1:
            strengths.append({
                "Evidence": f"{METHOD_LABELS[m]} rank",
                "Source": "Step 12",
                "Result": "Rank 1",
                "Interpretation": "Top-ranked on this MCDM method",
            })
        else:
            weaknesses.append({
                "Area": METHOD_LABELS[m],
                "Source": "Step 12",
                "Result": f"Rank {r}",
                "Impact": "Not top on all MCDM methods",
            })

    if l3_cats and methods:
        rank_grid, combo_labels = _compute_combo_rank_grid(l3_cats, cat_scores, methods)
        n_combos = len(combo_labels)
        rank1_count = int(np.sum(rank_grid[wi, :] == 1))
        if rank1_count > 0:
            strengths.append({
                "Evidence": "Category combinations",
                "Source": "Step 12",
                "Result": f"Rank 1 in {rank1_count} of {n_combos} combinations",
                "Interpretation": "Robust across stakeholder category views",
            })
        bad_combos = [
            (int(rank_grid[wi, ci]), combo_labels[ci])
            for ci in range(n_combos)
            if int(rank_grid[wi, ci]) > 1
        ]
        bad_combos.sort(key=lambda x: -x[0])
        for r, lbl in bad_combos[:5]:
            weaknesses.append({
                "Area": f"{lbl} combo",
                "Source": "Step 12",
                "Result": f"Rank {r}",
                "Impact": "Sensitive to this category subset",
            })
        if len(bad_combos) > 5:
            weaknesses.append({
                "Area": "Other combinations",
                "Source": "Step 12",
                "Result": f"{len(bad_combos) - 5} more combos with rank > 1",
                "Impact": "See Step 12 combination chart for full profile",
            })

    mc_psi = st.session_state.get("mc_psi_rank_counts")
    if mc_psi is not None:
        n_iter = st.session_state.get("mc_n_iter", 1) or 1
        pct_rank1 = 100.0 * mc_psi[wi, 0] / n_iter
        if pct_rank1 >= 50:
            strengths.append({
                "Evidence": "Monte Carlo PSI stability",
                "Source": "Step 13",
                "Result": f"Rank 1 in {pct_rank1:.1f}% of draws",
                "Interpretation": "Stable under weight uncertainty",
            })
        else:
            weaknesses.append({
                "Area": "Monte Carlo PSI stability",
                "Source": "Step 13",
                "Result": f"Rank 1 in only {pct_rank1:.1f}% of draws",
                "Impact": "Ranking sensitive to weight uncertainty",
            })

    val_choice = st.session_state.get("validation_choice", "None - skip validation")
    if val_choice and not val_choice.startswith("None"):
        val_label = val_choice.split(". ", 1)[-1] if ". " in val_choice else val_choice
        strengths.append({
            "Evidence": "Validation reviewed",
            "Source": "Step 13",
            "Result": val_label,
            "Interpretation": "Robustness check reviewed — see Step 13 for details",
        })

    aux_choice = st.session_state.get("auxiliary_radio", "None - skip")
    if aux_choice and not aux_choice.startswith("None"):
        aux_label = aux_choice.split(". ", 1)[-1] if ". " in aux_choice else aux_choice
        strengths.append({
            "Evidence": "Analytics reviewed",
            "Source": "Step 14",
            "Result": aux_label,
            "Interpretation": "Supplementary analysis reviewed — see Step 14 for details",
        })

    if not strengths:
        strengths.append({
            "Evidence": "—",
            "Source": "—",
            "Result": "—",
            "Interpretation": "No clear strength signals identified",
        })
    if not weaknesses:
        weaknesses.append({
            "Area": "—",
            "Source": "—",
            "Result": "—",
            "Impact": "No material weaknesses detected across categories, methods, or combinations",
        })

    return summary_rows, strengths, weaknesses, profile_rows


def compute_dirichlet_k(mat):
    k_cat = mat.shape[0]
    w_eq = np.full(k_cat, 1.0 / k_cat)
    w_en = entropy_weights(mat)
    w_cr = critic_weights(mat)
    W = np.array([w_eq, w_en, w_cr])
    n_methods = 3
    var_per_cat = W.var(axis=0, ddof=0)
    mean_var = var_per_cat.mean()
    max_var = ((n_methods - 1) * (1 / n_methods) ** 2 + (1 - 1 / n_methods) ** 2) / n_methods
    dispersion_ratio = min(mean_var / max_var, 1.0) if max_var > 0 else 0.0
    agreement = 1 - dispersion_ratio
    return agreement * 100, w_eq, w_en, w_cr


# ============================================================================
# SESSION STATE
# ============================================================================

def init_state():
    defaults = {
        "step": 0,
        "n_proc": 3,
        "proc_names": [],
        "sel_cats": set(),
        "sel_units": {},
        "indicator_values": {},
        "use_custom_indicators": None,
        "custom_indicator_counts": {},
        "custom_indicators": {},
        "l3_cats": set(),
        "sel_weight_methods": set(),
        "sel_mcdm_methods": set(),
        "computed": False,
        "corr_acknowledged": False,
        "outliers_acknowledged": False,
        "validation_choice": "None - skip validation",
        "disabled_indicators": set(),  # set of (ckey, j) tuples
        "current_assessment_id": None,
        "assessment_name": "",
        "share_code": "",
        "owner_label": "",
        "collaborator_mode": False,
        "assigned_categories": set(),
        "last_remote_updated_at": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def reset_all():
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    init_state()


# ============================================================================
# ASSESSMENT PERSISTENCE (serialize / Supabase / JSON fallback)
# ============================================================================

ASSESSMENT_SCHEMA_VERSION = 1

PERSIST_KEYS = [
    "step", "n_proc", "proc_names", "sel_cats", "sel_units",
    "indicator_values", "use_custom_indicators", "custom_indicator_counts",
    "custom_indicators", "l3_cats", "sel_weight_methods", "sel_mcdm_methods",
    "disabled_indicators", "corr_acknowledged", "validation_choice",
    "cat_scores", "nm_data", "n2_data", "merec_w", "final_cat_weights",
    "last_method_ranks", "last_psi_scores", "last_psi_combo_scores",
    "last_psi_combo_ranks", "computed",
]

SET_STATE_KEYS = {
    "sel_cats", "l3_cats", "sel_weight_methods", "sel_mcdm_methods",
    "disabled_indicators", "assigned_categories",
}
TUPLE_PAIR_KEYS = {"disabled_indicators", "custom_indicators"}
TUPLE_TRIPLE_KEYS = {"indicator_values"}

WIDGET_KEY_PREFIXES = (
    "editor_", "editor_seed_", "pname_", "catchk_", "l3chk_", "wmchk_",
    "mmchk_", "unitsel_", "unitcustom_", "ind_enabled_", "customcnt_",
    "custname_", "custunit_", "custben_", "use_custom_radio", "n_proc_slider",
)


def _encode_key(key):
    if isinstance(key, tuple):
        return "|".join(str(p) for p in key)
    return key


def _decode_key(key_str, n_parts):
    parts = key_str.split("|")
    if n_parts == 2:
        return parts[0], int(parts[1])
    if n_parts == 3:
        return parts[0], int(parts[1]), int(parts[2])
    return key_str


def _to_json_safe(value, state_key=None):
    if isinstance(value, set):
        items = []
        for item in value:
            if isinstance(item, tuple):
                items.append(_encode_key(item))
            else:
                items.append(item)
        return {"__type__": "set", "items": items}
    if isinstance(value, dict):
        out = {}
        for k, v in value.items():
            ek = _encode_key(k) if isinstance(k, tuple) else k
            out[ek] = _to_json_safe(v, state_key)
        return out
    if isinstance(value, np.ndarray):
        return {"__type__": "ndarray", "data": value.tolist()}
    if isinstance(value, (np.integer, np.floating)):
        return float(value) if isinstance(value, np.floating) else int(value)
    if isinstance(value, (list, tuple)):
        return [_to_json_safe(v, state_key) for v in value]
    return value


def _from_json_safe(value, state_key=None):
    if isinstance(value, dict):
        if value.get("__type__") == "set":
            items = value.get("items", [])
            decoded = []
            for item in items:
                if state_key in TUPLE_PAIR_KEYS and isinstance(item, str) and "|" in item:
                    decoded.append(_decode_key(item, 2))
                elif state_key in TUPLE_TRIPLE_KEYS and isinstance(item, str) and item.count("|") == 2:
                    decoded.append(_decode_key(item, 3))
                elif state_key == "custom_indicators" and isinstance(item, str) and "|" in item:
                    decoded.append(_decode_key(item, 2))
                else:
                    decoded.append(item)
            return set(decoded)
        if value.get("__type__") == "ndarray":
            return np.array(value.get("data", []), dtype=float)
        out = {}
        for k, v in value.items():
            if state_key in TUPLE_PAIR_KEYS and isinstance(k, str) and "|" in k:
                dk = _decode_key(k, 2)
            elif state_key in TUPLE_TRIPLE_KEYS and isinstance(k, str) and k.count("|") == 2:
                dk = _decode_key(k, 3)
            elif state_key == "custom_indicators" and isinstance(k, str) and "|" in k:
                dk = _decode_key(k, 2)
            else:
                dk = k
            out[dk] = _from_json_safe(v, state_key)
        return out
    if isinstance(value, list):
        return [_from_json_safe(v, state_key) for v in value]
    return value


def _clear_widget_state():
    for key in list(st.session_state.keys()):
        if any(key.startswith(p) for p in WIDGET_KEY_PREFIXES):
            del st.session_state[key]


def _rebuild_editor_seeds():
    names = st.session_state.get("proc_names") or []
    for ckey in ordered_sel_cats():
        ind_names, ind_units, _ = get_full_indicators(ckey)
        rows = [f"{ind_names[j]} ({ind_units[j]})" for j in range(len(ind_names))]
        seed = [
            [st.session_state.indicator_values.get((ckey, j, pi), 0.0)
             for pi in range(len(names))]
            for j in range(len(ind_names))
        ]
        st.session_state[f"editor_seed_{ckey}"] = pd.DataFrame(
            seed, index=rows, columns=names,
        )


def serialize_assessment(name=None):
    state = {}
    for k in PERSIST_KEYS:
        if k in st.session_state:
            state[k] = _to_json_safe(st.session_state[k], k)
    return {
        "schema_version": ASSESSMENT_SCHEMA_VERSION,
        "assessment_name": name or st.session_state.get("assessment_name") or "Untitled",
        "saved_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "state": state,
    }


def deserialize_assessment(payload):
    _clear_widget_state()
    inner = payload.get("state", payload)
    for k, v in inner.items():
        if k in PERSIST_KEYS or k in (
            "current_assessment_id", "assessment_name", "share_code",
            "owner_label", "collaborator_mode", "assigned_categories",
        ):
            st.session_state[k] = _from_json_safe(v, k)
    if payload.get("assessment_name"):
        st.session_state.assessment_name = payload["assessment_name"]
    _rebuild_editor_seeds()


def assessment_to_json_bytes(name=None):
    return json.dumps(serialize_assessment(name=name), indent=2).encode("utf-8")


def load_assessment_from_json_bytes(raw):
    payload = json.loads(raw.decode("utf-8"))
    deserialize_assessment(payload)
    st.session_state.current_assessment_id = None
    st.session_state.share_code = ""
    st.session_state.collaborator_mode = False
    st.session_state.assigned_categories = set()


def _supabase_config():
    url = key = None
    try:
        url = st.secrets.get("SUPABASE_URL")
        key = st.secrets.get("SUPABASE_KEY")
    except Exception:
        pass
    url = url or os.environ.get("SUPABASE_URL")
    key = key or os.environ.get("SUPABASE_KEY")
    return url, key


def supabase_available():
    url, key = _supabase_config()
    return SUPABASE_PKG_OK and bool(url) and bool(key)


def _get_supabase_client():
    if not supabase_available():
        return None
    url, key = _supabase_config()
    try:
        return create_client(url, key)
    except Exception:
        return None


def _generate_share_code():
    return pysecrets.token_hex(4).upper()


def list_cloud_assessments():
    client = _get_supabase_client()
    if not client:
        return []
    try:
        resp = (
            client.table("assessments")
            .select("id,name,share_code,updated_at,status,owner_label")
            .order("updated_at", desc=True)
            .execute()
        )
        return resp.data or []
    except Exception:
        return []


def fetch_assessment_by_id(assessment_id):
    client = _get_supabase_client()
    if not client:
        return None
    try:
        resp = (
            client.table("assessments")
            .select("*")
            .eq("id", assessment_id)
            .limit(1)
            .execute()
        )
        rows = resp.data or []
        return rows[0] if rows else None
    except Exception:
        return None


def fetch_assessment_by_share_code(share_code):
    client = _get_supabase_client()
    if not client:
        return None
    try:
        resp = (
            client.table("assessments")
            .select("*")
            .eq("share_code", share_code.strip().upper())
            .limit(1)
            .execute()
        )
        rows = resp.data or []
        return rows[0] if rows else None
    except Exception:
        return None


def save_assessment_to_cloud(name, owner_label="", assessment_id=None):
    client = _get_supabase_client()
    if not client:
        return None, "Supabase not configured"
    payload = serialize_assessment(name=name)
    now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    share_code = st.session_state.get("share_code") or _generate_share_code()
    owner_label = owner_label or st.session_state.get("owner_label") or "Lead"
    status = "complete" if st.session_state.get("last_method_ranks") else "in_progress"
    row = {
        "name": name,
        "owner_label": owner_label,
        "state_json": payload,
        "status": status,
        "updated_at": now,
        "share_code": share_code,
    }
    try:
        if assessment_id:
            existing = fetch_assessment_by_id(assessment_id)
            if existing and st.session_state.get("last_remote_updated_at"):
                if existing.get("updated_at") != st.session_state.last_remote_updated_at:
                    return None, "Another user saved since you loaded this assessment. Reload and merge."
            resp = client.table("assessments").update(row).eq("id", assessment_id).execute()
            saved = (resp.data or [{}])[0]
        else:
            row["id"] = str(uuid.uuid4())
            row["created_at"] = now
            resp = client.table("assessments").insert(row).execute()
            saved = (resp.data or [{}])[0]
        st.session_state.current_assessment_id = saved.get("id")
        st.session_state.assessment_name = name
        st.session_state.share_code = saved.get("share_code", share_code)
        st.session_state.owner_label = owner_label
        st.session_state.last_remote_updated_at = saved.get("updated_at")
        return saved, None
    except Exception as exc:
        return None, str(exc)


def load_assessment_from_cloud(assessment_id=None, share_code=None):
    row = None
    if assessment_id:
        row = fetch_assessment_by_id(assessment_id)
    elif share_code:
        row = fetch_assessment_by_share_code(share_code)
    if not row:
        return False, "Assessment not found"
    state_json = row.get("state_json") or {}
    deserialize_assessment(state_json)
    st.session_state.current_assessment_id = row.get("id")
    st.session_state.assessment_name = row.get("name", "")
    st.session_state.share_code = row.get("share_code", "")
    st.session_state.owner_label = row.get("owner_label", "")
    st.session_state.last_remote_updated_at = row.get("updated_at")
    return True, None


def save_collaborator_category_merge(assessment_id, assigned_categories):
    client = _get_supabase_client()
    if not client:
        return None, "Supabase not configured"
    row = fetch_assessment_by_id(assessment_id)
    if not row:
        return None, "Assessment not found"
    remote_payload = row.get("state_json") or {}
    remote_state = remote_payload.get("state", {})
    local_iv = st.session_state.indicator_values
    remote_iv_raw = remote_state.get("indicator_values", {})
    remote_iv = _from_json_safe(remote_iv_raw, "indicator_values")
    if not isinstance(remote_iv, dict):
        remote_iv = {}
    for ckey in assigned_categories:
        for key, val in local_iv.items():
            if isinstance(key, tuple) and key[0] == ckey:
                remote_iv[key] = val
    remote_state["indicator_values"] = _to_json_safe(remote_iv, "indicator_values")
    remote_payload["state"] = remote_state
    remote_payload["saved_at"] = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    try:
        resp = (
            client.table("assessments")
            .update({"state_json": remote_payload, "updated_at": now, "status": "in_progress"})
            .eq("id", assessment_id)
            .execute()
        )
        saved = (resp.data or [{}])[0]
        st.session_state.last_remote_updated_at = saved.get("updated_at")
        return saved, None
    except Exception as exc:
        return None, str(exc)


def list_category_assignments(assessment_id):
    client = _get_supabase_client()
    if not client:
        return []
    try:
        resp = (
            client.table("category_assignments")
            .select("*")
            .eq("assessment_id", assessment_id)
            .execute()
        )
        return resp.data or []
    except Exception:
        return []


def upsert_category_assignment(assessment_id, category_key, assignee_name, status="pending"):
    client = _get_supabase_client()
    if not client:
        return False
    now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    row = {
        "assessment_id": assessment_id,
        "category_key": category_key,
        "assignee_name": assignee_name,
        "status": status,
        "updated_at": now,
    }
    try:
        existing = [
            a for a in list_category_assignments(assessment_id)
            if a.get("category_key") == category_key
        ]
        if existing:
            client.table("category_assignments").update(row).eq("id", existing[0]["id"]).execute()
        else:
            row["id"] = str(uuid.uuid4())
            client.table("category_assignments").insert(row).execute()
        return True
    except Exception:
        return False


def ordered_sel_cats_for_editing():
    if st.session_state.get("collaborator_mode") and st.session_state.get("assigned_categories"):
        assigned = st.session_state.assigned_categories
        return [c for c in ordered_sel_cats() if c in assigned]
    return ordered_sel_cats()


def _ranks_from_snapshot(state_dict):
    names = state_dict.get("proc_names") or []
    method_ranks_raw = state_dict.get("last_method_ranks") or {}
    method_ranks = _from_json_safe(method_ranks_raw, "last_method_ranks")
    methods_raw = state_dict.get("sel_mcdm_methods") or []
    methods = _from_json_safe(methods_raw, "sel_mcdm_methods")
    if isinstance(methods, set):
        methods = list(methods)
    if not methods:
        methods = list(method_ranks.keys()) if method_ranks else list(ALL_MCDM_KEYS)
    methods = [m for m in methods if m in method_ranks] or list(method_ranks.keys())
    psi_raw = state_dict.get("last_psi_scores")
    psi_scores = _from_json_safe(psi_raw) if psi_raw else None
    if psi_scores is not None and len(methods) > 1:
        psi_arr = np.array(psi_scores, dtype=float)
        psi_ranks = rank_with_ties(psi_arr, ascending=False)
    elif method_ranks and methods:
        m0 = methods[0]
        ranks = np.array(method_ranks.get(m0, []), dtype=float)
        psi_ranks = rank_with_ties(ranks, ascending=True) if len(ranks) else np.array([])
        psi_arr = ranks
    else:
        psi_ranks = np.array([])
        psi_arr = np.array([])
    return names, methods, method_ranks, psi_arr, psi_ranks


def render_assessments_sidebar():
    st.markdown(
        f"<p style='font-size:10pt;font-weight:600;color:{_BRAND_NAVY};"
        f"text-transform:uppercase;letter-spacing:0.06em;"
        f"font-family:{_FONT_CSS};margin:0 0 8px 0;'>Assessments</p>",
        unsafe_allow_html=True,
    )
    if supabase_available():
        st.caption("Cloud save connected")
    else:
        st.caption("Cloud save offline — use JSON export/import")

    cur_name = st.session_state.get("assessment_name") or "Untitled"
    if st.session_state.get("current_assessment_id"):
        st.markdown(f"**Current:** {cur_name}")
        if st.session_state.get("share_code"):
            st.code(st.session_state.share_code, language=None)

    if st.button("New assessment", use_container_width=True, key="assess_new",
                 help="Start a fresh assessment (clears current session)."):
        reset_all()
        st.rerun()

    owner = st.text_input("Your name (lead)", value=st.session_state.get("owner_label") or "",
                          key="assess_owner_label", placeholder="e.g. Sam")
    save_name = st.text_input("Assessment name", value=cur_name, key="assess_save_name")

    sc1, sc2 = st.columns(2)
    with sc1:
        if st.button("Save", use_container_width=True, key="assess_save",
                     help="Save to cloud (or update current assessment)."):
            st.session_state.owner_label = owner.strip() or "Lead"
            if supabase_available():
                aid = st.session_state.get("current_assessment_id")
                saved, err = save_assessment_to_cloud(
                    save_name.strip() or "Untitled",
                    owner_label=st.session_state.owner_label,
                    assessment_id=aid,
                )
                if err:
                    st.error(err)
                else:
                    st.success("Saved to cloud.")
            else:
                st.session_state.assessment_name = save_name.strip() or "Untitled"
                st.info("Use Export JSON below (cloud not configured).")
    with sc2:
        st.download_button(
            "Export JSON",
            data=assessment_to_json_bytes(save_name.strip() or cur_name),
            file_name=f"PRISM_{save_name.strip() or 'assessment'}.json",
            mime="application/json",
            use_container_width=True,
            key="assess_export_json",
            help="Download assessment snapshot for offline backup.",
        )

    uploaded = st.file_uploader("Import JSON", type=["json"], key="assess_import_json",
                                help="Load a previously exported assessment file.")
    if uploaded is not None:
        if st.button("Load imported file", key="assess_import_btn", use_container_width=True):
            try:
                load_assessment_from_json_bytes(uploaded.getvalue())
                st.success("Assessment imported.")
                st.rerun()
            except Exception as exc:
                st.error(f"Import failed: {exc}")

    if supabase_available():
        cloud_list = list_cloud_assessments()
        if cloud_list:
            options = {f"{r['name']} ({r['updated_at'][:10]})": r["id"] for r in cloud_list}
            pick = st.selectbox("Load saved assessment", ["— select —"] + list(options.keys()),
                                key="assess_load_pick")
            if pick != "— select —" and st.button("Load", key="assess_load_btn", use_container_width=True):
                ok, err = load_assessment_from_cloud(assessment_id=options[pick])
                if ok:
                    st.session_state.collaborator_mode = False
                    st.session_state.assigned_categories = set()
                    st.success(f"Loaded {pick.split(' (')[0]}.")
                    st.rerun()
                else:
                    st.error(err or "Load failed")

        with st.expander("Join shared assessment"):
            join_code = st.text_input("Share code", key="assess_join_code", placeholder="8-char code")
            join_name = st.text_input("Your name", key="assess_join_name", placeholder="Your name")
            join_cats = st.multiselect(
                "Your category(ies)",
                options=CATEGORY_ORDER,
                format_func=lambda k: CATS[k]["label"],
                key="assess_join_cats",
                help="Categories you will enter data for.",
            )
            if st.button("Join & edit Step 5", key="assess_join_btn", use_container_width=True):
                if not join_code.strip():
                    st.error("Enter a share code.")
                elif not join_name.strip():
                    st.error("Enter your name.")
                elif not join_cats:
                    st.error("Select at least one category.")
                else:
                    ok, err = load_assessment_from_cloud(share_code=join_code.strip())
                    if ok:
                        st.session_state.collaborator_mode = True
                        st.session_state.assigned_categories = set(join_cats)
                        st.session_state.owner_label = join_name.strip()
                        for ckey in join_cats:
                            upsert_category_assignment(
                                st.session_state.current_assessment_id,
                                ckey, join_name.strip(), "in_progress",
                            )
                        st.session_state.step = 5
                        st.success("Joined — edit your assigned categories in Step 5.")
                        st.rerun()
                    else:
                        st.error(err or "Share code not found.")

        if st.session_state.get("current_assessment_id") and not st.session_state.get("collaborator_mode"):
            with st.expander("Category assignments"):
                for ckey in ordered_sel_cats():
                    assignee = st.text_input(
                        CATS[ckey]["label"],
                        key=f"assign_{ckey}",
                        placeholder="Assignee name",
                    )
                    if st.button(f"Assign {CATS[ckey]['label'][:3]}", key=f"assign_btn_{ckey}"):
                        if assignee.strip():
                            upsert_category_assignment(
                                st.session_state.current_assessment_id,
                                ckey, assignee.strip(), "pending",
                            )
                            st.success(f"Assigned {CATS[ckey]['label']}.")

    if st.button("Compare assessments", use_container_width=True, key="sidebar_compare",
                 help="Compare ranks and scores between two saved assessments."):
        st.session_state.step = -3
        st.rerun()

    st.divider()


def compare_assessments_page():
    st.header("Compare assessments")

    if not supabase_available():
        st.info(
            "Cloud compare requires Supabase. Import two JSON exports manually below, "
            "or configure Supabase secrets for saved assessments."
        )

    cloud_list = list_cloud_assessments() if supabase_available() else []
    options = {}
    for r in cloud_list:
        options[f"{r['name']} ({r['updated_at'][:10]})"] = r["id"]

    use_current = st.checkbox("Include current session as Assessment A", value=True, key="cmp_use_current")

    a_pick = b_pick = None
    if options:
        a_labels = (["Current session"] if use_current else []) + list(options.keys())
        b_labels = list(options.keys())
        a_pick = st.selectbox("Assessment A", a_labels, key="cmp_a") if a_labels else None
        b_pick = st.selectbox("Assessment B", b_labels, key="cmp_b") if b_labels else None

    ja = st.file_uploader("Or upload JSON for A", type=["json"], key="cmp_json_a")
    jb = st.file_uploader("Or upload JSON for B", type=["json"], key="cmp_json_b")

    if st.button("Run comparison", type="primary", key="cmp_run"):
        snap_a = snap_b = None

        if ja:
            snap_a = json.loads(ja.getvalue()).get("state", {})
        elif use_current and (not a_pick or a_pick == "Current session"):
            snap_a = serialize_assessment().get("state", {})
        elif a_pick and a_pick in options:
            row = fetch_assessment_by_id(options[a_pick])
            if row:
                snap_a = (row.get("state_json") or {}).get("state", {})

        if jb:
            snap_b = json.loads(jb.getvalue()).get("state", {})
        elif b_pick and b_pick in options:
            row = fetch_assessment_by_id(options[b_pick])
            if row:
                snap_b = (row.get("state_json") or {}).get("state", {})

        if not snap_a or not snap_b:
            st.error("Select or upload two assessments to compare.")
            return

        names_a, _, _, psi_a, rank_a = _ranks_from_snapshot(snap_a)
        names_b, _, _, psi_b, rank_b = _ranks_from_snapshot(snap_b)
        all_names = list(dict.fromkeys(list(names_a) + list(names_b)))

        rows = []
        for name in all_names:
            ia = names_a.index(name) if name in names_a else None
            ib = names_b.index(name) if name in names_b else None
            ra = int(rank_a[ia]) if ia is not None and len(rank_a) > ia else None
            rb = int(rank_b[ib]) if ib is not None and len(rank_b) > ib else None
            pa = float(psi_a[ia]) if ia is not None and len(psi_a) > ia else None
            pb = float(psi_b[ib]) if ib is not None and len(psi_b) > ib else None
            rows.append({
                "Alternative": name,
                "Rank A": ra,
                "Rank B": rb,
                "Rank Δ": (rb - ra) if ra is not None and rb is not None else None,
                "PSI A": round(pa, 4) if pa is not None else None,
                "PSI B": round(pb, 4) if pb is not None else None,
                "PSI Δ": round(pb - pa, 4) if pa is not None and pb is not None else None,
            })
        st.subheader("Rank & PSI comparison")
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        cs_a = _from_json_safe(snap_a.get("cat_scores", {}), "cat_scores")
        cs_b = _from_json_safe(snap_b.get("cat_scores", {}), "cat_scores")
        if isinstance(cs_a, dict) and isinstance(cs_b, dict) and cs_a and cs_b:
            cat_rows = []
            shared_cats = [c for c in CATEGORY_ORDER if c in cs_a and c in cs_b]
            for ckey in shared_cats:
                for pi, name in enumerate(all_names):
                    ia = names_a.index(name) if name in names_a else None
                    ib = names_b.index(name) if name in names_b else None
                    if ia is None or ib is None:
                        continue
                    va = float(np.array(cs_a[ckey])[ia]) if ia is not None else None
                    vb = float(np.array(cs_b[ckey])[ib]) if ib is not None else None
                    if va is not None and vb is not None:
                        cat_rows.append({
                            "Category": CATS[ckey]["label"],
                            "Alternative": name,
                            "Score A": round(va, 4),
                            "Score B": round(vb, 4),
                            "Δ": round(vb - va, 4),
                        })
            if cat_rows:
                st.subheader("Category score differences")
                st.dataframe(pd.DataFrame(cat_rows), use_container_width=True, hide_index=True)

    if st.button("Back to home", type="primary", key="cmp_back"):
        st.session_state.step = 0
        st.rerun()


init_state()


def ordered_sel_cats():
    return [c for c in CATEGORY_ORDER if c in st.session_state.sel_cats]


def ordered_l3_cats():
    return [c for c in CATEGORY_ORDER if c in st.session_state.l3_cats]


def get_full_indicators(ckey):
    cat = CATS[ckey]
    disabled = st.session_state.get("disabled_indicators", set())
    names, units, benefits = [], [], []
    for j, ind in enumerate(cat["indicators"]):
        if (ckey, j) not in disabled:
            names.append(ind)
            units.append(st.session_state.sel_units.get(f"{ckey}_{j}", cat["default_units"][j]))
            benefits.append(cat["benefit"][j])
    n_custom = st.session_state.custom_indicator_counts.get(ckey, 0)
    for ci in range(n_custom):
        info = st.session_state.custom_indicators.get((ckey, ci), {})
        names.append(info.get("name") or f"Custom indicator {ci+1}")
        units.append(info.get("unit") or "unit")
        benefits.append(info.get("benefit", True))
    return names, units, benefits


def get_raw_matrix(ckey):
    names, units, benefits = get_full_indicators(ckey)
    n_proc = st.session_state.n_proc
    raw = np.zeros((len(names), n_proc))
    for j in range(len(names)):
        for pi in range(n_proc):
            raw[j, pi] = st.session_state.indicator_values.get((ckey, j, pi), 0.0)
    return raw, benefits


OUTLIER_SD_THRESHOLD = 2.0
MIN_ALTS_FOR_OUTLIER = 3


def _indicator_values_fingerprint(cats=None):
    cats = cats or ordered_sel_cats()
    items = []
    for ckey in cats:
        ind_names, _, _ = get_full_indicators(ckey)
        for j in range(len(ind_names)):
            for pi in range(st.session_state.n_proc):
                val = st.session_state.indicator_values.get((ckey, j, pi), 0.0)
                items.append((ckey, j, pi, round(float(val), 6)))
    return tuple(sorted(items))


def detect_indicator_outliers(cats=None, threshold=OUTLIER_SD_THRESHOLD):
    """Flag indicator values more than threshold SD from the row mean across alternatives."""
    cats = cats or ordered_sel_cats()
    names = st.session_state.proc_names
    n_proc = len(names)
    if n_proc < MIN_ALTS_FOR_OUTLIER:
        return []

    outliers = []
    for ckey in cats:
        ind_names, _, _ = get_full_indicators(ckey)
        for j, ind_name in enumerate(ind_names):
            vals = np.array([
                st.session_state.indicator_values.get((ckey, j, pi), 0.0)
                for pi in range(n_proc)
            ], dtype=float)
            mean = float(vals.mean())
            std = float(vals.std(ddof=1)) if n_proc > 1 else 0.0
            if std == 0.0:
                continue
            for pi, val in enumerate(vals):
                z = abs(val - mean) / std
                if z > threshold:
                    outliers.append({
                        "Category": CATS[ckey]["label"],
                        "Indicator": ind_name,
                        "Alternative": names[pi],
                        "Value": round(float(val), 4),
                        "Mean": round(mean, 4),
                        "Std dev": round(std, 4),
                        "Z-score": round(z, 2),
                    })
    return outliers


def compute_category_score_from_raw(ckey, raw_override=None):
    names, units, benefits = get_full_indicators(ckey)
    n_ind = len(names)
    if raw_override is not None:
        raw = raw_override
    else:
        raw, _ = get_raw_matrix(ckey)
    nm = np.zeros_like(raw)
    n2 = np.zeros_like(raw)
    for j in range(n_ind):
        nm[j] = merec_norm(raw[j], benefits[j])
        n2[j] = n2_norm(raw[j], benefits[j])
    w = merec_weights(nm)
    score = (n2 * w[:, None]).sum(axis=0)
    return score


STEP_LABELS = [
    "1. Alternatives", "2. Categories", "3. Custom indicators", "4. Units",
    "5. Indicators", "6. Correlation check", "7. MEREC weights",
    "8. Category scores", "9. Level 3 categories", "10. Category weights",
    "11. MCDM methods", "12. Results", "13. Validation (optional)",
    "14. Analytics (optional)", "15. Decision support",
]

HOW_TO_USE_STEPS = [
    ("1–5", "Problem definition", "Name alternatives, select sustainability categories, configure indicators and units, then enter raw values in the decision matrix."),
    ("6", "Correlation check", "Review within-category indicator correlations and acknowledge before proceeding."),
    ("7–8", "Indicator processing", "Review MEREC weights, then inspect category scores (radar chart when multiple categories are selected)."),
    ("9–11", "Decision aggregation", "Choose Level 3 categories, select inter-category weighting methods, then pick MCDM methods for ranking."),
    ("12", "Results", "Compare MCDM rankings, PSI compromise scores, and category-combination sensitivity."),
    ("13", "Validation (optional)", "Run robustness checks — weight sensitivity, normalisation, Monte Carlo, bootstrap, and more."),
    ("14", "Analytics (optional)", "Contribution analysis, leave-one-out tests, and stakeholder preference simulation."),
    ("15", "Decision support", "Review the recommended winner with structured strengths and weaknesses tables."),
]

METHOD_DESCRIPTIONS = [
    {
        "Method": "MEREC",
        "Stage": "Indicator processing",
        "Plain-language description": "Assigns importance to indicators within each category based on how much each indicator influences the ranking when it changes.",
        "Role in PRISM": "Intra-category weighting — combines normalised indicators into a single category score.",
    },
    {
        "Method": "N2 normalisation",
        "Stage": "Indicator processing",
        "Plain-language description": "Scales raw indicator values to a common 0–1 range so benefit and cost indicators can be compared fairly.",
        "Role in PRISM": "Prepares the decision matrix before MEREC weighting and category scoring.",
    },
    {
        "Method": "Equal weights",
        "Stage": "Inter-category weighting",
        "Plain-language description": "Treats every sustainability category as equally important.",
        "Role in PRISM": "One of three objective methods used to derive category importance at Level 3.",
    },
    {
        "Method": "Entropy weights",
        "Stage": "Inter-category weighting",
        "Plain-language description": "Gives more weight to categories where alternatives differ most — higher variation means higher influence.",
        "Role in PRISM": "Objective weighting based on data dispersion across alternatives.",
    },
    {
        "Method": "CRITIC",
        "Stage": "Inter-category weighting",
        "Plain-language description": "Balances contrast between alternatives with correlation structure — categories that are both distinct and non-redundant receive higher weight.",
        "Role in PRISM": "Objective weighting that accounts for both contrast and inter-category correlation.",
    },
    {
        "Method": "RCW",
        "Stage": "Inter-category weighting",
        "Plain-language description": "Reciprocal Composite Weighting — harmonises multiple weighting methods into one consolidated category-weight vector.",
        "Role in PRISM": "Inter-category weighting and consolidation before MCDM aggregation.",
    },
    {
        "Method": "TOPSIS",
        "Stage": "MCDM aggregation",
        "Plain-language description": "Ranks alternatives by their distance to an ideal best solution and an ideal worst solution.",
        "Role in PRISM": "One of five MCDM methods whose ranks feed the PSI compromise index.",
    },
    {
        "Method": "VIKOR",
        "Stage": "MCDM aggregation",
        "Plain-language description": "Seeks a compromise solution closest to the group utility while minimising regret from not choosing the best individual criterion.",
        "Role in PRISM": "Provides a compromise-oriented ranking perspective for PSI.",
    },
    {
        "Method": "ELECTRE-Score",
        "Stage": "MCDM aggregation",
        "Plain-language description": "Uses outranking logic to score how strongly one alternative dominates others across criteria.",
        "Role in PRISM": "Adds an outranking-based view to the multi-method ensemble.",
    },
    {
        "Method": "MULTIMOORA",
        "Stage": "MCDM aggregation",
        "Plain-language description": "Combines three reference-point approaches (ratio, reference, and full multiplicative forms) for a balanced ranking.",
        "Role in PRISM": "Reference-point method contributing to rank diversity in PSI.",
    },
    {
        "Method": "WASPAS",
        "Stage": "MCDM aggregation",
        "Plain-language description": "Weighted Aggregated Sum Product Assessment — blends weighted sum and weighted product models.",
        "Role in PRISM": "Aggregated weighted model complementing the other MCDM techniques.",
    },
    {
        "Method": "PSI",
        "Stage": "Compromise ranking",
        "Plain-language description": "Performance Stability Index — balances average rank (performance) with rank consistency (stability) across MCDM methods.",
        "Role in PRISM": "Produces the final compromise ranking at p = 0.50 in Step 12.",
    },
    {
        "Method": "PSI-combo",
        "Stage": "Tie-breaker",
        "Plain-language description": "Applies PSI a second time using ranks across all category combinations when the headline PSI ranking is tied at rank 1.",
        "Role in PRISM": "Breaks ties when multiple alternatives share the top PSI rank.",
    },
]


def _render_method_descriptions_panel():
    """Render method descriptions table (intro + dataframe)."""
    st.markdown(
        "<p class='prism-about-body' style='font-family:Inter,sans-serif;"
        "font-size:11pt;line-height:1.75;color:#1A202C;margin-bottom:16px;'>"
        "Plain-language summary of each method used in PRISM — what it does and "
        "where it fits in the assessment workflow."
        "</p>",
        unsafe_allow_html=True,
    )
    st.dataframe(
        pd.DataFrame(METHOD_DESCRIPTIONS),
        use_container_width=True,
        hide_index=True,
    )


def method_descriptions_page():
    st.header("Method descriptions")
    _render_method_descriptions_panel()

    if st.button("Back to home", type="primary", key="md_back_home"):
        st.session_state.step = 0
        st.rerun()


def how_to_use_page():
    st.header("How to use PRISM")

    st.markdown(
        "<p class='prism-about-body' style='font-family:Inter,sans-serif;"
        "font-size:11pt;line-height:1.75;color:#1A202C;margin-bottom:16px;'>"
        "Follow the guided workflow in the sidebar from Step 1 through Step 12 "
        "for a complete assessment. Optional validation, analytics, and decision "
        "support steps are available after results are generated."
        "</p>",
        unsafe_allow_html=True,
    )

    guide_rows = [
        {"Steps": s, "Phase": phase, "What to do": action}
        for s, phase, action in HOW_TO_USE_STEPS
    ]
    st.dataframe(pd.DataFrame(guide_rows), use_container_width=True, hide_index=True)

    st.markdown(
        "<div class='prism-about-panel' style='margin-top:16px;'>"
        "<h3 class='prism-about-title'>Tips</h3>"
        "<p class='prism-about-body'>"
        "Save indicator drafts in Step 5 if you need to navigate away. "
        "Negative values and statistical outliers (&gt;2 SD) are flagged for review before proceeding. "
        "Run at least one validation check before finalising a recommendation. "
        "Use Step 15 Decision support for a structured winner summary."
        "</p></div>",
        unsafe_allow_html=True,
    )

    if st.button("Back to home", type="primary", key="htu_back_home"):
        st.session_state.step = 0
        st.rerun()


def build_excel_report():
    """Build a full Excel report from current session state."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HDR_FONT   = Font(name="Times New Roman", bold=True, color="FFFFFF", size=11)
    HDR_FILL   = PatternFill("solid", fgColor="0D2B5E")
    HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BODY_FONT  = Font(name="Times New Roman", size=10)
    BODY_ALIGN = Alignment(horizontal="center", vertical="center")
    TITLE_FONT = Font(name="Times New Roman", bold=True, size=12, color="0D2B5E")
    thin = Side(style="thin", color="D0D0D0")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = HDR_ALIGN
            cell.border = BORDER

    def style_body(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = BORDER

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

    def add_title(ws, title):
        ws.cell(1, 1, title).font = TITLE_FONT
        ws.row_dimensions[1].height = 20

    ss = st.session_state
    names    = ss.proc_names
    n_proc   = len(names)
    l3_cats  = ordered_l3_cats()

    # ── Sheet 1: Session Summary ──────────────────────────────────────────────
    ws1 = wb.create_sheet("1. Session Summary")
    add_title(ws1, "PRISM — Session Summary")
    rows_s = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Tool", "PRISM — Performance Ranking via Integrated Sustainability Metrics"),
        ("Institution", "Cranfield University"),
        ("Number of alternatives", n_proc),
        ("Alternatives", ", ".join(names)),
        ("Categories selected", ", ".join(CATS[c]["label"] for c in l3_cats)),
        ("MCDM methods", ", ".join(METHOD_LABELS[m] for m in (list(ss.sel_mcdm_methods) or ALL_MCDM_KEYS))),
        ("Weighting methods", ", ".join(ss.sel_weight_methods)),
    ]
    for ri, (k, v) in enumerate(rows_s, start=3):
        ws1.cell(ri, 1, k).font = Font(name="Times New Roman", bold=True, size=10)
        ws1.cell(ri, 2, str(v)).font = BODY_FONT
    auto_width(ws1)

    # ── Sheet 2: Decision Matrix ──────────────────────────────────────────────
    ws2 = wb.create_sheet("2. Decision Matrix")
    add_title(ws2, "Raw Indicator Values (Decision Matrix)")
    col = 1
    ws2.cell(3, 1, "Category").font = HDR_FONT
    ws2.cell(3, 1).fill = HDR_FILL
    ws2.cell(3, 1).alignment = HDR_ALIGN
    ws2.cell(3, 2, "Indicator").font = HDR_FONT
    ws2.cell(3, 2).fill = HDR_FILL
    ws2.cell(3, 2).alignment = HDR_ALIGN
    ws2.cell(3, 3, "Unit").font = HDR_FONT
    ws2.cell(3, 3).fill = HDR_FILL
    ws2.cell(3, 3).alignment = HDR_ALIGN
    ws2.cell(3, 4, "Type").font = HDR_FONT
    ws2.cell(3, 4).fill = HDR_FILL
    ws2.cell(3, 4).alignment = HDR_ALIGN
    for pi, name in enumerate(names):
        ws2.cell(3, 5 + pi, name).font = HDR_FONT
        ws2.cell(3, 5 + pi).fill = HDR_FILL
        ws2.cell(3, 5 + pi).alignment = HDR_ALIGN
    style_header(ws2, 3, 4 + n_proc)
    row = 4
    for ckey in l3_cats:
        ind_names, ind_units, benefits = get_full_indicators(ckey)
        for j in range(len(ind_names)):
            ws2.cell(row, 1, CATS[ckey]["label"]).font = BODY_FONT
            ws2.cell(row, 2, ind_names[j]).font = BODY_FONT
            ws2.cell(row, 3, ind_units[j]).font = BODY_FONT
            ws2.cell(row, 4, "Benefit" if benefits[j] else "Cost").font = BODY_FONT
            for pi in range(n_proc):
                val = ss.indicator_values.get((ckey, j, pi), 0.0)
                ws2.cell(row, 5 + pi, round(float(val), 4)).font = BODY_FONT
            style_body(ws2, row, 4 + n_proc)
            row += 1
    auto_width(ws2)

    # ── Sheet 3: MEREC Weights ────────────────────────────────────────────────
    if hasattr(ss, "merec_w") and ss.merec_w:
        ws3 = wb.create_sheet("3. MEREC Weights")
        add_title(ws3, "MEREC Indicator Weights")
        headers = ["Category", "Indicator", "Unit", "MEREC Weight"]
        for ci, h in enumerate(headers, 1):
            ws3.cell(3, ci, h).font = HDR_FONT
            ws3.cell(3, ci).fill = HDR_FILL
            ws3.cell(3, ci).alignment = HDR_ALIGN
        style_header(ws3, 3, 4)
        row = 4
        for ckey in l3_cats:
            ind_names, ind_units, _ = get_full_indicators(ckey)
            w_ind = ss.merec_w.get(ckey, np.zeros(len(ind_names)))
            for j in range(len(ind_names)):
                ws3.cell(row, 1, CATS[ckey]["label"]).font = BODY_FONT
                ws3.cell(row, 2, ind_names[j]).font = BODY_FONT
                ws3.cell(row, 3, ind_units[j]).font = BODY_FONT
                ws3.cell(row, 4, round(float(w_ind[j]), 4)).font = BODY_FONT
                style_body(ws3, row, 4)
                row += 1
        auto_width(ws3)

    # ── Sheet 4: Category Scores ──────────────────────────────────────────────
    if hasattr(ss, "cat_scores") and ss.cat_scores:
        ws4 = wb.create_sheet("4. Category Scores")
        add_title(ws4, "Category Scores (MEREC Weights × N2 Normalisation)")
        headers = ["Category"] + names
        for ci, h in enumerate(headers, 1):
            ws4.cell(3, ci, h).font = HDR_FONT
            ws4.cell(3, ci).fill = HDR_FILL
            ws4.cell(3, ci).alignment = HDR_ALIGN
        style_header(ws4, 3, 1 + n_proc)
        for ri, ckey in enumerate(l3_cats, start=4):
            ws4.cell(ri, 1, CATS[ckey]["label"]).font = BODY_FONT
            scores = ss.cat_scores.get(ckey, np.zeros(n_proc))
            for pi in range(n_proc):
                ws4.cell(ri, 2 + pi, round(float(scores[pi]), 4)).font = BODY_FONT
            style_body(ws4, ri, 1 + n_proc)
        auto_width(ws4)

    # ── Sheet 5: Category Weights ─────────────────────────────────────────────
    if hasattr(ss, "final_cat_weights") and ss.final_cat_weights is not None:
        ws5 = wb.create_sheet("5. Category Weights")
        add_title(ws5, "Category Weights (RCW Consolidation)")
        headers = ["Category", "RCW Weight", "RCW Weight (%)"]
        for ci, h in enumerate(headers, 1):
            ws5.cell(3, ci, h).font = HDR_FONT
            ws5.cell(3, ci).fill = HDR_FILL
            ws5.cell(3, ci).alignment = HDR_ALIGN
        style_header(ws5, 3, 3)
        for ri, ckey in enumerate(l3_cats, start=4):
            w = float(ss.final_cat_weights[ri - 4])
            ws5.cell(ri, 1, CATS[ckey]["label"]).font = BODY_FONT
            ws5.cell(ri, 2, round(w, 4)).font = BODY_FONT
            ws5.cell(ri, 3, round(w * 100, 1)).font = BODY_FONT
            style_body(ws5, ri, 3)
        auto_width(ws5)

    # ── Sheet 6: MCDM Results ─────────────────────────────────────────────────
    methods = list(ss.sel_mcdm_methods) or ALL_MCDM_KEYS
    mr = ss.get("last_method_ranks", {})
    if mr:
        ws6 = wb.create_sheet("6. MCDM Results")
        add_title(ws6, "MCDM Rankings and PSI Compromise Rank")
        headers = ["Alternative"] + [METHOD_LABELS[m] for m in methods]
        if len(methods) > 1:
            psi = calc_psi(mr, methods, 0.5)
            psi_ranks = rank_with_ties(psi, ascending=False)
            headers += ["PSI Score (p=0.5)", "PSI Rank"]
        psi_combo_scores = ss.get("last_psi_combo_scores")
        psi_combo_ranks = ss.get("last_psi_combo_ranks")
        has_combo = bool(psi_combo_scores) and len(psi_combo_scores) == n_proc
        if has_combo:
            headers += ["PSI-combo Score (combo tie-breaker)", "PSI-combo Rank"]
        for ci, h in enumerate(headers, 1):
            ws6.cell(3, ci, h).font = HDR_FONT
            ws6.cell(3, ci).fill = HDR_FILL
            ws6.cell(3, ci).alignment = HDR_ALIGN
        style_header(ws6, 3, len(headers))
        for pi, name in enumerate(names):
            row_data = [name] + [int(mr[m][pi]) for m in methods if m in mr]
            if len(methods) > 1:
                row_data += [round(float(psi[pi]), 4), int(psi_ranks[pi])]
            if has_combo:
                row_data += [round(float(psi_combo_scores[pi]), 4), int(psi_combo_ranks[pi])]
            for ci, val in enumerate(row_data, 1):
                ws6.cell(4 + pi, ci, val).font = BODY_FONT
            style_body(ws6, 4 + pi, len(headers))
        auto_width(ws6)

    # ── Sheet 7: PSI Curve Data ───────────────────────────────────────────────
    if mr and len(methods) > 1:
        ws7 = wb.create_sheet("7. PSI Curve Data")
        add_title(ws7, "PSI Score vs p Value")
        headers = ["p value"] + names
        for ci, h in enumerate(headers, 1):
            ws7.cell(3, ci, h).font = HDR_FONT
            ws7.cell(3, ci).fill = HDR_FILL
            ws7.cell(3, ci).alignment = HDR_ALIGN
        style_header(ws7, 3, len(headers))
        p_range = np.linspace(0, 1, 21)
        for ri, p in enumerate(p_range, start=4):
            psi_p = calc_psi(mr, methods, float(p))
            ws7.cell(ri, 1, round(float(p), 2)).font = BODY_FONT
            for pi in range(n_proc):
                ws7.cell(ri, 2 + pi, round(float(psi_p[pi]), 4)).font = BODY_FONT
            style_body(ws7, ri, len(headers))
        auto_width(ws7)

    # Finalise
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


with st.sidebar:
    st.markdown(
        f"<div style='margin-bottom:4px;'>"
        f"<span style='font-size:20px;font-weight:700;color:{_BRAND_NAVY};"
        f"font-family:{_FONT_CSS};letter-spacing:-0.02em;'>PRISM</span></div>"
        f"<p style='font-size:10pt;color:{_BRAND_MUTED};margin:0;line-height:1.4;"
        f"font-family:{_FONT_CSS};'>Performance Ranking via Integrated "
        f"Sustainability Metrics</p>"
        f"<span class='prism-sidebar-badge'>MCDM Assessment Tool</span>",
        unsafe_allow_html=True,
    )
    st.divider()

    step = st.session_state.step

    render_assessments_sidebar()

    st.markdown(
        f"<p style='font-size:10pt;font-weight:600;color:{_BRAND_NAVY};"
        f"text-transform:uppercase;letter-spacing:0.06em;"
        f"font-family:{_FONT_CSS};margin:0 0 8px 0;'>Documentation</p>",
        unsafe_allow_html=True,
    )
    if st.button("How to use", use_container_width=True, key="sidebar_how_to_use",
                 help="Step-by-step guide for the full PRISM assessment workflow."):
        st.session_state.step = -1
        st.rerun()
    if st.button("Method descriptions", use_container_width=True, key="sidebar_method_descriptions",
                 help="Plain-language summary of every method used in PRISM."):
        st.session_state.step = -2
        st.rerun()
    if step != 0 and st.button("Home", use_container_width=True, key="sidebar_home",
                               help="Return to the PRISM landing page."):
        st.session_state.step = 0
        st.rerun()

    st.divider()

    SECTIONS = {
        "Problem Definition": [1, 2, 3, 4, 5],
        "Indicator Processing": [6, 7, 8],
        "Decision Aggregation": [9, 10, 11, 12],
        "Validation": [13],
        "Analytics": [14],
        "Decision Support": [15],
    }

    total_steps = len(STEP_LABELS)
    if step == -3:
        st.caption("Compare assessments")
    elif step < 0:
        st.caption("Documentation")
        st.markdown(
            "<div class='prism-progress-wrap'>"
            "<div class='prism-progress-label'>"
            "<span>Assessment progress</span>"
            "<span class='prism-progress-pct'>—</span></div>"
            "<div class='prism-progress-track'>"
            "<div class='prism-progress-fill' style='width:0%'></div>"
            "</div>"
            "<p class='prism-progress-caption'>Reference material</p>"
            "</div>",
            unsafe_allow_html=True,
        )
    elif step == 0:
        st.caption("Landing page")
        st.markdown(
            "<div class='prism-progress-wrap'>"
            "<div class='prism-progress-label'>"
            "<span>Assessment progress</span>"
            "<span class='prism-progress-pct'>0%</span></div>"
            "<div class='prism-progress-track'>"
            "<div class='prism-progress-fill' style='width:0%'></div>"
            "</div>"
            "<p class='prism-progress-caption'>Ready to begin</p>"
            "</div>",
            unsafe_allow_html=True,
        )
    else:
        pct = min(step, total_steps) / total_steps if total_steps else 0
        pct_int = int(round(pct * 100))
        st.markdown(
            f"<div class='prism-progress-wrap'>"
            f"<div class='prism-progress-label'>"
            f"<span>Assessment progress</span>"
            f"<span class='prism-progress-pct'>{pct_int}%</span></div>"
            f"<div class='prism-progress-track'>"
            f"<div class='prism-progress-fill' style='width:{pct_int}%'></div>"
            f"</div>"
            f"<p class='prism-progress-caption'>"
            f"Step {min(step, total_steps)} of {total_steps}</p>"
            f"</div>",
            unsafe_allow_html=True,
        )

    for section, steps in SECTIONS.items():
        st.markdown(
            f"<p class='prism-section-head' style='font-size:10pt;font-weight:600;"
            f"color:{_BRAND_NAVY};text-transform:uppercase;letter-spacing:0.06em;"
            f"font-family:{_FONT_CSS};margin:10px 0 4px 0;'>{section}</p>",
            unsafe_allow_html=True,
        )
        for s in steps:
            if s >= len(STEP_LABELS) + 1:
                continue
            label = STEP_LABELS[s - 1]
            short = label.split(". ", 1)[1] if ". " in label else label
            if s < step:
                row_cls = "prism-step-row"
                check_cls = "prism-step-check done"
                check_mark = "&#10003;"
                label_cls = "prism-step-label done"
            elif s == step:
                row_cls = "prism-step-row current"
                check_cls = "prism-step-check current"
                check_mark = ""
                label_cls = "prism-step-label current"
            else:
                row_cls = "prism-step-row"
                check_cls = "prism-step-check pending"
                check_mark = ""
                label_cls = "prism-step-label pending"
            st.markdown(
                f"<div class='{row_cls}'>"
                f"<span class='{check_cls}'>{check_mark}</span>"
                f"<span class='{label_cls}'>{short}</span></div>",
                unsafe_allow_html=True,
            )

    st.divider()

    # Download report — always accessible
    if OPENPYXL_OK:
        mr_ready = bool(st.session_state.get("last_method_ranks", {}))
        if mr_ready:
            try:
                excel_buf = build_excel_report()
                fname = f"PRISM_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                st.download_button(
                    label="Download Excel report",
                    data=excel_buf,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    help="Export inputs, weights, rankings, and validation results to Excel.",
                )
            except Exception as e:
                st.caption(f"Report error: {e}")
        else:
            st.markdown(
                "<p style='font-size:11px;color:#9CA3AF;text-align:center;"
                "font-family:Times New Roman,serif;'>"
                "Run through Step 12 to enable report download.</p>",
                unsafe_allow_html=True,
            )
    else:
        st.caption("Install openpyxl to enable Excel export.")

    st.divider()
    if st.button("Reset", use_container_width=True,
                 help="Clear all session data and restart from the landing page."):
        reset_all()
        st.rerun()


# ============================================================================
# STEP 1 - PROCESSES (FIX 1: no example names pre-filled)
# ============================================================================

def landing_page():
    st.markdown(
        f"<div class='prism-hero'>"
        f"<h1>PRISM</h1>"
        f"<p>Performance Ranking via Integrated Sustainability Metrics</p></div>",
        unsafe_allow_html=True,
    )

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(
            "<div class='prism-card'>"
            "<span class='prism-step-pill'>Ranking</span>"
            "<h4>Multi-method MCDM aggregation</h4>"
            "<p>TOPSIS, VIKOR, ELECTRE-Score, MULTIMOORA, and WASPAS "
            "combined through the Performance Stability Index (PSI) "
            "for robust compromise ranking.</p></div>",
            unsafe_allow_html=True,
        )
    with c2:
        st.markdown(
            "<div class='prism-card'>"
            "<span class='prism-step-pill'>Validation</span>"
            "<h4>Integrated robustness checks</h4>"
            "<p>Seven validation modules covering weight sensitivity, "
            "normalisation, Monte Carlo uncertainty, bootstrap stability, "
            "and indicator-level propagation.</p></div>",
            unsafe_allow_html=True,
        )
    with c3:
        st.markdown(
            "<div class='prism-card'>"
            "<span class='prism-step-pill'>Reporting</span>"
            "<h4>Audit-ready outputs</h4>"
            "<p>Structured Excel export of inputs, weights, rankings, "
            "and validation results for internal review and "
            "stakeholder communication.</p></div>",
            unsafe_allow_html=True,
        )

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

    st.markdown(
        "<div class='prism-framework'>"
        "<p class='prism-framework-heading'>PRISM framework architecture</p>"
        "<div class='prism-framework-level level-01'>"
        "<div class='prism-framework-level-title'>Level 01 — System definition</div>"
        "<div class='prism-framework-row'>"
        "<div class='prism-framework-box wide'>"
        "<strong>Define goal and scope</strong>"
        "<span class='sub'>Manufacturing routes, materials, end-use applications, "
        "system boundaries, functional units</span></div>"
        "<div class='prism-framework-arrow'>&#8596;</div>"
        "<div class='prism-framework-box wide'>"
        "<strong>Define indicators</strong>"
        "<span class='sub'>Alternatives, categories, and sustainability criteria</span>"
        "</div></div></div>"
        "<div class='prism-framework-level level-02'>"
        "<div class='prism-framework-level-title'>Level 02 — Indicator processing</div>"
        "<div class='prism-framework-flow'>"
        "<div class='prism-framework-box'>"
        "<strong>Decision matrix</strong>"
        "<span class='sub'>Environmental, economic, social, quality, productivity</span>"
        "</div><div class='prism-framework-arrow'>&#8594;</div>"
        "<div class='prism-framework-box'>"
        "<strong>Normalisation</strong>"
        "<span class='sub'>Linear sum-based (N2)</span></div>"
        "<div class='prism-framework-arrow'>&#8594;</div>"
        "<div class='prism-framework-box'>"
        "<strong>Intra-category weighting</strong>"
        "<span class='sub'>MEREC weights</span></div>"
        "<div class='prism-framework-arrow'>&#8594;</div>"
        "<div class='prism-framework-box'>"
        "<strong>Category scores</strong>"
        "<span class='sub'>&Sigma; MEREC weights &times; normalisation</span>"
        "</div></div>"
        "<div class='prism-framework-side'>"
        "<div class='prism-framework-box dashed'>"
        "<strong>Data quality assessment</strong></div>"
        "<div class='prism-framework-box dashed'>"
        "<strong>Correlation analysis</strong></div>"
        "<div class='prism-framework-box dashed'>"
        "<strong>Category selection for Level 03</strong></div>"
        "</div></div>"
        "<div class='prism-framework-level level-03'>"
        "<div class='prism-framework-level-title'>Level 03 — Decision aggregation</div>"
        "<div class='prism-framework-flow'>"
        "<div class='prism-framework-box'>"
        "<strong>Inter-category weighting</strong>"
        "<span class='sub'>Equal, entropy, CRITIC</span></div>"
        "<div class='prism-framework-arrow'>&#8594;</div>"
        "<div class='prism-framework-box'>"
        "<strong>Weight consolidation</strong>"
        "<span class='sub'>Reciprocal composite weighting (RCW)</span></div>"
        "<div class='prism-framework-arrow'>&#8594;</div>"
        "<div class='prism-framework-box'>"
        "<strong>MCDM aggregation</strong>"
        "<span class='sub'>TOPSIS, VIKOR, ELECTRE-Score, MULTIMOORA, WASPAS</span>"
        "</div><div class='prism-framework-arrow'>&#8594;</div>"
        "<div class='prism-framework-box'>"
        "<strong>Compromise ranking</strong>"
        "<span class='sub'>Performance stability index (PSI)</span></div>"
        "</div>"
        "<div class='prism-framework-side'>"
        "<div class='prism-framework-box dashed'>"
        "<strong>Sensitivity and uncertainty analysis</strong>"
        "<span class='sub'>Weighting, normalisation, bootstrap, MCDM, rank reversal</span>"
        "</div>"
        "<div class='prism-framework-box dashed'>"
        "<strong>Category combination generation</strong>"
        "<span class='sub'>2<sup>n</sup>&nbsp;&minus;&nbsp;1 stakeholder combinations</span>"
        "</div>"
        "<div class='prism-framework-box dashed'>"
        "<strong>Decision support analytics</strong>"
        "<span class='sub'>Contribution, leave-one-out, simulation</span>"
        "</div></div></div></div>",
        unsafe_allow_html=True,
    )

    st.markdown(
        f"<p style='font-size:11pt;font-weight:600;color:{_BRAND_NAVY};"
        f"text-transform:uppercase;letter-spacing:0.05em;"
        f"font-family:{_FONT_CSS};margin-bottom:12px;'>Assessment workflow</p>",
        unsafe_allow_html=True,
    )
    w1, w2, w3, w4 = st.columns(4)
    workflow = [
        ("1", "Define", "Alternatives, categories, indicators"),
        ("2", "Process", "MEREC, N2, category scores"),
        ("3", "Aggregate", "RCW weights, MCDM, PSI ranking"),
        ("4", "Validate", "Sensitivity and uncertainty analysis"),
    ]
    for col, (num, lbl, sub) in zip([w1, w2, w3, w4], workflow):
        with col:
            st.markdown(
                f"<div class='prism-workflow-step'>"
                f"<div class='num'>{num}</div>"
                f"<div class='lbl'>{lbl}</div>"
                f"<div class='sub'>{sub}</div></div>",
                unsafe_allow_html=True,
            )

    st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

    st.markdown(
        "<div class='prism-about-panel'>"
        "<h3 class='prism-about-title'>About PRISM</h3>"
        "<p class='prism-about-body'>"
        "PRISM is an integrated multi-criteria decision-making framework "
        "for the comparative sustainability assessment of manufacturing "
        "alternatives. It combines intra-category weighting through MEREC, "
        "inter-category weighting and consolidation via Reciprocal Composite "
        "Weighting (RCW), and multi-method MCDM aggregation producing a final "
        "compromise ranking through the Performance Stability Index (PSI)."
        "</p>"
        "<p class='prism-about-body'>"
        "The framework evaluates alternatives across five sustainability "
        "dimensions:"
        "</p>"
        "<div class='prism-dimension-row'>"
        "<span class='prism-dimension-pill'>Environmental</span>"
        "<span class='prism-dimension-pill'>Economic</span>"
        "<span class='prism-dimension-pill'>Social</span>"
        "<span class='prism-dimension-pill'>Quality</span>"
        "<span class='prism-dimension-pill'>Productivity</span>"
        "</div>"
        "<p class='prism-about-body'>"
        "An integrated validation and analytics layer provides robustness "
        "evidence to support rigorous, evidence-based decision-making."
        "</p>"
        "</div>",
        unsafe_allow_html=True,
    )

    st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        if st.button("Begin assessment", type="primary",
                     use_container_width=True, key="start_btn"):
            st.session_state.step = 1
            st.rerun()



def step1():
    st.header("Step 1 - Define alternatives")

    n = st.slider("Number of alternatives", min_value=2, max_value=10,
                   value=st.session_state.n_proc, key="n_proc_slider",
                   help=FIELD_HELP["n_alternatives"])
    st.session_state.n_proc = n

    names = st.session_state.proc_names
    if len(names) < n:
        names = names + [""] * (n - len(names))
    elif len(names) > n:
        names = names[:n]
    st.session_state.proc_names = names

    st.subheader("Name each alternative")
    cols = st.columns(min(n, 5))
    new_names = []
    for i in range(n):
        with cols[i % len(cols)]:
            val = st.text_input(f"Alternative {i+1}", value=names[i], key=f"pname_{i}",
                                 placeholder="enter name",
                                 help=FIELD_HELP["alternative_name"])
            new_names.append(val.strip())
    st.session_state.proc_names = new_names

    st.divider()
    if st.button("Next ->", type="primary"):
        final_names = [n.strip() if n.strip() else f"P{i+1}" for i, n in enumerate(new_names)]
        st.session_state.proc_names = final_names
        st.session_state.step = 2
        st.rerun()


def step2():
    st.header("Step 2 - Select assessment categories")

    cols = st.columns(len(CATEGORY_ORDER))
    for i, key in enumerate(CATEGORY_ORDER):
        cat = CATS[key]
        with cols[i]:
            checked = key in st.session_state.sel_cats
            new_val = st.checkbox(cat["label"], value=checked, key=f"catchk_{key}",
                                  help=FIELD_HELP["category"])
            if new_val:
                st.session_state.sel_cats.add(key)
            else:
                st.session_state.sel_cats.discard(key)

    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("<- Back"):
            st.session_state.step = 1
            st.rerun()
    with c2:
        if st.button("Next ->", type="primary"):
            if not st.session_state.sel_cats:
                st.error("Select at least one category.")
            else:
                st.session_state.step = 3
                st.rerun()


def step3():
    st.header("Step 3 - Add custom indicators (optional)")

    use_custom = st.radio(
        "Do you want to add any indicators beyond the predefined list?",
        ["No, use the predefined indicators only", "Yes, I want to add custom indicators"],
        index=0 if st.session_state.use_custom_indicators in (None, False) else 1,
        key="use_custom_radio",
        help=FIELD_HELP["use_custom"],
    )
    st.session_state.use_custom_indicators = use_custom.startswith("Yes")

    if st.session_state.use_custom_indicators:
        st.divider()
        st.subheader("How many custom indicators per category?")
        cats = ordered_sel_cats()
        cols = st.columns(len(cats))
        for i, ckey in enumerate(cats):
            cat = CATS[ckey]
            with cols[i]:
                cnt = st.number_input(
                    cat["label"], min_value=0, max_value=5,
                    value=st.session_state.custom_indicator_counts.get(ckey, 0),
                    key=f"customcnt_{ckey}",
                    help=FIELD_HELP["custom_count"],
                )
                st.session_state.custom_indicator_counts[ckey] = int(cnt)

        any_custom = any(v > 0 for v in st.session_state.custom_indicator_counts.values())
        if any_custom:
            st.divider()
            st.subheader("Define each custom indicator")
            for ckey in cats:
                n_custom = st.session_state.custom_indicator_counts.get(ckey, 0)
                if n_custom == 0:
                    continue
                cat = CATS[ckey]
                st.markdown(f"**{cat['label']}**")
                for ci in range(n_custom):
                    info = st.session_state.custom_indicators.get((ckey, ci), {})
                    c1, c2, c3 = st.columns([2, 1, 1])
                    with c1:
                        name = st.text_input(
                            f"Indicator name #{ci+1}", value=info.get("name", ""),
                            key=f"custname_{ckey}_{ci}", placeholder="e.g. Noise level",
                            help=FIELD_HELP["indicator_name"],
                        )
                    with c2:
                        unit = st.text_input(
                            "Unit", value=info.get("unit", ""),
                            key=f"custunit_{ckey}_{ci}", placeholder="e.g. dB",
                            help=FIELD_HELP["indicator_unit"],
                        )
                    with c3:
                        benefit = st.selectbox(
                            "Direction", ["Cost (lower better)", "Benefit (higher better)"],
                            index=1 if info.get("benefit", False) else 0,
                            key=f"custben_{ckey}_{ci}",
                            help=FIELD_HELP["indicator_direction"],
                        )
                    st.session_state.custom_indicators[(ckey, ci)] = {
                        "name": name.strip(), "unit": unit.strip() or "unit",
                        "benefit": benefit.startswith("Benefit"),
                    }
                st.write("")

    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("<- Back"):
            st.session_state.step = 2
            st.rerun()
    with c2:
        if st.button("Next ->", type="primary"):
            st.session_state.step = 4
            st.rerun()


def step4():
    st.header("Step 4 - Select units and enable indicators")

    disabled = st.session_state.get("disabled_indicators", set())

    for ckey in ordered_sel_cats():
        cat = CATS[ckey]
        st.markdown(
            f"<span style='background:{cat['bg']};color:{cat['color']};"
            f"padding:2px 10px;border-radius:12px;font-size:13px;font-weight:600;"
            f"font-family:Times New Roman,Tinos,Times,serif;'>"
            f"{cat['label']}</span>", unsafe_allow_html=True,
        )
        st.write("")

        # Column headers
        hc1, hc2, hc3, hc4 = st.columns([0.4, 2, 1.5, 1.5])
        with hc1: st.markdown("**Use**")
        with hc2: st.markdown("**Indicator**")
        with hc3: st.markdown("**Unit**")
        with hc4: st.markdown("**Custom unit**")

        for j, ind in enumerate(cat["indicators"]):
            key = f"{ckey}_{j}"
            is_enabled = (ckey, j) not in disabled
            current = st.session_state.sel_units.get(key, cat["default_units"][j])
            options = list(cat["unit_options"][j])
            is_preset = current in options
            display_options = options + [CUSTOM_SENTINEL]
            default_index = options.index(current) if is_preset else len(options)

            c1, c2, c3, c4 = st.columns([0.4, 2, 1.5, 1.5])
            with c1:
                enabled = st.checkbox(
                    "use", value=is_enabled,
                    key=f"ind_enabled_{ckey}_{j}",
                    label_visibility="collapsed",
                    help=FIELD_HELP["indicator_enable"],
                )
                if enabled:
                    disabled.discard((ckey, j))
                else:
                    disabled.add((ckey, j))
            with c2:
                if not enabled:
                    tnr_label(ind, color="#aaaaaa", strike=True)
                else:
                    tnr_label(ind)
            with c3:
                chosen = st.selectbox(
                    "unit", display_options, index=default_index,
                    key=f"unitsel_{key}", label_visibility="collapsed",
                    disabled=not enabled,
                    help=FIELD_HELP["unit_select"],
                )
            with c4:
                if chosen == CUSTOM_SENTINEL and enabled:
                    custom_val = st.text_input(
                        "custom unit", value=current if not is_preset else "",
                        key=f"unitcustom_{key}", label_visibility="collapsed",
                        placeholder="type unit",
                        help=FIELD_HELP["custom_unit"],
                    )
                    final_unit = custom_val.strip() or "unit"
                else:
                    final_unit = chosen if chosen != CUSTOM_SENTINEL else current
                    st.caption(" ")
            st.session_state.sel_units[key] = final_unit

        st.session_state.disabled_indicators = disabled
        st.divider()

    c1, c2 = st.columns(2)
    with c1:
        if st.button("<- Back"):
            st.session_state.step = 3
            st.rerun()
    with c2:
        if st.button("Next ->", type="primary"):
            # Validate at least 1 indicator per selected category
            errors = []
            disabled = st.session_state.get("disabled_indicators", set())
            for ckey in ordered_sel_cats():
                cat = CATS[ckey]
                enabled_count = sum(
                    1 for j in range(len(cat["indicators"]))
                    if (ckey, j) not in disabled
                ) + st.session_state.custom_indicator_counts.get(ckey, 0)
                if enabled_count < 1:
                    errors.append(f"{cat['label']}: at least one indicator must be enabled.")
            if errors:
                for e in errors:
                    st.error(e)
            else:
                # Clear any stale seed keys for step5 so editor rebuilds
                for k in [k for k in st.session_state if k.startswith("editor_seed_")]:
                    del st.session_state[k]
                st.session_state.step = 5
                st.rerun()


# ============================================================================
# STEP 5 - INDICATOR VALUES
# FIX 2: values no longer reset to zero. The data_editor's own widget state
# (keyed per category, in st.session_state[editor_key]) is the single
# source of truth across reruns. We only build a seed DataFrame the FIRST
# time the key appears in session_state; on every subsequent rerun
# Streamlit reuses the live widget state instead of overwriting it.
# ============================================================================

def step5():
    st.header("Step 5 - Enter indicator values")

    if st.session_state.get("collaborator_mode"):
        aname = st.session_state.get("assessment_name") or "Shared assessment"
        cats = ", ".join(CATS[c]["label"] for c in ordered_sel_cats_for_editing())
        st.info(
            f"Collaborator mode — editing **{cats}** for assessment **{aname}**. "
            "Save your category data to cloud when finished."
        )

    names = st.session_state.proc_names
    edit_cats = ordered_sel_cats_for_editing()

    if st.session_state.get("collaborator_mode") and not edit_cats:
        st.warning("No assigned categories match this assessment. Re-join with valid categories.")
        return

    for ckey in edit_cats:
        cat = CATS[ckey]
        ind_names, ind_units, _ = get_full_indicators(ckey)

        st.markdown(
            f"<span style='background:{cat['bg']};color:{cat['color']};"
            f"padding:2px 10px;border-radius:12px;font-size:13px;font-weight:600;"
            f"font-family:Times New Roman,Tinos,Times,serif;'>"
            f"{cat['label']}</span>", unsafe_allow_html=True,
        )

        rows = [f"{ind_names[j]} ({ind_units[j]})" for j in range(len(ind_names))]
        editor_key = f"editor_{ckey}"
        seed_key = f"editor_seed_{ckey}"

        if seed_key not in st.session_state:
            seed = [[st.session_state.indicator_values.get((ckey, j, pi), 0.0)
                      for pi in range(len(names))] for j in range(len(ind_names))]
            st.session_state[seed_key] = pd.DataFrame(seed, index=rows, columns=names)

        # column_config disables sort arrows on process name columns
        col_cfg = {name: st.column_config.NumberColumn(
            name, help=FIELD_HELP["indicator_value"],
        ) for name in names}
        edited = st.data_editor(
            st.session_state[seed_key],
            key=editor_key,
            use_container_width=True,
            column_config=col_cfg,
        )

        for j in range(len(ind_names)):
            for pi in range(len(names)):
                st.session_state.indicator_values[(ckey, j, pi)] = float(edited.iloc[j, pi])

        negatives = []
        for j in range(len(ind_names)):
            for pi in range(len(names)):
                val = float(edited.iloc[j, pi])
                if val < 0:
                    negatives.append({
                        "Indicator": ind_names[j],
                        "Alternative": names[pi],
                        "Value": val,
                    })
        if negatives:
            st.warning(
                f"Negative values detected in **{cat['label']}** — "
                "verify entries are correct before proceeding."
            )
            st.dataframe(
                pd.DataFrame(negatives),
                use_container_width=True,
                hide_index=True,
            )

        st.write("")

    value_fp = _indicator_values_fingerprint(edit_cats)
    if st.session_state.get("_outlier_ack_fp") != value_fp:
        st.session_state.outliers_acknowledged = False

    outliers = detect_indicator_outliers(edit_cats)
    if outliers:
        st.warning(
            "**Statistical outliers detected** — one or more values fall more than "
            f"{OUTLIER_SD_THRESHOLD:.0f} standard deviations from the indicator mean "
            "across alternatives. Review before proceeding."
        )
        st.dataframe(
            pd.DataFrame(outliers),
            use_container_width=True,
            hide_index=True,
        )
        if not st.session_state.get("collaborator_mode"):
            ack = st.checkbox(
                "I have reviewed the flagged outliers and confirm they are correct",
                value=st.session_state.get("outliers_acknowledged", False),
                key="outlier_ack_cb",
            )
            st.session_state.outliers_acknowledged = ack
            if ack:
                st.session_state._outlier_ack_fp = value_fp

    if st.session_state.get("collaborator_mode") and st.session_state.get("current_assessment_id"):
        st.divider()
        if st.button("Save my categories to cloud", type="primary", key="collab_cloud_save"):
            saved, err = save_collaborator_category_merge(
                st.session_state.current_assessment_id,
                st.session_state.assigned_categories,
            )
            if err:
                st.error(err)
            else:
                for ckey in st.session_state.assigned_categories:
                    upsert_category_assignment(
                        st.session_state.current_assessment_id,
                        ckey,
                        st.session_state.get("owner_label") or "Collaborator",
                        "complete",
                    )
                st.success("Your category data has been saved to the shared assessment.")

    # ── Save / Load Draft ────────────────────────────────────────────────────
    if not st.session_state.get("collaborator_mode"):
        st.divider()
        dc1, dc2 = st.columns(2)
        with dc1:
            if st.button("💾 Save draft", key="save_draft_btn",
                         help="Saves all entered values in memory. Restores them if you navigate away."):
                draft = {}
                for ckey in ordered_sel_cats():
                    editor_key = f"editor_{ckey}"
                    seed_key   = f"editor_seed_{ckey}"
                    if editor_key in st.session_state:
                        df = st.session_state[editor_key]
                    elif seed_key in st.session_state:
                        df = st.session_state[seed_key]
                    else:
                        continue
                    draft[ckey] = df.values.tolist()
                st.session_state["_step5_draft"] = draft
                st.success("Draft saved — values preserved across navigation.")
        with dc2:
            has_draft = "_step5_draft" in st.session_state
            if st.button("📂 Load draft", key="load_draft_btn",
                         disabled=not has_draft,
                         help="Restore previously saved values."):
                draft = st.session_state["_step5_draft"]
                for ckey, values in draft.items():
                    ind_names, ind_units, _ = get_full_indicators(ckey)
                    rows = [f"{ind_names[j]} ({ind_units[j]})"
                            for j in range(len(ind_names))]
                    df = pd.DataFrame(values, index=rows,
                                      columns=st.session_state.proc_names)
                    st.session_state[f"editor_seed_{ckey}"] = df
                    for j in range(len(ind_names)):
                        for pi in range(len(st.session_state.proc_names)):
                            st.session_state.indicator_values[(ckey, j, pi)] = float(values[j][pi])
                st.success("Draft loaded.")
                st.rerun()
            if not has_draft:
                st.caption("No draft saved yet.")
        st.divider()
        c1, c2 = st.columns(2)
        with c1:
            if st.button("<- Back"):
                st.session_state.step = 4
                st.rerun()
        with c2:
            if st.button("Next ->", type="primary"):
                has_values = any(v != 0 for v in st.session_state.indicator_values.values())
                if not has_values:
                    st.error("Enter at least some values before proceeding.")
                elif outliers and not st.session_state.get("outliers_acknowledged"):
                    st.error("Review flagged outliers before proceeding.")
                else:
                    st.session_state.corr_acknowledged = False
                    st.session_state.step = 6
                    st.rerun()


def step6():
    st.header("Step 6 - Within-category correlation analysis")

    names = st.session_state.proc_names
    n_proc = len(names)
    flagged_any = False

    for ckey in ordered_sel_cats():
        cat = CATS[ckey]
        ind_names, _, _ = get_full_indicators(ckey)
        n_ind = len(ind_names)

        st.markdown(
            f"<span style='background:{cat['bg']};color:{cat['color']};"
            f"padding:2px 10px;border-radius:12px;font-size:13px;font-weight:600;"
            f"font-family:Times New Roman,Tinos,Times,serif;'>"
            f"{cat['label']}</span>", unsafe_allow_html=True,
        )

        if n_ind < 2:
            st.write("")
            continue
        if n_proc < 3:
            st.write("")
            continue

        raw = np.zeros((n_ind, n_proc))
        for j in range(n_ind):
            for pi in range(n_proc):
                raw[j, pi] = st.session_state.indicator_values.get((ckey, j, pi), 0.0)

        corr_mat = np.eye(n_ind)
        for a in range(n_ind):
            for b in range(n_ind):
                if a == b:
                    continue
                rho, _ = spearmanr(raw[a], raw[b])
                corr_mat[a, b] = rho if not np.isnan(rho) else 0.0

        df_corr = pd.DataFrame(corr_mat, index=ind_names, columns=ind_names).round(3)

        st.dataframe(df_corr, use_container_width=True)

        pairs_flagged = []
        for a in range(n_ind):
            for b in range(a + 1, n_ind):
                if abs(corr_mat[a, b]) > 0.8:
                    pairs_flagged.append((ind_names[a], ind_names[b], corr_mat[a, b]))

        if pairs_flagged:
            flagged_any = True
            for a, b, r in pairs_flagged:
                st.warning(f"**{a}** and **{b}** are highly correlated (rho = {r:.3f}).")

        st.write("")

    if flagged_any:
        st.info(
            "Highly correlated indicator pairs were found. This is shown for your "
            "awareness - both indicators remain in the analysis."
        )

    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("<- Back"):
            st.session_state.step = 5
            st.rerun()
    with c2:
        if st.button("Accept and proceed ->", type="primary"):
            st.session_state.corr_acknowledged = True
            compute_level2()
            st.session_state.step = 7
            st.rerun()


def compute_level2():
    n_proc = st.session_state.n_proc
    nm_data, n2_data, merec_w, cat_scores = {}, {}, {}, {}

    for ckey in ordered_sel_cats():
        ind_names, ind_units, benefits = get_full_indicators(ckey)
        n_ind = len(ind_names)

        raw = np.zeros((n_ind, n_proc))
        for j in range(n_ind):
            for pi in range(n_proc):
                raw[j, pi] = st.session_state.indicator_values.get((ckey, j, pi), 0.0)

        nm = np.zeros((n_ind, n_proc))
        n2 = np.zeros((n_ind, n_proc))
        for j in range(n_ind):
            nm[j] = merec_norm(raw[j], benefits[j])
            n2[j] = n2_norm(raw[j], benefits[j])

        w = merec_weights(nm)
        scores = (n2 * w[:, None]).sum(axis=0)

        nm_data[ckey] = nm
        n2_data[ckey] = n2
        merec_w[ckey] = w
        cat_scores[ckey] = scores

    st.session_state.nm_data = nm_data
    st.session_state.n2_data = n2_data
    st.session_state.merec_w = merec_w
    st.session_state.cat_scores = cat_scores
    st.session_state.computed = True


# ============================================================================
# STEP 7 - MEREC WEIGHTS
# FIX 4: only the single, final MEREC weight column is shown; per-process
# intermediate normalisation columns are no longer displayed.
# ============================================================================

def step7():
    st.header("Step 7 - MEREC weights")

    # ── Data-driven normalisation recommendation ──────────────────────────────
    names_s7 = st.session_state.proc_names
    n_proc_s7 = len(names_s7)
    all_raw_s7 = []
    has_zeros_s7 = False
    for ckey_s7 in ordered_sel_cats():
        ind_names_s7, _, benefits_s7 = get_full_indicators(ckey_s7)
        for j in range(len(ind_names_s7)):
            row_s7 = [st.session_state.indicator_values.get((ckey_s7,j,pi), 0.0)
                      for pi in range(n_proc_s7)]
            all_raw_s7.extend(row_s7)
            if any(v == 0.0 for v in row_s7):
                has_zeros_s7 = True

    if all_raw_s7:
        arr_s7    = np.array(all_raw_s7)
        nz_s7     = arr_s7[arr_s7 > 0]
        cv_s7     = float(np.std(nz_s7)/np.mean(nz_s7)) if len(nz_s7)>1 else 0.0
        rng_r_s7  = float(nz_s7.max()/nz_s7.min()) if len(nz_s7)>1 else 1.0

        if has_zeros_s7:
            rec_title = "N2 (sum-based) — recommended"
            rec_text  = ("Your data contains zero values. Min-max normalisation "
                         "produces division by zero for cost indicators. "
                         "N2 sum-based normalisation handles zeros correctly by "
                         "using reciprocal transformation with a floor value.")
            rec_col   = "green"
        elif rng_r_s7 > 100:
            rec_title = "N2 (sum-based) — recommended"
            rec_text  = (f"Your data spans a large range (max/min ratio = {rng_r_s7:.0f}×). "
                         "Min-max normalisation compresses small differences near the minimum. "
                         "N2 sum-based normalisation preserves proportional differences "
                         "between alternatives more faithfully.")
            rec_col   = "green"
        elif cv_s7 < 0.15:
            rec_title = "Any normalisation — low sensitivity expected"
            rec_text  = (f"Your indicator values show low variation (CV = {cv_s7:.2f}). "
                         "All normalisation methods will produce similar results. "
                         "N2 is the PRISM default. Run Validation Check 5 "
                         "(normalisation sensitivity) to confirm.")
            rec_col   = "blue"
        else:
            rec_title = "N2 (sum-based) — recommended"
            rec_text  = (f"Standard dataset (CV = {cv_s7:.2f}, range ratio = {rng_r_s7:.1f}×). "
                         "N2 sum-based normalisation is appropriate as the PRISM default. "
                         "Verify with Validation Check 5 (normalisation sensitivity).")
            rec_col   = "green"

        bg_map  = {"green":"#E8F5E9","blue":"#E3F0FC"}
        bd_map  = {"green":"#16A34A","blue":"#1565C0"}
        tx_map  = {"green":"#1B5E20","blue":"#0D2B5E"}
        st.markdown(
            f'<div style="background:{bg_map[rec_col]};border-left:3px solid {bd_map[rec_col]};'
            f'padding:8px 12px;border-radius:4px;margin:0 0 12px 0;'
            f'font-family:Times New Roman,Tinos,serif;font-size:13px;color:{tx_map[rec_col]};">'
            f'<b>Normalisation recommendation:</b> {rec_title}<br>'
            f'<span style="font-size:12px;">{rec_text}</span></div>',
            unsafe_allow_html=True,
        )

    for ckey in ordered_sel_cats():
        cat = CATS[ckey]
        ind_names, ind_units, _ = get_full_indicators(ckey)
        w = st.session_state.merec_w[ckey]

        st.markdown(
            f"<span style='background:{cat['bg']};color:{cat['color']};"
            f"padding:2px 10px;border-radius:12px;font-size:13px;font-weight:600;"
            f"font-family:Times New Roman,Tinos,Times,serif;'>"
            f"{cat['label']}</span>", unsafe_allow_html=True,
        )

        rows = [[ind, round(w[j], 4)] for j, ind in enumerate(ind_names)]
        st.dataframe(pd.DataFrame(rows, columns=["Indicator", "MEREC weight"]),
                     use_container_width=True, hide_index=True)
        st.write("")

    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("<- Back"):
            st.session_state.step = 6
            st.rerun()
    with c2:
        if st.button("View category scores ->", type="primary"):
            st.session_state.step = 8
            st.rerun()


def _step8_category_radar_figure(names, cat_keys, cat_scores):
    """Build a Plotly radar chart: categories as axes, alternatives as traces."""
    theta = [CATS[c]["label"] for c in cat_keys]
    theta_closed = theta + [theta[0]]
    max_r = max(
        float(cat_scores[c][pi])
        for c in cat_keys
        for pi in range(len(names))
    ) or 1.0

    fig = go.Figure()
    for pi, name in enumerate(names):
        r = [float(cat_scores[c][pi]) for c in cat_keys]
        r_closed = r + [r[0]]
        color = PROC_COLORS[pi % len(PROC_COLORS)]
        hex_c = color.lstrip("#")
        rv, gv, bv = (int(hex_c[i:i + 2], 16) for i in (0, 2, 4))
        fig.add_trace(go.Scatterpolar(
            r=r_closed,
            theta=theta_closed,
            fill="toself",
            name=name,
            line=dict(color=color, width=2),
            fillcolor=f"rgba({rv},{gv},{bv},0.15)",
        ))
    fig.update_layout(
        polar=dict(
            radialaxis=dict(
                visible=True,
                range=[0, max_r * 1.05],
                tickfont=dict(size=9, family="Times New Roman"),
                gridcolor="rgba(0,0,0,0.1)",
            ),
            angularaxis=dict(tickfont=dict(size=10, family="Times New Roman")),
        ),
        legend=dict(
            orientation="h", yanchor="bottom", y=1.05,
            font=dict(family="Times New Roman"),
        ),
        height=480,
        margin=dict(l=60, r=60, t=60, b=40),
        paper_bgcolor="white",
        font=dict(family="Times New Roman", color="#0D2B5E"),
        hoverlabel=dict(font_family="Times New Roman"),
    )
    return fig


def step8():
    st.header("Step 8 - Category scores")

    names = st.session_state.proc_names
    cat_keys = ordered_sel_cats()
    cat_scores = st.session_state.cat_scores

    if len(cat_keys) == 1:
        ckey = cat_keys[0]
        cat = CATS[ckey]
        scores = cat_scores[ckey]
        apply_mpl_style()
        shorts = proc_short_labels(names)
        fig, ax = plt.subplots(figsize=(6, max(1.2, 0.45 * len(names))))
        ax.barh(shorts, scores, color=cat["color"], edgecolor="white")
        ax.set_xlim(0, max(scores) * 1.05)
        ax.tick_params(axis="both", labelsize=9)
        plt.tight_layout()
        mpl_show(fig)
    else:
        st.plotly_chart(
            _step8_category_radar_figure(names, cat_keys, cat_scores),
            use_container_width=True,
        )

    score_rows = []
    for pi, name in enumerate(names):
        row = {"Alternative": name}
        for ckey in cat_keys:
            row[CATS[ckey]["label"]] = round(float(cat_scores[ckey][pi]), 4)
        score_rows.append(row)
    st.dataframe(pd.DataFrame(score_rows), use_container_width=True,
                 hide_index=True)

    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("<- Back"):
            st.session_state.step = 7
            st.rerun()
    with c2:
        if st.button("Select for Level 3 ->", type="primary"):
            st.session_state.step = 9
            st.rerun()


def step9():
    st.header("Step 9 - Select categories for Level 3")

    cats = ordered_sel_cats()
    cols = st.columns(len(cats))
    for i, ckey in enumerate(cats):
        cat = CATS[ckey]
        with cols[i]:
            checked = ckey in st.session_state.l3_cats
            new_val = st.checkbox(cat["label"], value=checked, key=f"l3chk_{ckey}",
                                  help=FIELD_HELP["l3_category"])
            if new_val:
                st.session_state.l3_cats.add(ckey)
            else:
                st.session_state.l3_cats.discard(ckey)

    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("<- Back"):
            st.session_state.step = 8
            st.rerun()
    with c2:
        if st.button("Level 3 ->", type="primary"):
            if not st.session_state.l3_cats:
                st.error("Select at least one category.")
            else:
                st.session_state.step = 10
                st.rerun()


def step10():
    st.header("Step 10 - Category weighting methods")

    options = {"equal": "Equal weights", "entropy": "Entropy weights", "critic": "CRITIC weights"}
    weight_help = {
        "equal": FIELD_HELP["weight_equal"],
        "entropy": FIELD_HELP["weight_entropy"],
        "critic": FIELD_HELP["weight_critic"],
    }
    for key, label in options.items():
        checked = key in st.session_state.sel_weight_methods
        new_val = st.checkbox(label, value=checked, key=f"wmchk_{key}",
                              help=weight_help[key])
        if new_val:
            st.session_state.sel_weight_methods.add(key)
        else:
            st.session_state.sel_weight_methods.discard(key)

    if st.session_state.sel_weight_methods:
        l3_cats = ordered_l3_cats()
        mat = np.array([st.session_state.cat_scores[c] for c in l3_cats])
        k = len(l3_cats)

        sets, labels = [], []
        if "equal" in st.session_state.sel_weight_methods:
            sets.append(np.full(k, 1.0 / k)); labels.append("Equal")
        if "entropy" in st.session_state.sel_weight_methods:
            sets.append(entropy_weights(mat)); labels.append("Entropy")
        if "critic" in st.session_state.sel_weight_methods:
            sets.append(critic_weights(mat)); labels.append("CRITIC")

        final_w = rcw_consolidate(sets) if len(sets) > 1 else sets[0]
        st.session_state.final_cat_weights = final_w

        rows = []
        for i, ckey in enumerate(l3_cats):
            row = [CATS[ckey]["label"]]
            for s in sets:
                row.append(round(s[i], 4))
            row.append(round(final_w[i], 4))
            rows.append(row)
        cols = ["Category"] + labels + (["RCW"] if len(sets) > 1 else ["Final"])
        st.dataframe(pd.DataFrame(rows, columns=cols), use_container_width=True, hide_index=True)

    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("<- Back"):
            st.session_state.step = 9
            st.rerun()
    with c2:
        if st.button("Select MCDM ->", type="primary"):
            if not st.session_state.sel_weight_methods:
                st.error("Select at least one weighting method.")
            else:
                st.session_state.step = 11
                st.rerun()


def step11():
    st.header("Step 11 - Select MCDM methods")


    MCDM_FAMILIES = {
        "Distance-based": {
            "description": "Evaluate alternatives by distance from ideal or reference solutions.",
            "methods": {
                "topsis": ("TOPSIS", "Technique for Order of Preference by Similarity to Ideal Solution"),
            },
        },
        "Compromise-based": {
            "description": "Seek a compromise balancing group utility and individual regret.",
            "methods": {
                "vikor": ("VIKOR", "VIšekriterijumsko KOmpromisno Rangiranje"),
            },
        },
        "Outranking": {
            "description": "Compare alternatives pairwise using concordance and discordance.",
            "methods": {
                "electre": ("ELECTRE-Score", "Outranking-based continuous scoring"),
            },
        },
        "Aggregation": {
            "description": "Combine weighted scores through mathematical aggregation functions.",
            "methods": {
                "waspas": ("WASPAS", "Weighted Aggregated Sum Product ASsessment"),
            },
        },
        "Multi-subordinate": {
            "description": "Aggregate multiple subordinate ranking models into one.",
            "methods": {
                "multimoora": ("MULTIMOORA", "Multi-Objective Optimisation by Ratio Analysis"),
            },
        },
    }

    families_selected = set()
    for family_name, family_data in MCDM_FAMILIES.items():
        with st.expander(family_name, expanded=True):
            st.caption(family_data["description"])
            for key, (short, full) in family_data["methods"].items():
                checked = key in st.session_state.sel_mcdm_methods
                new_val = st.checkbox(
                    f"{short} — {full}",
                    value=checked,
                    key=f"mmchk_{key}",
                    help=FIELD_HELP["mcdm_method"],
                )
                if new_val:
                    st.session_state.sel_mcdm_methods.add(key)
                    families_selected.add(family_name)
                else:
                    st.session_state.sel_mcdm_methods.discard(key)

    # Recount families after loop
    families_selected = set()
    for family_name, family_data in MCDM_FAMILIES.items():
        for key in family_data["methods"]:
            if key in st.session_state.sel_mcdm_methods:
                families_selected.add(family_name)

    n_selected = len(st.session_state.sel_mcdm_methods)
    n_families = len(families_selected)



    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("<- Back"):
            st.session_state.step = 10
            st.rerun()
    with c2:
        if st.button("Calculate results ->", type="primary"):
            if n_selected == 0:
                st.error("Select at least one MCDM method.")
            else:
                st.session_state.step = 12
                st.rerun()



def step12():
    st.header("Step 12 - Results")

    names = st.session_state.proc_names
    n_proc = len(names)
    l3_cats = ordered_l3_cats()
    final_w = st.session_state.final_cat_weights
    cat_scores = st.session_state.cat_scores
    methods = list(st.session_state.sel_mcdm_methods)
    multi = len(methods) > 1

    weighted_mat = np.array([cat_scores[c] * final_w[i] for i, c in enumerate(l3_cats)])

    method_ranks = run_mcdm_suite(weighted_mat, final_w, methods)
    st.session_state.last_method_ranks = method_ranks

    # Store PSI scores for use in Check 7
    if multi:
        psi_stored = calc_psi(method_ranks, methods, 0.5)
        st.session_state["last_psi_scores"] = psi_stored.tolist()

    st.subheader("MCDM rankings")

    cols = ["Alternative"] + [METHOD_LABELS[m] for m in methods]
    if multi:
        cols.append("PSI Rank (p=0.50)")
        psi_05 = calc_psi(method_ranks, methods, 0.5)
        psi_rank_05 = rank_with_ties(psi_05, ascending=False)

    rows = []
    for pi, name in enumerate(names):
        row = [name] + [int(method_ranks[m][pi]) for m in methods]
        if multi:
            row.append(int(psi_rank_05[pi]))
        rows.append(row)

    st.dataframe(pd.DataFrame(rows, columns=cols), use_container_width=True, hide_index=True)

    if multi:
        st.divider()
        st.subheader("PSI curve")
        st.markdown(
            "<p style='font-family:Times New Roman,serif;font-size:13px;"
            "color:#0D2B5E;margin:4px 0;text-align:center;'>"
            "<b>PSI</b><sub>i</sub> = "
            "<i>M</i><sub>i</sub><sup>p</sup> × "
            "<i>A</i><sub>i</sub><sup>(1−p)</sup> &nbsp;|&nbsp; "
            "<i>M</i><sub>i</sub> = 1 / <i>R̄</i><sub>i</sub> &nbsp;|&nbsp; "
            "<i>A</i><sub>i</sub> = 1 / (1 + <i>CV</i><sub>i</sub>) &nbsp;|&nbsp; "
            "<i>p</i> ∈ (0, 1)"
            "</p>",
            unsafe_allow_html=True,
        )

        p_val = st.slider("p (stability <-> performance)", min_value=0.01, max_value=0.99,
                           value=0.5, step=0.01, key="psi_p_slider",
                           help=FIELD_HELP["psi_p"])

        psi_vals = calc_psi(method_ranks, methods, p_val)
        psi_ranks = rank_with_ties(psi_vals, ascending=False)

        bar_cols = st.columns(len(names))
        for i, name in enumerate(names):
            with bar_cols[i]:
                st.metric(name, f"{psi_vals[i]:.4f}", f"rank {psi_ranks[i]}")

        p_range = np.linspace(0.01, 0.99, 99)
        apply_mpl_style()
        fig, ax = plt.subplots(figsize=(7, 3.8))
        for i, name in enumerate(names):
            series = [calc_psi(method_ranks, methods, p)[i] for p in p_range]
            ax.plot(p_range, series, label=name, linewidth=2,
                    color=PROC_COLORS[i % len(PROC_COLORS)])
        ax.axvline(x=p_val, color="gray", linestyle="--", linewidth=1)
        ax.set_xlabel("p", fontsize=10)
        ax.set_ylabel("PSI", fontsize=10)
        ax.legend(loc="upper center", bbox_to_anchor=(0.5, 1.12),
                  ncol=len(names), fontsize=9, frameon=False)
        ax.tick_params(labelsize=9)
        plt.tight_layout()
        mpl_show(fig)

        st.divider()
        st.subheader("Ranking across category combinations")
        n_combo = 2 ** len(l3_cats) - 1

        combo_p = st.slider("p for combination view", min_value=0.0, max_value=1.0,
                             value=0.5, step=0.01, key="combo_p_slider",
                             help=FIELD_HELP["combo_p"])

        combos = get_combinations(l3_cats)
        cat_initial = {c: CAT_SHORT.get(c, CATS[c]["label"][:1]) for c in l3_cats}
        combo_labels = ["+".join(cat_initial[c] for c in combo) for combo in combos]

        rank_grid = np.zeros((n_proc, len(combos)), dtype=int)
        for ci, combo in enumerate(combos):
            if len(combo) == 1:
                rank_grid[:, ci] = rank_with_ties(cat_scores[combo[0]], ascending=False)
            else:
                sub_mat = np.array([cat_scores[c] for c in combo])
                sub_w = get_category_weights(sub_mat, st.session_state.sel_weight_methods)
                sub_weighted = sub_mat * sub_w[:, None]
                sub_ranks = run_mcdm_suite(sub_weighted, sub_w, methods)
                psi_combo = calc_psi(sub_ranks, methods, combo_p)
                rank_grid[:, ci] = rank_with_ties(psi_combo, ascending=False)

        apply_mpl_style()
        n_combos_b = len(combos)
        # Jitter overlapping points slightly so all are visible
        jitter_step = 0.12
        fig_combo = go.Figure()
        for pi, name in enumerate(names):
            offset = (pi - (len(names)-1)/2) * jitter_step
            x_jit = [ci + offset for ci in range(n_combos_b)]
            hover = [
                f"<b>{name}</b><br>Combination: {combo_labels[ci]}<br>Rank: {rank_grid[pi, ci]}"
                for ci in range(n_combos_b)
            ]
            fig_combo.add_trace(go.Scatter(
                x=x_jit,
                y=rank_grid[pi, :],
                mode="markers",
                name=name,
                marker=dict(size=9, symbol="square",
                            color=PROC_COLORS[pi % len(PROC_COLORS)]),
                hovertemplate="%{customdata}<extra></extra>",
                customdata=hover,
            ))
        fig_combo.update_layout(
            xaxis=dict(
                tickmode="array",
                tickvals=list(range(n_combos_b)),
                ticktext=combo_labels,
                tickangle=-90,
                tickfont=dict(size=9, family="Times New Roman"),
                title="Category combination",
                title_font=dict(family="Times New Roman", color="#0D2B5E"),
            ),
            yaxis=dict(
                title="Rank",
                autorange="reversed",
                dtick=1, tick0=1,
                range=[0.5, n_proc + 0.5],
                gridcolor="rgba(0,0,0,0.1)",
                gridwidth=1,
                griddash="dash",
                title_font=dict(family="Times New Roman", color="#0D2B5E"),
                tickfont=dict(family="Times New Roman"),
            ),
            legend=dict(orientation="h", yanchor="bottom", y=1.02,
                        font=dict(family="Times New Roman")),
            height=420,
            margin=dict(l=10, r=10, t=40, b=120),
            plot_bgcolor="white",
            paper_bgcolor="white",
            font=dict(family="Times New Roman", color="#0D2B5E"),
            hoverlabel=dict(font_family="Times New Roman"),
        )
        st.plotly_chart(fig_combo, use_container_width=True)

        n_combos_total = len(combos)

        # Only offer the PSI-combo tie-breaker if the headline PSI ranking
        # (p = 0.50, shown in the MCDM rankings table at the top of this step)
        # actually has more than one alternative sharing rank 1.
        top_tie_count = int(np.sum(psi_rank_05 == psi_rank_05.min()))
        if top_tie_count > 1:
            st.divider()
            st.subheader("Tie-breaker: PSI-combo (recursive PSI across category combinations)")
            tied_top_names = ", ".join(names[i] for i in range(n_proc) if psi_rank_05[i] == psi_rank_05.min())
            st.markdown(
                "<p style='font-family:Times New Roman,serif;font-size:13px;"
                "color:#0D2B5E;margin:4px 0;text-align:center;'>"
                f"The headline PSI ranking (p = 0.50) is tied for rank 1 between "
                f"<b>{tied_top_names}</b>. The PSI equation can be applied a second "
                "time — using each alternative's rank across the "
                f"{n_combos_total} category combinations above as the input set, "
                "in place of the ranks from the individual MCDM methods — to break the tie."
                "</p>",
                unsafe_allow_html=True,
            )
            st.markdown(
                "<p style='font-family:Times New Roman,serif;font-size:13px;"
                "color:#0D2B5E;margin:4px 0;text-align:center;'>"
                "<b>PSI</b><sup>combo</sup><sub>i</sub> = "
                "<i>M</i><sub>i</sub><sup>p</sup> × "
                "<i>A</i><sub>i</sub><sup>(1−p)</sup> &nbsp;|&nbsp; "
                "<i>M</i><sub>i</sub> = 1 / <i>R̄</i><sup>combo</sup><sub>i</sub> &nbsp;|&nbsp; "
                "<i>A</i><sub>i</sub> = 1 / (1 + <i>CV</i><sup>combo</sup><sub>i</sub>) &nbsp;|&nbsp; "
                f"ranks drawn from all {n_combos_total} category combinations"
                "</p>",
                unsafe_allow_html=True,
            )

            psi_combo_p = st.slider("p for PSI-combo tie-breaker", min_value=0.01, max_value=0.99,
                                     value=0.5, step=0.01, key="psi_combo_p_slider",
                                     help=FIELD_HELP["psi_combo_p"])

            combo_rank_dict = {ci: rank_grid[:, ci] for ci in range(n_combos_total)}
            combo_keys_list = list(range(n_combos_total))
            psi_combo_scores = calc_psi(combo_rank_dict, combo_keys_list, psi_combo_p)
            psi_combo_ranks = rank_with_ties(psi_combo_scores, ascending=False)
            st.session_state["last_psi_combo_scores"] = psi_combo_scores.tolist()
            st.session_state["last_psi_combo_ranks"] = psi_combo_ranks.tolist()

            combo_bar_cols = st.columns(len(names))
            for i, name in enumerate(names):
                with combo_bar_cols[i]:
                    st.metric(name, f"{psi_combo_scores[i]:.4f}", f"rank {psi_combo_ranks[i]}")

            winner_idx = int(np.argmin(psi_combo_ranks))
            n_leaders = int(np.sum(psi_combo_ranks == psi_combo_ranks.min()))
            if n_leaders == 1:
                st.success(f"**Overall winner (PSI-combo tie-breaker):** {names[winner_idx]}")
            else:
                tied_names = ", ".join(names[i] for i in range(n_proc) if psi_combo_ranks[i] == psi_combo_ranks.min())
                st.warning(f"PSI-combo tie-breaker still balanced between: {tied_names}")
        else:
            st.session_state.pop("last_psi_combo_scores", None)
            st.session_state.pop("last_psi_combo_ranks", None)

    st.divider()
    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        if st.button("<- Back"):
            st.session_state.step = 11
            st.rerun()
    with c2:
        if st.button("Validation (optional) ->", type="primary"):
            st.session_state.step = 13
            st.rerun()
    with c3:
        if st.button("Analytics ->", type="primary"):
            st.session_state.step = 14
            st.rerun()
    with c4:
        if st.button("Decision support ->", type="primary"):
            st.session_state.step = 15
            st.rerun()
    with c5:
        if st.button("Reset all"):
            reset_all()
            st.rerun()


# ============================================================================
# STEP 13 - VALIDATION (optional)
# ============================================================================

def validation_rank_reversal():
    st.subheader("4. Rank-reversal test")
    
    names = st.session_state.proc_names
    n_proc = len(names)
    l3_cats = ordered_l3_cats()
    sel_weight_methods = st.session_state.sel_weight_methods
    sel_mcdm_methods = list(st.session_state.sel_mcdm_methods) or ALL_MCDM_KEYS
    method_ranks = st.session_state.get("last_method_ranks", {})

    if n_proc < 3:
        st.warning("Rank-reversal test requires at least 3 alternatives (you currently have 2).")
        return

    p_val = st.slider("p value (PSI)", min_value=0.0, max_value=1.0,
                       value=0.5, step=0.01, key="rr_p_slider",
                       help=FIELD_HELP["psi_p"])

    excluded = st.multiselect(
        "Select alternative(s) to temporarily exclude",
        options=names,
        default=[],
        key="rr_exclude",
        help=FIELD_HELP["rr_exclude"],
    )

    if not excluded:
        return

    keep_idx = [i for i, n in enumerate(names) if n not in excluded]
    keep_names = [names[i] for i in keep_idx]

    if len(keep_idx) < 2:
        st.warning("At least 2 alternatives must remain after exclusion.")
        return

    # --- Baseline ranks for the kept alternatives ---
    multi = len(sel_mcdm_methods) > 1
    base_method_ranks_kept = {}
    for m in sel_mcdm_methods:
        if m in method_ranks:
            # Re-rank among kept subset only (to get correct relative positions)
            scores_kept = np.array([method_ranks[m][i] for i in keep_idx], dtype=float)
            base_method_ranks_kept[m] = rank_with_ties(scores_kept, ascending=True)

    if multi and base_method_ranks_kept:
        base_psi_kept = calc_psi(base_method_ranks_kept, sel_mcdm_methods, p_val)
        base_psi_rank_kept = rank_with_ties(base_psi_kept, ascending=False)

    # --- Perturbed pipeline: re-run full PRISM on kept alternatives only ---
    # Rebuild category scores from raw indicator values, restricted to kept_idx
    cat_scores_pert = {}
    for ckey in l3_cats:
        ind_names, ind_units, benefits = get_full_indicators(ckey)
        n_ind = len(ind_names)
        raw_full = np.zeros((n_ind, n_proc))
        for j in range(n_ind):
            for pi in range(n_proc):
                raw_full[j, pi] = st.session_state.indicator_values.get((ckey, j, pi), 0.0)
        raw_kept = raw_full[:, keep_idx]
        # MEREC norm + weights on reduced set
        nm = np.zeros_like(raw_kept)
        n2 = np.zeros_like(raw_kept)
        for j in range(n_ind):
            nm[j] = merec_norm(raw_kept[j], benefits[j])
            n2[j] = n2_norm(raw_kept[j], benefits[j])
        w_ind = merec_weights(nm)
        cat_scores_pert[ckey] = (n2 * w_ind[:, None]).sum(axis=0)

    mat_pert = np.array([cat_scores_pert[c] for c in l3_cats])
    w_cat_pert = get_category_weights(mat_pert, sel_weight_methods)
    weighted_mat_pert = mat_pert * w_cat_pert[:, None]
    ranks_pert = run_mcdm_suite(weighted_mat_pert, w_cat_pert, sel_mcdm_methods)

    if multi:
        psi_pert = calc_psi(ranks_pert, sel_mcdm_methods, p_val)
        psi_rank_pert = rank_with_ties(psi_pert, ascending=False)

    # --- Build comparison table ---
    cols_table = ["Alternative"]
    for m in sel_mcdm_methods:
        cols_table += [f"{METHOD_LABELS[m]} (baseline)", f"{METHOD_LABELS[m]} (reduced)"]
    if multi:
        cols_table += [f"PSI Rank (baseline, p={p_val:.2f})", f"PSI Rank (reduced, p={p_val:.2f})"]

    rows_table = []
    reversal_found = False
    for idx_out, i in enumerate(keep_idx):
        row = [keep_names[idx_out]]
        method_reversal = False
        for m in sel_mcdm_methods:
            base_r = int(base_method_ranks_kept[m][idx_out]) if m in base_method_ranks_kept else "—"
            pert_r = int(ranks_pert[m][idx_out])
            row += [base_r, pert_r]
            if base_r != pert_r:
                method_reversal = True
        if multi:
            base_psi_r = int(base_psi_rank_kept[idx_out]) if base_method_ranks_kept else "—"
            pert_psi_r = int(psi_rank_pert[idx_out])
            row += [base_psi_r, pert_psi_r]
            if base_psi_r != pert_psi_r:
                method_reversal = True
        if method_reversal:
            reversal_found = True
        rows_table.append(row)

    df_compare = pd.DataFrame(rows_table, columns=cols_table)
    st.session_state["export_rank_reversal"] = {
        "excluded": excluded, "df": df_compare
    }

    st.markdown(
        f"**Excluded:** {', '.join(excluded)}  |  "
        f"**Remaining:** {', '.join(keep_names)}  |  p = {p_val:.2f}"
    )
    st.dataframe(df_compare, use_container_width=True, hide_index=True)

    if reversal_found:
        st.markdown(
            '<div style="background:#FFF8E1;border-left:3px solid #D97706;'
            'padding:8px 12px;border-radius:4px;margin:6px 0;'
            'font-family:Times New Roman,Tinos,Times,serif;color:#7B5800;font-size:13px;">'
            '⚠️ Rank-reversal detected: one or more alternatives changed position after exclusion.'
            '</div>', unsafe_allow_html=True)

        st.markdown("**Why did this happen? — RCW weight changes upon exclusion**")
        st.markdown(
            "<p style='font-family:Times New Roman,Tinos,serif;font-size:12px;color:#555;'>"
            "PRISM computes MEREC and RCW weights from the performance data of "
            "<b>all alternatives simultaneously</b>. When an alternative is removed, "
            "the normalised indicator values change → MEREC weights change → "
            "category scores change → Entropy and CRITIC weights change → "
            "RCW consolidated weights change → rankings change. "
            "This is an inherent property of data-driven objective weighting — "
            "not a framework deficiency.</p>",
            unsafe_allow_html=True,
        )

        # Show RCW weight change table
        w_full = st.session_state.final_cat_weights
        l3_cats_rr = ordered_l3_cats()
        rows_w = []
        for ci, ckey in enumerate(l3_cats_rr):
            rows_w.append({
                "Category": CATS[ckey]["label"],
                "RCW weight (all alternatives)": round(float(w_full[ci]), 4),
                "RCW weight (after exclusion)":  round(float(w_cat_pert[ci]), 4),
                "Δ Change": f"{float(w_cat_pert[ci]-w_full[ci]):+.4f}",
            })
        st.dataframe(pd.DataFrame(rows_w), use_container_width=True, hide_index=True)
        st.caption(
            "Categories with the largest Δ Change drove the rank reversal. "
            "This is expected behaviour of objective weighting — weights are "
            "recalculated from the remaining alternatives' data."
        )
    else:
        st.success(
            "✅ No rank-reversal detected: the relative ranking of the remaining alternatives "
            "is identical to their baseline positions. The result is robust to the exclusion "
            f"of {', '.join(excluded)}."
        )

    



def validation_normalisation_sensitivity():
    st.subheader("5. Normalisation sensitivity")
    
    names = st.session_state.proc_names
    l3_cats = ordered_l3_cats()
    sel_weight_methods = st.session_state.sel_weight_methods
    sel_mcdm_methods = list(st.session_state.sel_mcdm_methods) or ALL_MCDM_KEYS
    method_ranks = st.session_state.get("last_method_ranks", {})
    multi = len(sel_mcdm_methods) > 1

    # --- Alternative normalisation functions ---
    def minmax_norm(vals, benefit):
        vals = np.asarray(vals, dtype=float)
        mn, mx = vals.min(), vals.max()
        rng = (mx - mn) or 1.0
        if benefit:
            return (vals - mn) / rng
        else:
            return (mx - vals) / rng

    def vector_norm(vals, benefit):
        vals = np.asarray(vals, dtype=float)
        denom = np.sqrt(np.sum(vals ** 2)) or 1.0
        normed = vals / denom
        if benefit:
            return normed
        else:
            mx = normed.max() or 1.0
            return mx - normed

    NORM_OPTIONS = {
        "Min-max": minmax_norm,
        "Vector": vector_norm,
        "N2 (baseline)": n2_norm,
    }

    alt_norm_label = st.selectbox(
        "Alternative normalisation method",
        [k for k in NORM_OPTIONS if k != "N2 (baseline)"],
        key="norm_sens_select",
        help=FIELD_HELP["alt_norm"],
    )
    p_val = st.slider("p value (PSI)", min_value=0.0, max_value=1.0,
                       value=0.5, step=0.01, key="norm_sens_p_slider",
                       help=FIELD_HELP["psi_p"])

    alt_norm_func = NORM_OPTIONS[alt_norm_label]

    # --- Baseline ranks (from session state) ---
    base_psi_rank = None
    if multi and method_ranks:
        base_psi = calc_psi(method_ranks, sel_mcdm_methods, p_val)
        base_psi_rank = rank_with_ties(base_psi, ascending=False)

    # --- Alternative pipeline ---
    cat_scores_alt = {}
    for ckey in l3_cats:
        ind_names, ind_units, benefits = get_full_indicators(ckey)
        n_ind = len(ind_names)
        raw, _ = get_raw_matrix(ckey)
        # MEREC-norm + weights unchanged
        nm = np.zeros_like(raw)
        for j in range(n_ind):
            nm[j] = merec_norm(raw[j], benefits[j])
        w_ind = merec_weights(nm)
        # Swap N2 for chosen normalisation
        alt_n = np.zeros_like(raw)
        for j in range(n_ind):
            alt_n[j] = alt_norm_func(raw[j], benefits[j])
        cat_scores_alt[ckey] = (alt_n * w_ind[:, None]).sum(axis=0)

    mat_alt = np.array([cat_scores_alt[c] for c in l3_cats])
    w_cat_alt = get_category_weights(mat_alt, sel_weight_methods)
    weighted_mat_alt = mat_alt * w_cat_alt[:, None]
    ranks_alt = run_mcdm_suite(weighted_mat_alt, w_cat_alt, sel_mcdm_methods)

    alt_psi_rank = None
    if multi:
        psi_alt = calc_psi(ranks_alt, sel_mcdm_methods, p_val)
        alt_psi_rank = rank_with_ties(psi_alt, ascending=False)

    # --- Comparison table ---
    cols_table = ["Alternative"]
    for m in sel_mcdm_methods:
        cols_table += [f"{METHOD_LABELS[m]} (N2 baseline)", f"{METHOD_LABELS[m]} ({alt_norm_label})"]
    if multi:
        cols_table += [f"PSI Rank (N2, p={p_val:.2f})", f"PSI Rank ({alt_norm_label}, p={p_val:.2f})"]

    rows_table = []
    reversal_found = False
    for pi, name in enumerate(names):
        row = [name]
        for m in sel_mcdm_methods:
            base_r = int(method_ranks[m][pi]) if m in method_ranks else "—"
            alt_r  = int(ranks_alt[m][pi])
            row += [base_r, alt_r]
            if base_r != "—" and base_r != alt_r:
                reversal_found = True
        if multi:
            base_psi_r = int(base_psi_rank[pi]) if base_psi_rank is not None else "—"
            alt_psi_r  = int(alt_psi_rank[pi]) if alt_psi_rank is not None else "—"
            row += [base_psi_r, alt_psi_r]
            if base_psi_r != "—" and base_psi_r != alt_psi_r:
                reversal_found = True
        rows_table.append(row)

    st.markdown(
        f"**Baseline normalisation:** N2  |  "
        f"**Alternative:** {alt_norm_label}  |  p = {p_val:.2f}"
    )
    df_norm_export = pd.DataFrame(rows_table, columns=cols_table)
    st.session_state["export_norm_sens"] = {"alt_norm": alt_norm_label, "df": df_norm_export}
    st.dataframe(df_norm_export, use_container_width=True, hide_index=True)

    if reversal_found:
        st.warning(
            f"⚠️ Rank change detected: switching from N2 to {alt_norm_label} normalisation "
            "changes one or more rankings. The result has some sensitivity to normalisation choice."
        )
    else:
        st.markdown(f'<div style="background:#E8F5E9;border-left:3px solid #16A34A;padding:8px 12px;border-radius:4px;margin:6px 0;font-family:Times New Roman,Tinos,Times,serif;color:#1B5E20;font-size:13px;">No rank change under {alt_norm_label} normalisation. The result is robust to normalisation method choice.</div>', unsafe_allow_html=True)

    


def validation_bootstrap_merec_rcw():
    st.subheader("6. Bootstrap stability analysis — MEREC and RCW")

    names    = st.session_state.proc_names
    l3_cats  = ordered_l3_cats()
    final_w  = st.session_state.final_cat_weights
    sel_wm   = st.session_state.sel_weight_methods

    c1, c2 = st.columns(2)
    with c1:
        unc_pct = st.slider(
            "Indicator value uncertainty (±%)",
            min_value=1, max_value=50, value=10, step=1,
            key="bs_unc_pct",
            help=FIELD_HELP["unc_pct"],
        )
    with c2:
        n_iter = st.slider(
            "Number of bootstrap iterations",
            min_value=100, max_value=2000, value=500, step=100,
            key="bs_n_iter",
            help=FIELD_HELP["n_iter"],
        )

    if not st.button("Run bootstrap", type="primary", key="bs_run"):
        st.info("Set parameters above and click Run bootstrap.")
        return

    st.session_state.dash_bs_n_iter = int(n_iter)
    st.session_state.dash_bs_unc_pct = int(unc_pct)
    st.session_state.pop("validation_dashboard_results", None)

    rng = np.random.default_rng(42)
    progress = st.progress(0, text="Running bootstrap...")

    # Storage
    # MEREC: per category → list of weight arrays (n_iter, n_ind)
    merec_boot = {ckey: [] for ckey in l3_cats}
    # RCW: list of weight arrays (n_iter, n_cats)
    rcw_boot = []

    for b in range(n_iter):
        # ── Perturb indicator values ──────────────────────────────────────────
        cat_scores_b = {}
        merec_w_b = {}

        for ckey in l3_cats:
            raw, benefits = get_raw_matrix(ckey)
            n_ind, n_proc = raw.shape
            # Sample each indicator value within ±unc_pct%
            noise = rng.uniform(
                1 - unc_pct/100,
                1 + unc_pct/100,
                size=raw.shape
            )
            raw_b = np.maximum(raw * noise, 1e-9)

            # Recompute MEREC and N2
            nm_b = np.zeros_like(raw_b)
            n2_b = np.zeros_like(raw_b)
            for j in range(n_ind):
                nm_b[j] = merec_norm(raw_b[j], benefits[j])
                n2_b[j] = n2_norm(raw_b[j], benefits[j])

            w_ind_b = merec_weights(nm_b)
            merec_w_b[ckey] = w_ind_b
            merec_boot[ckey].append(w_ind_b.copy())

            cat_scores_b[ckey] = (n2_b * w_ind_b[:, None]).sum(axis=0)

        # ── Recompute RCW ─────────────────────────────────────────────────────
        mat_b = np.array([cat_scores_b[c] for c in l3_cats])
        w_rcw_b = get_category_weights(mat_b, sel_wm)
        rcw_boot.append(w_rcw_b.copy())

        if (b + 1) % 50 == 0:
            progress.progress((b+1)/n_iter,
                               text=f"Bootstrap iteration {b+1}/{n_iter}...")

    progress.empty()

    # ── MEREC results ─────────────────────────────────────────────────────────
    st.markdown("### MEREC Weight Stability")
    st.markdown(
        f"<p style='font-family:Times New Roman,Tinos,serif;font-size:12px;"
        f"color:#555;'>±{unc_pct}% indicator uncertainty | "
        f"{n_iter} iterations | 95% CI shown</p>",
        unsafe_allow_html=True,
    )

    for ckey in l3_cats:
        cat = CATS[ckey]
        ind_names, ind_units, _ = get_full_indicators(ckey)
        boot_arr = np.array(merec_boot[ckey])  # (n_iter, n_ind)
        observed = st.session_state.merec_w[ckey]

        st.markdown(
            f"<span style='background:{cat['bg']};color:{cat['color']};"
            f"padding:2px 10px;border-radius:12px;font-size:13px;"
            f"font-weight:600;font-family:Times New Roman,Tinos,serif;'>"
            f"{cat['label']}</span>",
            unsafe_allow_html=True,
        )

        rows_m = []
        all_stable = True
        for j, (iname, iunit) in enumerate(zip(ind_names, ind_units)):
            obs  = float(observed[j])
            lo   = float(np.percentile(boot_arr[:, j], 2.5))
            hi   = float(np.percentile(boot_arr[:, j], 97.5))
            mean = float(boot_arr[:, j].mean())
            cv   = float(boot_arr[:, j].std(ddof=0) / mean * 100) if mean > 0 else 0.0
            # Stable if observed rank matches bootstrap mean rank
            obs_rank   = int(np.argsort(-observed)[j] + 1) if len(observed) > 1 else 1
            stable = lo <= obs <= hi
            if not stable:
                all_stable = False
            rows_m.append({
                "Indicator": f"{iname} ({iunit})",
                "Observed weight": round(obs, 4),
                "Bootstrap mean": round(mean, 4),
                "95% CI lower": round(lo, 4),
                "95% CI upper": round(hi, 4),
                "CV (%)": round(cv, 1),
                "Stable?": "✅ Yes" if stable else "⚠️ No",
            })

        st.dataframe(pd.DataFrame(rows_m), use_container_width=True,
                     hide_index=True)

        # Matplotlib CI chart
        apply_mpl_style()
        n_ind = len(ind_names)
        short_names = [f"{ind_names[j][:12]}" for j in range(n_ind)]
        obs_vals = [float(observed[j]) for j in range(n_ind)]
        lo_vals  = [float(np.percentile(boot_arr[:, j], 2.5)) for j in range(n_ind)]
        hi_vals  = [float(np.percentile(boot_arr[:, j], 97.5)) for j in range(n_ind)]
        means    = [float(boot_arr[:, j].mean()) for j in range(n_ind)]

        fig, ax = plt.subplots(figsize=(max(5, n_ind * 1.2), 2.8))
        x = np.arange(n_ind)
        # CI bars
        ax.bar(x, [h - l for h, l in zip(hi_vals, lo_vals)],
               bottom=lo_vals, color=cat["color"], alpha=0.3,
               width=0.5, label="95% CI", edgecolor="none")
        # Observed
        ax.scatter(x, obs_vals, color=cat["color"], zorder=5,
                   s=60, label="Observed", marker="D")
        # Bootstrap mean
        ax.scatter(x, means, color="#0D2B5E", zorder=4,
                   s=30, label="Bootstrap mean", marker="o")
        ax.set_xticks(x)
        ax.set_xticklabels(short_names, rotation=20, ha="right", fontsize=8)
        ax.set_ylabel("MEREC weight", fontsize=9)
        ax.legend(fontsize=8, frameon=False,
                  loc="upper center", bbox_to_anchor=(0.5, 1.18), ncol=3)
        ax.tick_params(labelsize=8)
        plt.tight_layout()
        mpl_show(fig)

        if all_stable:
            st.markdown(
                f'<div style="background:#E8F5E9;border-left:3px solid #16A34A;'
                f'padding:8px 12px;border-radius:4px;margin:4px 0;'
                f'font-family:Times New Roman,Tinos,serif;font-size:13px;color:#1B5E20;">'
                f'✅ All MEREC weights for {cat["label"]} are stable within ±{unc_pct}% '
                f'indicator uncertainty — observed weights lie within the 95% CI.</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f'<div style="background:#FFF8E1;border-left:3px solid #D97706;'
                f'padding:8px 12px;border-radius:4px;margin:4px 0;'
                f'font-family:Times New Roman,Tinos,serif;font-size:13px;color:#7B5800;">'
                f'⚠️ Some MEREC weights for {cat["label"]} fall outside the 95% CI '
                f'under ±{unc_pct}% uncertainty — consider reviewing indicator data quality.</div>',
                unsafe_allow_html=True,
            )

    # ── RCW results ───────────────────────────────────────────────────────────
    st.markdown("### RCW Weight Stability")

    rcw_arr = np.array(rcw_boot)  # (n_iter, n_cats)
    cat_labels = [CATS[c]["label"] for c in l3_cats]

    rows_r = []
    all_stable_rcw = True
    for ci, ckey in enumerate(l3_cats):
        obs  = float(final_w[ci])
        lo   = float(np.percentile(rcw_arr[:, ci], 2.5))
        hi   = float(np.percentile(rcw_arr[:, ci], 97.5))
        mean = float(rcw_arr[:, ci].mean())
        cv   = float(rcw_arr[:, ci].std(ddof=0) / mean * 100) if mean > 0 else 0.0
        stable = lo <= obs <= hi
        if not stable:
            all_stable_rcw = False
        rows_r.append({
            "Category": CATS[ckey]["label"],
            "Observed RCW weight": round(obs, 4),
            "Bootstrap mean": round(mean, 4),
            "95% CI lower": round(lo, 4),
            "95% CI upper": round(hi, 4),
            "CV (%)": round(cv, 1),
            "Stable?": "✅ Yes" if stable else "⚠️ No",
        })

    st.dataframe(pd.DataFrame(rows_r), use_container_width=True,
                 hide_index=True)

    # RCW CI chart
    apply_mpl_style()
    n_cats = len(l3_cats)
    obs_r  = [float(final_w[ci]) for ci in range(n_cats)]
    lo_r   = [float(np.percentile(rcw_arr[:, ci], 2.5)) for ci in range(n_cats)]
    hi_r   = [float(np.percentile(rcw_arr[:, ci], 97.5)) for ci in range(n_cats)]
    mn_r   = [float(rcw_arr[:, ci].mean()) for ci in range(n_cats)]
    colors = [CATS[c]["color"] for c in l3_cats]

    fig, ax = plt.subplots(figsize=(max(5, n_cats * 1.4), 3.2))
    x = np.arange(n_cats)
    for ci in range(n_cats):
        ax.bar(ci, hi_r[ci] - lo_r[ci], bottom=lo_r[ci],
               color=colors[ci], alpha=0.3, width=0.5, edgecolor="none")
        ax.scatter(ci, obs_r[ci], color=colors[ci], zorder=5,
                   s=70, marker="D")
        ax.scatter(ci, mn_r[ci], color="#0D2B5E", zorder=4,
                   s=35, marker="o")
    ax.set_xticks(x)
    ax.set_xticklabels([CATS[c]["label"][:6] for c in l3_cats],
                       rotation=20, ha="right", fontsize=9)
    ax.set_ylabel("RCW weight", fontsize=10)
    from matplotlib.lines import Line2D
    legend_elements = [
        Line2D([0],[0], marker="D", color="grey", markersize=7,
               linestyle="none", label="Observed"),
        Line2D([0],[0], marker="o", color="#0D2B5E", markersize=5,
               linestyle="none", label="Bootstrap mean"),
        plt.Rectangle((0,0),1,1, facecolor="grey", alpha=0.3, label="95% CI"),
    ]
    ax.legend(handles=legend_elements, fontsize=8, frameon=False,
              loc="upper center", bbox_to_anchor=(0.5, 1.14), ncol=3)
    ax.tick_params(labelsize=9)
    plt.tight_layout()
    mpl_show(fig)

    # Weight order stability
    st.markdown("**Weight ordering stability**")
    obs_order   = [cat_labels[i] for i in np.argsort(-np.array(obs_r))]
    order_stable = 0
    for b in range(n_iter):
        boot_order = [cat_labels[i]
                      for i in np.argsort(-rcw_arr[b])]
        if boot_order == obs_order:
            order_stable += 1
    pct_stable = order_stable / n_iter * 100

    if all_stable_rcw:
        st.markdown(
            f'<div style="background:#E8F5E9;border-left:3px solid #16A34A;'
            f'padding:8px 12px;border-radius:4px;margin:4px 0;'
            f'font-family:Times New Roman,Tinos,serif;font-size:13px;color:#1B5E20;">'
            f'✅ All RCW weights are stable within ±{unc_pct}% indicator uncertainty. '
            f'The observed weight ordering ({" > ".join(obs_order)}) is preserved in '
            f'{pct_stable:.1f}% of bootstrap iterations.</div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f'<div style="background:#FFF8E1;border-left:3px solid #D97706;'
            f'padding:8px 12px;border-radius:4px;margin:4px 0;'
            f'font-family:Times New Roman,Tinos,serif;font-size:13px;color:#7B5800;">'
            f'⚠️ Some RCW weights fall outside the 95% CI. '
            f'The observed weight ordering is preserved in '
            f'{pct_stable:.1f}% of bootstrap iterations.</div>',
            unsafe_allow_html=True,
        )

    # Store for export
    st.session_state["export_bootstrap"] = {
        "unc_pct": unc_pct, "n_iter": n_iter,
        "merec_boot": merec_boot, "rcw_boot": rcw_arr,
        "pct_stable": pct_stable,
    }


def _indicator_uncertainty_psi(l3_cats, sel_wm, sel_mcdm, multi, p_psi, rng=None,
                               unc_pct=0, perturb=False):
    """Run one PRISM pipeline pass and return PSI vector."""
    cat_scores_b = {}
    for ckey in l3_cats:
        raw, benefits = get_raw_matrix(ckey)
        if perturb and unc_pct > 0:
            noise = rng.uniform(
                1 - unc_pct / 100, 1 + unc_pct / 100, size=raw.shape,
            )
            raw_b = np.maximum(raw * noise, 1e-9)
        else:
            raw_b = np.maximum(raw, 1e-9)

        nm = np.zeros_like(raw_b)
        n2 = np.zeros_like(raw_b)
        for j in range(raw_b.shape[0]):
            nm[j] = merec_norm(raw_b[j], benefits[j])
            n2[j] = n2_norm(raw_b[j], benefits[j])
        w_ind = merec_weights(nm)
        cat_scores_b[ckey] = (n2 * w_ind[:, None]).sum(axis=0)

    mat_b = np.array([cat_scores_b[c] for c in l3_cats])
    w_cat_b = get_category_weights(mat_b, sel_wm)
    wm_b = mat_b * w_cat_b[:, None]
    ranks_b = run_mcdm_suite(wm_b, w_cat_b, sel_mcdm)

    if multi:
        psi_b = calc_psi(ranks_b, sel_mcdm, p_psi)
    else:
        psi_b = 1.0 / ranks_b[sel_mcdm[0]].astype(float)

    return np.clip(psi_b, 0.0, 1.0), ranks_b


def _psi_percentile_ci(samples):
    """Return (lo, hi, mean) from finite bootstrap PSI samples."""
    valid = samples[np.isfinite(samples)]
    if valid.size == 0:
        return 0.0, 0.0, 0.0
    lo = float(np.percentile(valid, 2.5))
    hi = float(np.percentile(valid, 97.5))
    mn = float(valid.mean())
    return lo, hi, mn


def validation_indicator_uncertainty():
    """
    Validation Check 7 — Indicator-level uncertainty propagation.

    Propagates ±% uncertainty in raw indicator values through the full
    PRISM pipeline (MEREC → N2 → CategoryScores → RCW → MCDM → PSI)
    using Monte Carlo sampling.

    Addresses Limitation L3: uncertainty in raw indicator values is not
    captured by weight-level Monte Carlo (Check 3) alone.
    """
    st.subheader("7. Indicator-level uncertainty propagation")

    names    = st.session_state.proc_names
    n_proc   = len(names)
    l3_cats  = ordered_l3_cats()
    sel_wm   = st.session_state.sel_weight_methods
    sel_mcdm = list(st.session_state.sel_mcdm_methods) or ALL_MCDM_KEYS
    multi    = len(sel_mcdm) > 1

    c1, c2 = st.columns(2)
    with c1:
        unc_pct = st.slider(
            "Indicator value uncertainty (±%)",
            min_value=1, max_value=50, value=10, step=1,
            key="ind_unc_pct",
            help=FIELD_HELP["unc_pct"],
        )
    with c2:
        n_iter = st.slider(
            "Monte Carlo iterations",
            min_value=100, max_value=10000, value=500, step=100,
            key="ind_unc_iter",
            help=FIELD_HELP["n_iter"],
        )

    p_psi = st.slider(
        "p value (PSI)", min_value=0.0, max_value=1.0,
        value=0.5, step=0.01, key="ind_unc_p",
        help=FIELD_HELP["psi_p"],
    )

    if not st.button("Run uncertainty propagation", type="primary",
                     key="ind_unc_run"):
        st.info(
            "This check propagates measurement uncertainty in indicator values "
            "through the full PRISM pipeline. Set ±% uncertainty and click Run."
        )
        return

    st.session_state.dash_ind_unc_iter = int(n_iter)
    st.session_state.dash_ind_unc_pct = int(unc_pct)
    st.session_state.pop("validation_dashboard_results", None)

    rng      = np.random.default_rng(42)
    progress = st.progress(0, text="Running indicator uncertainty propagation...")

    obs_psi, _ = _indicator_uncertainty_psi(
        l3_cats, sel_wm, sel_mcdm, multi, p_psi, perturb=False,
    )

    psi_boot     = np.full((n_iter, n_proc), np.nan)
    rank_counts  = np.zeros((n_proc, n_proc), dtype=int)
    invalid_iters = 0

    for b in range(n_iter):
        psi_b, ranks_b = _indicator_uncertainty_psi(
            l3_cats, sel_wm, sel_mcdm, multi, p_psi,
            rng=rng, unc_pct=unc_pct, perturb=True,
        )
        if not np.all(np.isfinite(psi_b)):
            invalid_iters += 1
            continue

        psi_boot[b] = psi_b
        psi_ranks_b = rank_with_ties(psi_b, ascending=False)
        for pi in range(n_proc):
            r = int(psi_ranks_b[pi]) - 1
            if 0 <= r < n_proc:
                rank_counts[pi, r] += 1

        if (b + 1) % 50 == 0:
            progress.progress(
                (b + 1) / n_iter,
                text=f"Iteration {b + 1}/{n_iter}...",
            )

    progress.empty()

    valid_iters = int(np.isfinite(psi_boot).all(axis=1).sum())
    if invalid_iters:
        st.caption(
            f"Note: {invalid_iters} iteration(s) excluded due to non-finite PSI."
        )
    if valid_iters == 0:
        st.error("Uncertainty propagation failed — no valid iterations.")
        return

    invariant_alts = []

    # ── PSI CI table ─────────────────────────────────────────────────────────
    st.markdown("**PSI score confidence intervals under ±" + str(unc_pct) + "% indicator uncertainty**")
    rows_ci = []
    ci_stats = []
    x_vals = []
    for pi, name in enumerate(names):
        col = psi_boot[:, pi]
        col = col[np.isfinite(col)]
        lo, hi, mn = _psi_percentile_ci(col)
        obs = float(obs_psi[pi])
        if col.size > 1 and float(col.std(ddof=0)) < 1e-9:
            invariant_alts.append(name)
        rows_ci.append({
            "Alternative":    name,
            "Observed PSI":   round(obs, 4),
            "Bootstrap mean": round(mn, 4),
            "95% CI lower":   round(lo, 4),
            "95% CI upper":   round(hi, 4),
            "CI width":       round(hi - lo, 4),
        })
        ci_stats.append((lo, hi, mn, obs))
        x_vals.extend([lo, hi, mn, obs])
    st.dataframe(pd.DataFrame(rows_ci),
                 use_container_width=True, hide_index=True)
    if invariant_alts:
        st.caption(
            "Zero-width CI for "
            + ", ".join(invariant_alts)
            + ": MCDM ranks (hence PSI) did not change across any valid iteration "
            f"under ±{unc_pct}% indicator uncertainty."
        )

    # ── Rank probability table ────────────────────────────────────────────────
    st.markdown("**Rank probability (% of iterations each alternative held each rank)**")
    rows_rp = []
    for pi, name in enumerate(names):
        row = {"Alternative": name}
        for r in range(n_proc):
            row[f"Rank {r+1} (%)"] = round(rank_counts[pi, r] / valid_iters * 100, 1)
        rows_rp.append(row)
    st.dataframe(pd.DataFrame(rows_rp),
                 use_container_width=True, hide_index=True)

    # ── PSI distribution chart ────────────────────────────────────────────────
    st.markdown("**PSI score distributions — 95% CI across iterations**")
    apply_mpl_style()
    shorts = proc_short_labels(names)
    fig, ax = plt.subplots(
        figsize=(max(6, n_proc * 2.0), max(3.5, n_proc * 1.1)),
    )

    for pi, (lo, hi, mn, obs) in enumerate(ci_stats):
        ax.barh(
            pi, hi - lo, left=lo, height=0.5,
            color="#0D2B5E", alpha=0.3,
            label="95% CI" if pi == 0 else "",
        )
        ax.scatter(
            mn, pi, color="#0D2B5E", s=55, zorder=5,
            marker="D", label="Bootstrap mean" if pi == 0 else "",
        )
        ax.scatter(
            obs, pi, color="#E85C1A", s=55, zorder=6,
            marker="o", label="Observed PSI" if pi == 0 else "",
        )

    xmin = max(0.0, min(x_vals) - 0.05)
    xmax = min(1.05, max(x_vals) + 0.05)
    if xmax - xmin < 0.05:
        xmax = xmin + 0.05
    ax.set_xlim(xmin, xmax)
    ax.set_yticks(range(n_proc))
    ax.set_yticklabels(shorts, fontsize=9)
    ax.set_xlabel("PSI score", fontsize=10)
    ax.legend(fontsize=8, frameon=False, loc="lower right")
    ax.tick_params(labelsize=9)
    plt.tight_layout()
    mpl_show(fig)

    # ── Summary message ───────────────────────────────────────────────────────
    top_pi  = int(np.argmax([rank_counts[p, 0] for p in range(n_proc)]))
    pct_top = rank_counts[top_pi, 0] / valid_iters * 100

    if pct_top >= 90:
        st.markdown(
            f'<div style="background:#E8F5E9;border-left:3px solid #16A34A;'
            f'padding:8px 12px;border-radius:4px;margin:6px 0;'
            f'font-family:Times New Roman,Tinos,serif;font-size:13px;color:#1B5E20;">'
            f'✅ <b>{names[top_pi]}</b> holds Rank 1 in <b>{pct_top:.1f}%</b> of '
            f'{valid_iters} valid iterations under ±{unc_pct}% indicator uncertainty. '
            f'The ranking is robust to realistic data uncertainty.</div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f'<div style="background:#FFF8E1;border-left:3px solid #D97706;'
            f'padding:8px 12px;border-radius:4px;margin:6px 0;'
            f'font-family:Times New Roman,Tinos,serif;font-size:13px;color:#7B5800;">'
            f'⚠️ <b>{names[top_pi]}</b> holds Rank 1 in only <b>{pct_top:.1f}%</b> of '
            f'{valid_iters} valid iterations under ±{unc_pct}% indicator uncertainty. '
            f'The ranking is sensitive to data uncertainty — improve data quality '
            f'for indicators with the widest CI ranges.</div>',
            unsafe_allow_html=True,
        )



TRAFFIC_LABELS = {
    "green": "Robust",
    "amber": "Partially sensitive",
    "red": "Sensitive",
    "na": "Not applicable",
}


def _status_from_rank1_pct(pct):
    if pct >= 70:
        return "green"
    if pct >= 40:
        return "amber"
    return "red"


def _winner_headline_rank(method_ranks, methods, winner_idx, p=0.5):
    if len(methods) > 1:
        psi = calc_psi(method_ranks, methods, p)
        return int(rank_with_ties(psi, ascending=False)[winner_idx])
    return int(method_ranks[methods[0]][winner_idx])


def _assess_weight_sensitivity(winner_idx, l3_cats, cat_scores, methods):
    mat = np.array([cat_scores[c] for c in l3_cats])
    winner_ranks = []
    for combo_methods, _ in WEIGHT_COMBO_SETS:
        w = get_category_weights(mat, set(combo_methods))
        weighted_mat = mat * w[:, None]
        ranks = run_mcdm_suite(weighted_mat, w, methods)
        if len(methods) > 1:
            psi = calc_psi(ranks, methods, 0.5)
            winner_ranks.append(int(rank_with_ties(psi, ascending=False)[winner_idx]))
        else:
            winner_ranks.append(int(ranks[methods[0]][winner_idx]))
    if all(r == 1 for r in winner_ranks):
        status, detail = "green", "Winner Rank 1 across all category weighting schemes"
    elif max(winner_ranks) <= 2:
        status = "amber"
        detail = f"Winner rank varies (ranks {min(winner_ranks)}–{max(winner_ranks)}) but remains competitive"
    else:
        status = "red"
        detail = f"Winner loses top rank under some weighting schemes (worst rank {max(winner_ranks)})"
    return {
        "id": "weighting", "name": "Weighting-method sensitivity",
        "status": status, "detail": detail, "status_label": TRAFFIC_LABELS[status],
    }


def _assess_bc_sensitivity(winner_idx, l3_cats, methods, method_ranks, sel_wm):
    names = st.session_state.proc_names
    multi = len(methods) > 1
    base_rank = _winner_headline_rank(method_ranks, methods, winner_idx, 0.5)

    rng_bc = np.random.default_rng(42)
    ben_pct, cost_pct = 10, 10
    cat_scores_pert = {}
    for ckey in l3_cats:
        raw, benefits = get_raw_matrix(ckey)
        raw_pert = raw.copy()
        for j in range(raw.shape[0]):
            pct = ben_pct if benefits[j] else cost_pct
            half = abs(pct) * 0.5
            noise = rng_bc.uniform(-half, half, size=raw.shape[1])
            factors = 1 + (pct + noise) / 100.0
            raw_pert[j] = np.maximum(raw[j] * factors, 1e-9)
        cat_scores_pert[ckey] = compute_category_score_from_raw(ckey, raw_pert)

    mat = np.array([cat_scores_pert[c] for c in l3_cats])
    w = get_category_weights(mat, sel_wm)
    weighted_mat = mat * w[:, None]
    ranks = run_mcdm_suite(weighted_mat, w, methods)
    if multi:
        pert_rank = int(rank_with_ties(calc_psi(ranks, methods, 0.5), ascending=False)[winner_idx])
    else:
        pert_rank = int(ranks[methods[0]][winner_idx])

    if pert_rank == base_rank:
        status, detail = "green", f"No winner rank change under ±{ben_pct}% benefit / ±{cost_pct}% cost perturbation"
    elif pert_rank <= 2 and base_rank == 1:
        status = "amber"
        detail = f"Winner rank shifts from {base_rank} to {pert_rank} under ±10% indicator perturbation"
    else:
        status = "red"
        detail = f"Winner rank changes from {base_rank} to {pert_rank} under ±10% indicator perturbation"
    return {
        "id": "bc", "name": "Benefit/Cost indicator sensitivity",
        "status": status, "detail": detail, "status_label": TRAFFIC_LABELS[status],
    }


def _assess_monte_carlo(winner_idx, l3_cats, cat_scores, methods, final_w, n_iter=1000):
    names = st.session_state.proc_names
    n_proc = len(names)
    mat = np.array([cat_scores[c] for c in l3_cats])
    k_value, _, _, _ = compute_dirichlet_k(mat)
    alpha_scale = 1.0 + (k_value / 100.0) * 49.0
    alpha = np.maximum(final_w * alpha_scale, 0.05)
    rng = np.random.default_rng(42)
    draws = rng.dirichlet(alpha, size=n_iter)
    rank1_count = 0
    p_val = 0.5
    for it in range(n_iter):
        w_draw = draws[it]
        weighted_mat = mat * w_draw[:, None]
        ranks_draw = run_mcdm_suite(weighted_mat, w_draw, methods)
        if len(methods) > 1:
            psi_ranks = rank_with_ties(calc_psi(ranks_draw, methods, p_val), ascending=False)
        else:
            psi_ranks = ranks_draw[methods[0]]
        if int(psi_ranks[winner_idx]) == 1:
            rank1_count += 1
    pct = rank1_count / n_iter * 100
    status = _status_from_rank1_pct(pct)
    detail = f"Winner holds Rank 1 in {pct:.1f}% of {n_iter:,} Monte Carlo weight draws"
    return {
        "id": "monte_carlo", "name": "Monte Carlo uncertainty (Dirichlet)",
        "status": status, "detail": detail, "status_label": TRAFFIC_LABELS[status],
    }


def _assess_rank_reversal(winner_idx, l3_cats, methods, method_ranks, sel_wm, names):
    n_proc = len(names)
    if n_proc < 3:
        return {
            "id": "rank_reversal", "name": "Rank-reversal test",
            "status": "na", "detail": "Requires at least 3 alternatives",
            "status_label": TRAFFIC_LABELS["na"],
        }

    multi = len(methods) > 1
    if multi:
        baseline_all = rank_with_ties(calc_psi(method_ranks, methods, 0.5), ascending=False)
    else:
        baseline_all = method_ranks[methods[0]]

    exclude_idx = max(
        (pi for pi in range(n_proc) if pi != winner_idx),
        key=lambda pi: int(baseline_all[pi]),
    )
    keep_idx = [i for i in range(n_proc) if i != exclude_idx]
    wi_kept = keep_idx.index(winner_idx)

    base_method_ranks_kept = {}
    for m in methods:
        scores_kept = np.array([method_ranks[m][i] for i in keep_idx], dtype=float)
        base_method_ranks_kept[m] = rank_with_ties(scores_kept, ascending=True)

    cat_scores_pert = {}
    for ckey in l3_cats:
        ind_names, _, benefits = get_full_indicators(ckey)
        n_ind = len(ind_names)
        raw_full = np.zeros((n_ind, n_proc))
        for j in range(n_ind):
            for pi in range(n_proc):
                raw_full[j, pi] = st.session_state.indicator_values.get((ckey, j, pi), 0.0)
        raw_kept = raw_full[:, keep_idx]
        nm = np.zeros_like(raw_kept)
        n2 = np.zeros_like(raw_kept)
        for j in range(n_ind):
            nm[j] = merec_norm(raw_kept[j], benefits[j])
            n2[j] = n2_norm(raw_kept[j], benefits[j])
        w_ind = merec_weights(nm)
        cat_scores_pert[ckey] = (n2 * w_ind[:, None]).sum(axis=0)

    mat_pert = np.array([cat_scores_pert[c] for c in l3_cats])
    w_cat_pert = get_category_weights(mat_pert, sel_wm)
    weighted_mat_pert = mat_pert * w_cat_pert[:, None]
    ranks_pert = run_mcdm_suite(weighted_mat_pert, w_cat_pert, methods)

    if multi:
        base_rank = int(rank_with_ties(
            calc_psi(base_method_ranks_kept, methods, 0.5), ascending=False,
        )[wi_kept])
        pert_rank = int(rank_with_ties(calc_psi(ranks_pert, methods, 0.5), ascending=False)[wi_kept])
    else:
        base_rank = int(base_method_ranks_kept[methods[0]][wi_kept])
        pert_rank = int(ranks_pert[methods[0]][wi_kept])

    excluded_name = names[exclude_idx]
    if pert_rank == base_rank:
        status, detail = "green", f"No rank reversal for winner after excluding {excluded_name}"
    else:
        status, detail = "red", f"Winner rank changes ({base_rank} → {pert_rank}) after excluding {excluded_name}"
    return {
        "id": "rank_reversal", "name": "Rank-reversal test",
        "status": status, "detail": detail, "status_label": TRAFFIC_LABELS[status],
    }


def _assess_normalisation(winner_idx, l3_cats, methods, method_ranks, sel_wm):
    def minmax_norm(vals, benefit):
        vals = np.asarray(vals, dtype=float)
        mn, mx = vals.min(), vals.max()
        rng = (mx - mn) or 1.0
        return (vals - mn) / rng if benefit else (mx - vals) / rng

    multi = len(methods) > 1
    base_rank = _winner_headline_rank(method_ranks, methods, winner_idx, 0.5)

    cat_scores_alt = {}
    for ckey in l3_cats:
        ind_names, _, benefits = get_full_indicators(ckey)
        raw, _ = get_raw_matrix(ckey)
        nm = np.zeros_like(raw)
        for j in range(len(ind_names)):
            nm[j] = merec_norm(raw[j], benefits[j])
        w_ind = merec_weights(nm)
        alt_n = np.zeros_like(raw)
        for j in range(len(ind_names)):
            alt_n[j] = minmax_norm(raw[j], benefits[j])
        cat_scores_alt[ckey] = (alt_n * w_ind[:, None]).sum(axis=0)

    mat_alt = np.array([cat_scores_alt[c] for c in l3_cats])
    w_cat_alt = get_category_weights(mat_alt, sel_wm)
    weighted_mat_alt = mat_alt * w_cat_alt[:, None]
    ranks_alt = run_mcdm_suite(weighted_mat_alt, w_cat_alt, methods)
    if multi:
        alt_rank = int(rank_with_ties(calc_psi(ranks_alt, methods, 0.5), ascending=False)[winner_idx])
    else:
        alt_rank = int(ranks_alt[methods[0]][winner_idx])

    if alt_rank == base_rank:
        status, detail = "green", "Winner rank unchanged when switching N2 to Min-max normalisation"
    elif alt_rank <= 2 and base_rank == 1:
        status, detail = "amber", f"Winner rank shifts from {base_rank} to {alt_rank} under Min-max normalisation"
    else:
        status, detail = "red", f"Winner rank changes from {base_rank} to {alt_rank} under Min-max normalisation"
    return {
        "id": "normalisation", "name": "Normalisation sensitivity",
        "status": status, "detail": detail, "status_label": TRAFFIC_LABELS[status],
    }


def _assess_bootstrap(winner_idx, l3_cats, methods, sel_wm, n_iter=500, unc_pct=10):
    rng = np.random.default_rng(42)
    multi = len(methods) > 1
    rank1_count = 0
    for _ in range(n_iter):
        cat_scores_b = {}
        for ckey in l3_cats:
            raw, benefits = get_raw_matrix(ckey)
            noise = rng.uniform(1 - unc_pct / 100, 1 + unc_pct / 100, size=raw.shape)
            raw_b = np.maximum(raw * noise, 1e-9)
            nm_b = np.zeros_like(raw_b)
            n2_b = np.zeros_like(raw_b)
            for j in range(raw.shape[0]):
                nm_b[j] = merec_norm(raw_b[j], benefits[j])
                n2_b[j] = n2_norm(raw_b[j], benefits[j])
            w_ind_b = merec_weights(nm_b)
            cat_scores_b[ckey] = (n2_b * w_ind_b[:, None]).sum(axis=0)
        mat_b = np.array([cat_scores_b[c] for c in l3_cats])
        w_rcw_b = get_category_weights(mat_b, sel_wm)
        wm_b = mat_b * w_rcw_b[:, None]
        ranks_b = run_mcdm_suite(wm_b, w_rcw_b, methods)
        if multi:
            psi_ranks = rank_with_ties(calc_psi(ranks_b, methods, 0.5), ascending=False)
        else:
            psi_ranks = ranks_b[methods[0]]
        if int(psi_ranks[winner_idx]) == 1:
            rank1_count += 1
    pct = rank1_count / n_iter * 100
    status = _status_from_rank1_pct(pct)
    detail = f"Winner holds Rank 1 in {pct:.1f}% of {n_iter} bootstrap iterations (±{unc_pct}% noise)"
    return {
        "id": "bootstrap", "name": "Bootstrap stability (MEREC and RCW)",
        "status": status, "detail": detail, "status_label": TRAFFIC_LABELS[status],
    }


def _assess_indicator_uncertainty(winner_idx, l3_cats, methods, sel_wm, n_iter=500, unc_pct=10):
    multi = len(methods) > 1
    rng = np.random.default_rng(42)
    rank1_count = 0
    valid = 0
    for _ in range(n_iter):
        psi_b, ranks_b = _indicator_uncertainty_psi(
            l3_cats, sel_wm, methods, multi, 0.5,
            rng=rng, unc_pct=unc_pct, perturb=True,
        )
        if not np.all(np.isfinite(psi_b)):
            continue
        valid += 1
        if multi:
            psi_ranks = rank_with_ties(psi_b, ascending=False)
        else:
            psi_ranks = ranks_b[methods[0]]
        if int(psi_ranks[winner_idx]) == 1:
            rank1_count += 1
    if valid == 0:
        return {
            "id": "indicator_unc", "name": "Indicator-level uncertainty propagation",
            "status": "red", "detail": "Uncertainty propagation produced no valid iterations",
            "status_label": TRAFFIC_LABELS["red"],
        }
    pct = rank1_count / valid * 100
    status = _status_from_rank1_pct(pct)
    detail = f"Winner holds Rank 1 in {pct:.1f}% of {valid} valid iterations (±{unc_pct}% indicator noise)"
    return {
        "id": "indicator_unc", "name": "Indicator-level uncertainty propagation",
        "status": status, "detail": detail, "status_label": TRAFFIC_LABELS[status],
    }


def _build_dashboard_summary(checks, winner_name):
    scored = [c for c in checks if c["status"] != "na"]
    robust = [c for c in scored if c["status"] == "green"]
    partial = [c for c in scored if c["status"] == "amber"]
    sensitive = [c for c in scored if c["status"] == "red"]
    n_scored = len(scored)
    n_robust = len(robust)

    parts = [
        f"The recommendation for **{winner_name}** is robust to "
        f"**{n_robust} of {n_scored}** analyses."
    ]
    if sensitive:
        names = ", ".join(c["name"].lower() for c in sensitive)
        parts.append(f"It is sensitive to **{names}**.")
    elif partial:
        names = ", ".join(c["name"].lower() for c in partial)
        parts.append(f"It is partially sensitive to **{names}**.")
    else:
        parts.append("No material sensitivity was detected across the validation suite.")
    return " ".join(parts)


def _overall_dashboard_status(checks):
    scored = [c for c in checks if c["status"] != "na"]
    if not scored:
        return "na"
    reds = sum(1 for c in scored if c["status"] == "red")
    ambers = sum(1 for c in scored if c["status"] == "amber")
    greens = sum(1 for c in scored if c["status"] == "green")
    if reds >= 2 or (reds >= 1 and greens <= len(scored) // 2):
        return "red"
    if reds >= 1 or ambers >= 2:
        return "amber"
    if ambers >= 1:
        return "amber"
    return "green"


def _dashboard_iter_settings():
    """Iteration counts used by the Step 13 sensitivity dashboard."""
    return {
        "bs_n_iter": int(st.session_state.get("dash_bs_n_iter", st.session_state.get("bs_n_iter", 500))),
        "bs_unc_pct": int(st.session_state.get("dash_bs_unc_pct", st.session_state.get("bs_unc_pct", 10))),
        "ind_n_iter": int(st.session_state.get("dash_ind_unc_iter", st.session_state.get("ind_unc_iter", 500))),
        "ind_unc_pct": int(st.session_state.get("dash_ind_unc_pct", st.session_state.get("ind_unc_pct", 10))),
    }


def compute_sensitivity_dashboard(force=False):
    cache_key = "validation_dashboard_results"
    settings = _dashboard_iter_settings()
    cached = st.session_state.get(cache_key)
    if not force and cached and cached.get("settings") == settings:
        return cached

    method_ranks = st.session_state.last_method_ranks
    names = st.session_state.proc_names
    methods = list(st.session_state.sel_mcdm_methods) or list(method_ranks.keys())
    l3_cats = ordered_l3_cats()
    cat_scores = st.session_state.cat_scores
    sel_wm = st.session_state.sel_weight_methods
    final_w = st.session_state.final_cat_weights

    winner_info = _resolve_overall_winner(names, method_ranks, methods)
    wi = winner_info["winner_idx"]

    checks = [
        _assess_weight_sensitivity(wi, l3_cats, cat_scores, methods),
        _assess_bc_sensitivity(wi, l3_cats, methods, method_ranks, sel_wm),
        _assess_monte_carlo(wi, l3_cats, cat_scores, methods, final_w),
        _assess_rank_reversal(wi, l3_cats, methods, method_ranks, sel_wm, names),
        _assess_normalisation(wi, l3_cats, methods, method_ranks, sel_wm),
        _assess_bootstrap(
            wi, l3_cats, methods, sel_wm,
            n_iter=settings["bs_n_iter"], unc_pct=settings["bs_unc_pct"],
        ),
        _assess_indicator_uncertainty(
            wi, l3_cats, methods, sel_wm,
            n_iter=settings["ind_n_iter"], unc_pct=settings["ind_unc_pct"],
        ),
    ]

    overall = _overall_dashboard_status(checks)
    summary = _build_dashboard_summary(checks, winner_info["winner_name"])
    scored = [c for c in checks if c["status"] != "na"]
    result = {
        "checks": checks,
        "winner_name": winner_info["winner_name"],
        "winner_basis": winner_info["basis_label"],
        "robust_count": sum(1 for c in scored if c["status"] == "green"),
        "scored_count": len(scored),
        "overall_status": overall,
        "overall_label": TRAFFIC_LABELS[overall],
        "summary": summary,
        "settings": settings,
    }
    st.session_state[cache_key] = result
    st.session_state.validation_choice = "Sensitivity dashboard"
    return result


def validation_intro():
    st.header("Step 13 - Validation (optional)")

    method_ranks = st.session_state.get("last_method_ranks")
    if not method_ranks:
        st.warning("Complete Step 12 first to generate rankings before running validation.")
        st.divider()
        c1, c2, c3 = st.columns(3)
        with c1:
            if st.button("<- Back to results"):
                st.session_state.step = 12
                st.rerun()
        with c2:
            if st.button("Decision support ->", type="primary"):
                st.session_state.step = 15
                st.rerun()
        with c3:
            if st.button("Reset all"):
                reset_all()
                st.rerun()
        return

    c_refresh, c_info = st.columns([1, 3])
    with c_refresh:
        force = st.button("Refresh dashboard", type="primary", key="val_dash_refresh")
    with c_info:
        st.caption(
            "All seven validation checks run automatically and focus on the "
            "recommended winner from Step 12."
        )

    st.markdown("**Dashboard iteration settings**")
    dc1, dc2, dc3, dc4 = st.columns(4)
    with dc1:
        st.slider(
            "Bootstrap iterations",
            min_value=100, max_value=2000,
            value=int(st.session_state.get("dash_bs_n_iter", 500)), step=100,
            key="dash_bs_n_iter",
            help="Iterations for check 6 in the dashboard summary.",
        )
    with dc2:
        st.slider(
            "Bootstrap ±%",
            min_value=1, max_value=50,
            value=int(st.session_state.get("dash_bs_unc_pct", 10)), step=1,
            key="dash_bs_unc_pct",
            help="Indicator noise range for dashboard bootstrap check.",
        )
    with dc3:
        st.slider(
            "Indicator uncertainty iterations",
            min_value=100, max_value=10000,
            value=int(st.session_state.get("dash_ind_unc_iter", 500)), step=100,
            key="dash_ind_unc_iter",
            help="Iterations for check 7 in the dashboard summary.",
        )
    with dc4:
        st.slider(
            "Indicator uncertainty ±%",
            min_value=1, max_value=50,
            value=int(st.session_state.get("dash_ind_unc_pct", 10)), step=1,
            key="dash_ind_unc_pct",
            help="Indicator noise range for dashboard uncertainty check.",
        )

    with st.spinner("Running sensitivity dashboard..."):
        dashboard = compute_sensitivity_dashboard(force=force)

    overall = dashboard["overall_status"]
    st.markdown(
        f"<span class='prism-traffic-badge prism-traffic-{overall}'>"
        f"Overall recommendation: {dashboard['overall_label']}</span>",
        unsafe_allow_html=True,
    )
    st.markdown(
        f"<div class='prism-about-panel'>"
        f"<p class='prism-about-body'>{dashboard['summary']}</p>"
        f"<p class='prism-about-body' style='margin-top:8px;color:{_BRAND_MUTED};'>"
        f"Recommended alternative: <strong>{dashboard['winner_name']}</strong> "
        f"({dashboard['winner_basis']})</p></div>",
        unsafe_allow_html=True,
    )

    table_rows = [
        {
            "Analysis": c["name"],
            "Status": c["status_label"],
            "Finding": c["detail"],
        }
        for c in dashboard["checks"]
    ]
    st.dataframe(pd.DataFrame(table_rows), use_container_width=True, hide_index=True)

    st.markdown("##### Detailed validation views")
    with st.expander("1. Weighting-method sensitivity"):
        validation_weight_sensitivity()
    with st.expander("2. Benefit/Cost indicator sensitivity"):
        validation_bc_sensitivity()
    with st.expander("3. Monte Carlo uncertainty (Dirichlet)"):
        validation_monte_carlo()
    with st.expander("4. Rank-reversal test"):
        validation_rank_reversal()
    with st.expander("5. Normalisation sensitivity"):
        validation_normalisation_sensitivity()
    with st.expander("6. Bootstrap stability (MEREC and RCW)"):
        validation_bootstrap_merec_rcw()
    with st.expander("7. Indicator-level uncertainty propagation"):
        validation_indicator_uncertainty()

    st.divider()
    c1, c2, c3 = st.columns(3)
    with c1:
        if st.button("<- Back to results"):
            st.session_state.step = 12
            st.rerun()
    with c2:
        if st.button("Decision support ->", type="primary"):
            st.session_state.step = 15
            st.rerun()
    with c3:
        if st.button("Reset all"):
            reset_all()
            st.rerun()


def validation_weight_sensitivity():
    st.subheader("1. Weighting-method sensitivity")

    names = st.session_state.proc_names
    l3_cats = ordered_l3_cats()
    cat_scores = st.session_state.cat_scores
    mat = np.array([cat_scores[c] for c in l3_cats])

    rows = [{"Alternative": name} for name in names]
    for combo_methods, combo_label in WEIGHT_COMBO_SETS:
        w = get_category_weights(mat, set(combo_methods))
        weighted_mat = mat * w[:, None]
        ranks = run_mcdm_suite(weighted_mat, w, ALL_MCDM_KEYS)
        for pi in range(len(names)):
            for mkey in ALL_MCDM_KEYS:
                col = f"{combo_label} - {METHOD_LABELS[mkey]}"
                rows[pi][col] = int(ranks[mkey][pi])

    df = pd.DataFrame(rows)
    st.session_state["export_weight_sens"] = df
    st.dataframe(df, use_container_width=True, hide_index=True)
    rank_cols = [c for c in df.columns if c != "Alternative"]
    any_change = any(
        len(set(df.loc[i, rank_cols].tolist())) > 1
        for i in range(len(df))
    )
    if any_change:
        st.markdown(
            '<div style="background:#FFF8E1;border-left:3px solid #D97706;'
            'padding:8px 12px;border-radius:4px;margin:6px 0;'
            'font-family:Times New Roman,serif;color:#7B5800;font-size:13px;">'
            'Rank variation detected across weighting schemes.</div>',
            unsafe_allow_html=True)
    else:
        st.markdown(
            '<div style="background:#E8F5E9;border-left:3px solid #16A34A;'
            'padding:8px 12px;border-radius:4px;margin:6px 0;'
            'font-family:Times New Roman,serif;color:#1B5E20;font-size:13px;">'
            'No rank changes detected. The ranking is robust to category '
            'weighting method choice.</div>',
            unsafe_allow_html=True)

    st.divider()
    st.markdown("**Overall ranking across category combinations (this weighting scheme set)**")

    wm_options = {"equal": "Equal", "entropy": "Entropy", "critic": "CRITIC"}
    sens1_weight_help = {
        "equal": FIELD_HELP["weight_equal"],
        "entropy": FIELD_HELP["weight_entropy"],
        "critic": FIELD_HELP["weight_critic"],
    }
    if "sens1_wm" not in st.session_state:
        st.session_state.sens1_wm = set(st.session_state.sel_weight_methods) or {"equal"}
    cols = st.columns(3)
    for i, (k, lbl) in enumerate(wm_options.items()):
        with cols[i]:
            checked = k in st.session_state.sens1_wm
            new_val = st.checkbox(lbl, value=checked, key=f"sens1wm_{k}",
                                  help=sens1_weight_help.get(k))
            if new_val:
                st.session_state.sens1_wm.add(k)
            else:
                st.session_state.sens1_wm.discard(k)
    if not st.session_state.sens1_wm:
        st.session_state.sens1_wm = {"equal"}

    sens1_p = st.slider("p for combination view", min_value=0.0, max_value=1.0,
                         value=0.5, step=0.01, key="sens1_p_slider",
                         help=FIELD_HELP["combo_p"])

    combos = get_combinations(l3_cats)
    cat_initial = {c: CAT_SHORT.get(c, CATS[c]["label"][:1]) for c in l3_cats}
    combo_labels = ["+".join(cat_initial[c] for c in combo) for combo in combos]
    n_proc = len(names)

    rank_grid = np.zeros((n_proc, len(combos)), dtype=int)
    for ci, combo in enumerate(combos):
        if len(combo) == 1:
            rank_grid[:, ci] = rank_with_ties(cat_scores[combo[0]], ascending=False)
        else:
            sub_mat = np.array([cat_scores[c] for c in combo])
            sub_w = get_category_weights(sub_mat, st.session_state.sens1_wm)
            sub_weighted = sub_mat * sub_w[:, None]
            sub_ranks = run_mcdm_suite(sub_weighted, sub_w, ALL_MCDM_KEYS)
            psi_combo = calc_psi(sub_ranks, ALL_MCDM_KEYS, sens1_p)
            rank_grid[:, ci] = rank_with_ties(psi_combo, ascending=False)

    apply_mpl_style()
    n_combos_c = len(combos)
    jitter_step = 0.12
    fig_val_combo = go.Figure()
    for pi, name in enumerate(names):
        offset = (pi - (len(names)-1)/2) * jitter_step
        x_jit = [ci + offset for ci in range(n_combos_c)]
        hover = [
            f"<b>{name}</b><br>Combination: {combo_labels[ci]}<br>Rank: {rank_grid[pi, ci]}"
            for ci in range(n_combos_c)
        ]
        fig_val_combo.add_trace(go.Scatter(
            x=x_jit,
            y=rank_grid[pi, :],
            mode="markers",
            name=name,
            marker=dict(size=9, symbol="square",
                        color=PROC_COLORS[pi % len(PROC_COLORS)]),
            hovertemplate="%{customdata}<extra></extra>",
            customdata=hover,
        ))
    fig_val_combo.update_layout(
        xaxis=dict(
            tickmode="array",
            tickvals=list(range(n_combos_c)),
            ticktext=combo_labels,
            tickangle=-90,
            tickfont=dict(size=9, family="Times New Roman"),
            title="Category combination",
            title_font=dict(family="Times New Roman", color="#0D2B5E"),
        ),
        yaxis=dict(
            title="Rank",
            autorange="reversed",
            dtick=1, tick0=1,
            range=[0.5, n_proc + 0.5],
            gridcolor="rgba(0,0,0,0.1)",
            gridwidth=1,
            griddash="dash",
            title_font=dict(family="Times New Roman", color="#0D2B5E"),
            tickfont=dict(family="Times New Roman"),
        ),
        legend=dict(orientation="h", yanchor="bottom", y=1.02,
                    font=dict(family="Times New Roman")),
        height=420,
        margin=dict(l=10, r=10, t=40, b=120),
        plot_bgcolor="white",
        paper_bgcolor="white",
        font=dict(family="Times New Roman", color="#0D2B5E"),
        hoverlabel=dict(font_family="Times New Roman"),
    )
    st.plotly_chart(fig_val_combo, use_container_width=True)


def validation_bc_sensitivity():
    st.subheader("2. Benefit/Cost indicator sensitivity")

    l3_cats = ordered_l3_cats()
    names = st.session_state.proc_names
    sel_weight_methods = st.session_state.sel_weight_methods
    sel_mcdm_methods = list(st.session_state.sel_mcdm_methods) or ALL_MCDM_KEYS
    method_ranks = st.session_state.get("last_method_ranks", {})

    c1, c2 = st.columns(2)
    with c1:
        ben_pct = st.slider("Benefit-type indicators (%)", min_value=-25, max_value=25,
                             value=0, step=5, key="bc_ben_slider",
                             help=FIELD_HELP["ben_pct"])
    with c2:
        cost_pct = st.slider("Cost-type indicators (%)", min_value=-25, max_value=25,
                              value=0, step=5, key="bc_cost_slider",
                              help=FIELD_HELP["cost_pct"])

    p_val = st.slider("p value", min_value=0.0, max_value=1.0, value=0.5, step=0.01,
                       key="bc_p_slider", help=FIELD_HELP["psi_p"])

    multi = len(sel_mcdm_methods) > 1
    cols = ["Alternative"] + [METHOD_LABELS[m] for m in sel_mcdm_methods]
    if multi and method_ranks:
        cols.append("PSI Rank (p=0.50)")
        base_psi = calc_psi(method_ranks, sel_mcdm_methods, 0.5)
        base_psi_rank = rank_with_ties(base_psi, ascending=False)

    base_rows = []
    for pi, name in enumerate(names):
        row = [name] + [int(method_ranks[m][pi]) for m in sel_mcdm_methods if m in method_ranks]
        if multi and method_ranks:
            row.append(int(base_psi_rank[pi]))
        base_rows.append(row)

    st.markdown("**Baseline results (Step 12, unperturbed)**")
    st.dataframe(pd.DataFrame(base_rows, columns=cols), use_container_width=True, hide_index=True)

    rng_bc = np.random.default_rng(42)
    cat_scores_pert = {}
    for ckey in l3_cats:
        raw, benefits = get_raw_matrix(ckey)
        raw_pert = raw.copy()
        n_ind_bc, n_p_bc = raw.shape
        for j in range(n_ind_bc):
            pct = ben_pct if benefits[j] else cost_pct
            if pct == 0:
                raw_pert[j] = raw[j].copy()  # no perturbation at 0%
            else:
                half = abs(pct) * 0.5  # ±50% of the stated % as random spread
                noise = rng_bc.uniform(-half, half, size=n_p_bc)
                factors = 1 + (pct + noise) / 100.0
                raw_pert[j] = np.maximum(raw[j] * factors, 1e-9)
        cat_scores_pert[ckey] = compute_category_score_from_raw(ckey, raw_pert)

    mat = np.array([cat_scores_pert[c] for c in l3_cats])
    w = get_category_weights(mat, sel_weight_methods)
    weighted_mat = mat * w[:, None]
    ranks = run_mcdm_suite(weighted_mat, w, sel_mcdm_methods)

    p_cols = ["Alternative"] + [METHOD_LABELS[m] for m in sel_mcdm_methods]
    if multi:
        p_cols.append(f"PSI Rank (p={p_val:.2f})")
        psi_vals = calc_psi(ranks, sel_mcdm_methods, p_val)
        psi_ranks = rank_with_ties(psi_vals, ascending=False)

    p_rows = []
    for pi, name in enumerate(names):
        row = [name] + [int(ranks[m][pi]) for m in sel_mcdm_methods]
        if multi:
            row.append(int(psi_ranks[pi]))
        p_rows.append(row)

    st.markdown(
        f"**Perturbed results** — Benefit indicators: **{ben_pct:+d}%**, "
        f"Cost indicators: **{cost_pct:+d}%**, p = **{p_val:.2f}**"
    )
    df_bc_export = pd.DataFrame(p_rows, columns=p_cols)
    st.session_state["export_bc_sens"] = df_bc_export
    st.dataframe(df_bc_export, use_container_width=True, hide_index=True)

    # ── Rank change summary ───────────────────────────────────────────────────
    base_df  = pd.DataFrame(base_rows, columns=cols)
    pert_df  = df_bc_export
    any_rank_change = False
    for m in sel_mcdm_methods:
        ml = METHOD_LABELS[m]
        if ml in base_df.columns and ml in pert_df.columns:
            if not (base_df[ml].values == pert_df[ml].values).all():
                any_rank_change = True
                break
    if multi and method_ranks:
        psi_col_base = "PSI Rank (p=0.50)"
        psi_col_pert = f"PSI Rank (p={p_val:.2f})"
        if psi_col_base in base_df.columns and psi_col_pert in pert_df.columns:
            if not (base_df[psi_col_base].values == pert_df[psi_col_pert].values).all():
                any_rank_change = True

    if not any_rank_change:
        st.markdown(
            f'<div style="background:#E8F5E9;border-left:3px solid #16A34A;'
            f'padding:8px 12px;border-radius:4px;margin:6px 0;'
            f'font-family:Times New Roman,Tinos,serif;font-size:13px;color:#1B5E20;">'
            f'✅ No rank changes detected under {ben_pct:+d}% benefit / '
            f'{cost_pct:+d}% cost perturbation. '
            f'The ranking is robust to this level of indicator uncertainty. '
            f'This is expected when absolute performance differences between '
            f'alternatives are large relative to the perturbation magnitude.</div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f'<div style="background:#FFF8E1;border-left:3px solid #D97706;'
            f'padding:8px 12px;border-radius:4px;margin:6px 0;'
            f'font-family:Times New Roman,Tinos,serif;font-size:13px;color:#7B5800;">'
            f'⚠️ Rank changes detected under {ben_pct:+d}% benefit / '
            f'{cost_pct:+d}% cost perturbation. '
            f'Review which alternatives are sensitive.</div>',
            unsafe_allow_html=True,
        )

    # ── Category score changes ────────────────────────────────────────────────
    st.markdown("**Category score changes under perturbation**")
    st.caption("Even when ranks are unchanged, the magnitude of category scores shifts — "
               "showing the perturbation was applied and the analysis is working correctly.")
    base_scores = st.session_state.cat_scores
    score_rows = []
    for pi, name in enumerate(names):
        row = {"Process": name}
        for ckey in l3_cats:
            orig = float(base_scores[ckey][pi])
            pert = float(cat_scores_pert[ckey][pi])
            delta = pert - orig
            pct_chg = (delta / orig * 100) if orig > 1e-10 else 0.0
            row[f"{CATS[ckey]['label']} (orig)"]   = round(orig, 4)
            row[f"{CATS[ckey]['label']} (pert)"]   = round(pert, 4)
            row[f"{CATS[ckey]['label']} (Δ%)"]     = f"{pct_chg:+.1f}%"
        score_rows.append(row)
    st.dataframe(pd.DataFrame(score_rows), use_container_width=True,
                 hide_index=True)


def validation_monte_carlo():
    st.subheader("3. Monte Carlo uncertainty (Dirichlet)")

    l3_cats = ordered_l3_cats()
    names = st.session_state.proc_names
    n_proc = len(names)
    cat_scores = st.session_state.cat_scores
    final_w = st.session_state.final_cat_weights
    sel_mcdm_methods = list(st.session_state.sel_mcdm_methods) or ALL_MCDM_KEYS

    mat = np.array([cat_scores[c] for c in l3_cats])
    k_value, w_eq, w_en, w_cr = compute_dirichlet_k(mat)

    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("Data-driven k (inter-method agreement)", f"{k_value:.1f} / 100")
    with c2:
        use_custom_k = st.checkbox("Override k manually", value=False, key="mc_use_custom_k",
                                   help=FIELD_HELP["mc_override"])
        if use_custom_k:
            k_value = st.slider("k value", min_value=1.0, max_value=100.0,
                                 value=float(min(k_value, 50.0)), step=1.0, key="mc_k_override",
                                 help=FIELD_HELP["mc_k"])
    with c3:
        p_val = st.slider("p value (for PSI compromise rank)", min_value=0.0, max_value=1.0,
                           value=0.5, step=0.01, key="mc_p_slider",
                           help=FIELD_HELP["psi_p"])

    with st.expander("Show Equal/Entropy/CRITIC weights used to compute k"):
        wdf = pd.DataFrame(
            [w_eq, w_en, w_cr], index=["Equal", "Entropy", "CRITIC"],
            columns=[CATS[c]["label"] for c in l3_cats],
        ).round(4)
        st.dataframe(wdf, use_container_width=True)

    run_mc = st.button("Run Monte Carlo simulation (10,000 iterations)", type="primary")

    if run_mc:
        n_iter = 10000
        alpha_scale = 1.0 + (k_value / 100.0) * 49.0  # range 1-50
        alpha = np.maximum(final_w * alpha_scale, 0.05)

        rng = np.random.default_rng()
        draws = rng.dirichlet(alpha, size=n_iter)

        rank_counts = {m: np.zeros((n_proc, n_proc), dtype=int) for m in sel_mcdm_methods}
        psi_rank_counts = np.zeros((n_proc, n_proc), dtype=int) if len(sel_mcdm_methods) > 1 else None

        progress = st.progress(0, text="Running Monte Carlo simulation...")
        batch = max(1, n_iter // 20)

        for it in range(n_iter):
            w_draw = draws[it]
            weighted_mat = mat * w_draw[:, None]
            ranks_draw = run_mcdm_suite(weighted_mat, w_draw, sel_mcdm_methods)
            for m in sel_mcdm_methods:
                for pi in range(n_proc):
                    r = ranks_draw[m][pi]
                    rank_counts[m][pi, r - 1] += 1
            if psi_rank_counts is not None:
                psi_vals = calc_psi(ranks_draw, sel_mcdm_methods, p_val)
                psi_ranks = rank_with_ties(psi_vals, ascending=False)
                for pi in range(n_proc):
                    psi_rank_counts[pi, psi_ranks[pi] - 1] += 1
            if it % batch == 0:
                progress.progress(min(1.0, it / n_iter), text=f"Running Monte Carlo simulation... {it}/{n_iter}")
        progress.progress(1.0, text="Done.")

        st.session_state.mc_rank_counts = rank_counts
        st.session_state["export_mc"] = rank_counts
        st.session_state.mc_psi_rank_counts = psi_rank_counts
        st.session_state.mc_n_iter = n_iter
        st.session_state.mc_methods_used = sel_mcdm_methods
        st.session_state.mc_k_value = k_value

    if "mc_rank_counts" in st.session_state:
        rank_counts = st.session_state.mc_rank_counts
        psi_rank_counts = st.session_state.mc_psi_rank_counts
        n_iter = st.session_state.mc_n_iter
        methods_used = st.session_state.mc_methods_used

        st.markdown(
            f"**Results from {n_iter:,} iterations** "
            f"(k = {st.session_state.mc_k_value:.1f}, rank distribution as % of draws)"
        )

        for m in methods_used:
            st.markdown(f"**{METHOD_LABELS[m]}**")
            pct_table = (rank_counts[m] / n_iter * 100).round(1)
            df = pd.DataFrame(pct_table, index=names, columns=[f"Rank {r+1}" for r in range(n_proc)])
            st.dataframe(df, use_container_width=True)

        if psi_rank_counts is not None:
            st.markdown(f"**PSI compromise rank (p={p_val:.2f})**")
            pct_table = (psi_rank_counts / n_iter * 100).round(1)
            df = pd.DataFrame(pct_table, index=names, columns=[f"Rank {r+1}" for r in range(n_proc)])
            st.dataframe(df, use_container_width=True)



# ============================================================================
# STEP 14 - AUXILIARY ASSESSMENT (optional)
# ============================================================================





def analytics_category_weighted_contribution():
    """
    Category contribution to weighted MCDM input analysis.

    Formulae (from PRISM methodology):
      S_c,i  = CategoryScore_c,i × w_c^RCW
      T_i    = Σ_c S_c,i
      Pct^PSI_c,i = (S_c,i / T_i) × 100
      PSI_c,i = PSI_i × (S_c,i / T_i)

    Each row sums to 100% — directly linked to PSI.
    """
    st.subheader("A. Category contribution to weighted MCDM input")

    names     = st.session_state.proc_names
    n_proc    = len(names)
    l3_cats   = ordered_l3_cats()
    cat_scores = st.session_state.cat_scores
    final_w   = st.session_state.final_cat_weights
    method_ranks = st.session_state.get("last_method_ranks", {})
    sel_mcdm  = list(st.session_state.sel_mcdm_methods) or ALL_MCDM_KEYS
    multi     = len(sel_mcdm) > 1

    p_val = st.slider("p value (PSI)", min_value=0.0, max_value=1.0,
                       value=0.5, step=0.01, key="cat_contrib_p",
                       help=FIELD_HELP["psi_p"])

    # ── Step 1: S_c,i = CategoryScore × w_c^RCW ─────────────────────────────
    S = np.array([cat_scores[c] * final_w[ci]
                  for ci, c in enumerate(l3_cats)])  # (n_cats, n_proc)

    # ── Step 2: T_i = Σ_c S_c,i ─────────────────────────────────────────────
    T = S.sum(axis=0)  # (n_proc,)

    # ── Step 3: Pct^PSI_c,i = S_c,i / T_i × 100 ─────────────────────────────
    pct = np.where(T > 0, S / T * 100, 0.0)  # (n_cats, n_proc)

    # ── Step 4: PSI scores for reference ─────────────────────────────────────
    if multi and method_ranks:
        psi_scores = calc_psi(method_ranks, sel_mcdm, p_val)
    else:
        psi_scores = np.ones(n_proc)

    shorts = proc_short_labels(names)

    # ── Chart A: Stacked 100% bar — Pct^PSI per category ─────────────────────
    st.markdown(
        "<p style='font-family:Times New Roman,Tinos,serif;font-size:13px;"
        "color:#0D2B5E;margin:4px 0;'>"
        "<b>Pct<sub>c,i</sub></b> = "
        "(<i>S<sub>c,i</sub></i> / <i>T<sub>i</sub></i>) × 100 &nbsp;|&nbsp; "
        "<i>S<sub>c,i</sub></i> = CategoryScore<sub>c,i</sub> × <i>w<sub>c</sub><sup>RCW</sup></i>"
        "</p>",
        unsafe_allow_html=True,
    )

    apply_mpl_style()
    fig, ax = plt.subplots(figsize=(max(5, n_proc * 1.5), 3.8))
    bottoms = np.zeros(n_proc)
    for ci, ckey in enumerate(l3_cats):
        ax.bar(shorts, pct[ci], bottom=bottoms,
               color=CATS[ckey]["color"], label=CATS[ckey]["label"],
               edgecolor="white", linewidth=0.5)
        bottoms += pct[ci]
    ax.set_ylabel("% contribution to weighted MCDM input", fontsize=10)
    ax.set_ylim(0, 105)
    ax.axhline(y=100, color="#0D2B5E", linewidth=0.5, linestyle="--", alpha=0.4)
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, 1.14),
              ncol=min(len(l3_cats), 3), fontsize=8, frameon=False)
    ax.tick_params(labelsize=9)
    plt.tight_layout()
    mpl_show(fig)

    # ── Table A: Pct^PSI per category ────────────────────────────────────────
    st.markdown("**% contribution of each category to the weighted MCDM input $T_i$ (each row sums to 100%)**")
    rows_pct = []
    for pi, name in enumerate(names):
        row = {"Process": name, "PSI rank": int(rank_with_ties(psi_scores, ascending=False)[pi]), "PSI score (ref)": round(float(psi_scores[pi]), 4)}
        for ci, ckey in enumerate(l3_cats):
            row[f"{CATS[ckey]['label']} (%)"] = round(float(pct[ci, pi]), 1)
        row["Total (%)"] = 100.0
        rows_pct.append(row)
    df_pct = pd.DataFrame(rows_pct)
    st.session_state["export_cat_contrib"] = df_pct
    st.dataframe(df_pct, use_container_width=True, hide_index=True)



    # ── RCW weights reference ─────────────────────────────────────────────────
    with st.expander("RCW category weights"):
        w_rows = [{"Category": CATS[l3_cats[ci]]["label"],
                   "RCW weight": round(float(final_w[ci]), 4),
                   "Weight (%)": round(float(final_w[ci]) * 100, 1)}
                  for ci in range(len(l3_cats))]
        st.dataframe(pd.DataFrame(w_rows), use_container_width=True, hide_index=True)


def analytics_combined_contribution():
    """
    Combined contribution analysis — nested donut per process.

    Inner ring: categories sized by Pct_c,i = S_c,i / T_i × 100
                % label shown inside each slice
    Outer ring: indicators sized by Pct^PSI_j,i = n2×w^MEREC×w^RCW / T × 100
                leader lines to labels outside the ring for ALL indicators
    Centre:     short process name + PSI score
    """
    st.subheader("1. Contribution analysis")

    names    = st.session_state.proc_names
    n_proc   = len(names)
    l3_cats  = ordered_l3_cats()
    n2_data  = st.session_state.n2_data
    merec_w  = st.session_state.merec_w
    final_w  = st.session_state.final_cat_weights
    method_ranks = st.session_state.get("last_method_ranks", {})
    sel_mcdm = list(st.session_state.sel_mcdm_methods) or ALL_MCDM_KEYS
    multi    = len(sel_mcdm) > 1

    p_val = st.slider("p value (PSI)", min_value=0.0, max_value=1.0,
                       value=0.5, step=0.01, key="combined_contrib_p",
                       help=FIELD_HELP["psi_p"])

    if multi and method_ranks:
        psi_scores = calc_psi(method_ranks, sel_mcdm, p_val)
    else:
        psi_scores = np.ones(n_proc)

    # ── Pre-compute S_c,i and T_i ─────────────────────────────────────────────
    S_cat = np.array([
        st.session_state.cat_scores[c] * final_w[ci]
        for ci, c in enumerate(l3_cats)
    ])
    T = S_cat.sum(axis=0)
    cat_pct = np.where(T > 0, S_cat / T * 100, 0.0)

    # ── Build flat indicator lists ─────────────────────────────────────────────
    ind_labels  = []
    ind_units_l = []
    ind_cat_key = []
    ind_psi_pct = []   # (n_ind_total, n_proc)

    for ci, ckey in enumerate(l3_cats):
        ind_names, ind_units, _ = get_full_indicators(ckey)
        n_ind = len(ind_names)
        n2    = n2_data[ckey]
        w_ind = merec_w[ckey]
        u     = n2 * w_ind[:, None]
        psi_pct = np.where(T > 0, u * final_w[ci] / T * 100, 0.0)
        for j in range(n_ind):
            ind_labels.append(ind_names[j])
            ind_units_l.append(ind_units[j])
            ind_cat_key.append(ckey)
            ind_psi_pct.append(psi_pct[j])

    ind_psi_pct = np.array(ind_psi_pct)  # (n_total, n_proc)
    n_total = len(ind_labels)
    n_cats  = len(l3_cats)

    shorts_c = proc_short_labels(names)

    # ── Short name map for chart labels ───────────────────────────────────────
    _SHORT = {
        # Environmental
        "cumulative energy demand": "CED",
        "co2 emissions":            "CO₂",
        "co₂ emissions":            "CO₂",
        "water consumption":        "Water",
        # Economic
        "material cost":            "Mat. cost",
        "machine cost":             "Mach. cost",
        "labour cost":              "Labour",
        "consumables cost":         "Consumab.",
        "energy cost":              "Energy",
        # Social
        "recordable injury rate":   "Injury",
        "job satisfaction":         "Wage",
        "average operator wage":    "Wage",
        # Quality
        "tensile strength":         "Tensile",
        "yield strength":           "Yield",
        "elongation":               "Elongation",
        "surface roughness":        "Roughness",
        "hardness":                 "Hardness",
        # Productivity
        "production time":          "Prod. time",
        "material utilisation":     "Mat. util.",
        "material utilization":     "Mat. util.",
        "build rate":               "Build rate",
        "deposition rate":          "Dep. rate",
    }

    def short_name(full):
        return _SHORT.get(full.lower().strip(), full[:10])

    # ── Pre-build outer colours ────────────────────────────────────────────────
    cat_counts = {}
    for ck in ind_cat_key:
        cat_counts[ck] = cat_counts.get(ck, 0) + 1
    cat_idx = {}
    outer_colors_base = []
    for j in range(n_total):
        ck    = ind_cat_key[j]
        idx   = cat_idx.get(ck, 0)
        count = cat_counts[ck]
        alpha = 0.42 + 0.58 * (idx / max(count - 1, 1))
        hex_c = CATS[ck]["color"].lstrip("#")
        r,g,b = (int(hex_c[i:i+2],16)/255 for i in (0,2,4))
        outer_colors_base.append((r, g, b, alpha))
        cat_idx[ck] = idx + 1

    from matplotlib.patches import Patch
    import matplotlib.patheffects as pe

    for pi, name in enumerate(names):
        apply_mpl_style()
        # Large figure with extra right margin for legend
        fig, ax = plt.subplots(figsize=(10.0, 9.5))
        ax.set_aspect("equal")

        inner_sizes  = [float(cat_pct[ci, pi]) for ci in range(n_cats)]
        inner_colors = [CATS[ckey]["color"] for ckey in l3_cats]
        outer_sizes  = [max(float(ind_psi_pct[j, pi]), 0.001)
                        for j in range(n_total)]

        # ── Inner ring — categories ──────────────────────────────────────────
        wedges_inner, _ = ax.pie(
            inner_sizes,
            radius=0.62,
            colors=inner_colors,
            startangle=90,
            wedgeprops=dict(width=0.30, edgecolor="white", linewidth=2.0),
            labels=None,
            counterclock=False,
        )

        # ── Outer ring — indicators ──────────────────────────────────────────
        wedges_outer, _ = ax.pie(
            outer_sizes,
            radius=1.18,
            colors=outer_colors_base,
            startangle=90,
            wedgeprops=dict(width=0.52, edgecolor="white", linewidth=1.0),
            labels=None,
            counterclock=False,
        )

        # ── Centre — short process name + PSI ─────────────────────────────────
        ax.text(0, 0.09, shorts_c[pi],
                ha="center", va="center", fontsize=16, fontweight="bold",
                fontfamily="Times New Roman", color="#0D2B5E")
        if multi and method_ranks:
            ax.text(0, -0.12, "PSI  " + f"{psi_scores[pi]:.4f}",
                    ha="center", va="center", fontsize=8.5,
                    fontfamily="Times New Roman", color="#444444")

        # ── Category % labels inside inner ring ─────────────────────────────
        for ci, (wedge, sz) in enumerate(zip(wedges_inner, inner_sizes)):
            if sz < 1.0:   # only skip truly invisible slices
                continue
            ang = (wedge.theta1 + wedge.theta2) / 2
            r_lbl = 0.47
            x = r_lbl * np.cos(np.deg2rad(ang))
            y = r_lbl * np.sin(np.deg2rad(ang))
            # For small slices show only %, for larger show abbrev + %
            if sz < 5:
                lbl = f'{sz:.1f}%'
            else:
                lbl = CATS[l3_cats[ci]]['label'][:3] + '\n' + f'{sz:.1f}%'
            ax.text(x, y, lbl, ha="center", va="center",
                    fontsize=7.5 if sz >= 5 else 6.5,
                    fontweight="bold",
                    fontfamily="Times New Roman", color="white")

        # ── Leader lines + labels for ALL outer indicators ────────────────────
        label_radius = 1.48  # where label anchor sits
        line_start   = 1.20  # just outside outer ring edge
        line_end     = 1.36  # end of leader line

        for j, (wedge, sz) in enumerate(zip(wedges_outer, outer_sizes)):
            ang_deg = (wedge.theta1 + wedge.theta2) / 2
            ang_rad = np.deg2rad(ang_deg)
            cos_a   = np.cos(ang_rad)
            sin_a   = np.sin(ang_rad)

            # Leader line: start at outer edge → elbow → label
            x0 = line_start * cos_a
            y0 = line_start * sin_a
            x1 = line_end   * cos_a
            y1 = line_end   * sin_a

            ax.annotate(
                "",
                xy=(x1, y1), xytext=(x0, y0),
                arrowprops=dict(arrowstyle="-", color="#888888",
                                lw=0.7, shrinkA=0, shrinkB=0),
            )

            # Label position — push further out left/right based on angle
            ha = "left" if cos_a >= 0 else "right"
            xl = label_radius * cos_a
            yl = label_radius * sin_a

            short_nm = short_name(ind_labels[j])
            lbl_txt  = f"{short_nm}  {float(ind_psi_pct[j,pi]):.1f}%"

            ax.text(xl, yl, lbl_txt,
                    ha=ha, va="center",
                    fontsize=12.0,
                    fontfamily="Times New Roman",
                    color="#111111",
                    bbox=dict(boxstyle="round,pad=0.18",
                              fc="white", ec="none", alpha=0.85))

        ax.set_xlim(-1.72, 1.72)
        ax.set_ylim(-1.72, 1.72)
        ax.set_title(
            f"{name}",
            fontsize=13, fontweight="bold",
            fontfamily="Times New Roman", color="#0D2B5E", pad=10,
        )
        plt.tight_layout()
        mpl_show(fig)

    # ── Summary tables in expander ────────────────────────────────────────────
    with st.expander("Detailed contribution tables"):
        # Category table
        st.markdown("**Category contribution to weighted MCDM input**")
        rows_cat = []
        for pi, pname in enumerate(names):
            row = {"Process": pname,
                   "PSI rank": int(rank_with_ties(psi_scores,
                                                   ascending=False)[pi])}
            for ci, ckey in enumerate(l3_cats):
                row[f"{CATS[ckey]['label']} (%)"] = round(
                    float(cat_pct[ci, pi]), 1)
            row["Total (%)"] = 100.0
            rows_cat.append(row)
        st.dataframe(pd.DataFrame(rows_cat),
                     use_container_width=True, hide_index=True)
        st.divider()

        # Indicator table per category
        for ci, ckey in enumerate(l3_cats):
            cat = CATS[ckey]
            ind_names, ind_units, _ = get_full_indicators(ckey)
            n_ind = len(ind_names)
            n2    = n2_data[ckey]
            w_ind = merec_w[ckey]
            u     = n2 * w_ind[:, None]
            cat_total = u.sum(axis=0)
            shares = np.where(cat_total > 0, u / cat_total * 100, 0.0)
            psi_pct_this = np.where(T > 0,
                                    u * final_w[ci] / T * 100, 0.0)

            st.markdown(
                f"<span style='background:{cat['bg']};color:{cat['color']};"
                f"padding:2px 10px;border-radius:12px;font-size:13px;"
                f"font-weight:600;font-family:Times New Roman,Tinos,serif;'>"
                f"{cat['label']}</span>", unsafe_allow_html=True,
            )
            rows_ind = []
            for j in range(n_ind):
                row = {"Indicator": f"{ind_names[j]} ({ind_units[j]})",
                       "MEREC weight": round(float(w_ind[j]), 4)}
                for pi, pname in enumerate(names):
                    row[f"{pname} within-cat (%)"] = round(
                        float(shares[j, pi]), 1)
                for pi, pname in enumerate(names):
                    row[f"{pname} full (%)"] = round(
                        float(psi_pct_this[j, pi]), 1)
                rows_ind.append(row)
            tot_row = {"Indicator": "Total", "MEREC weight": ""}
            for pi, pname in enumerate(names):
                tot_row[f"{pname} within-cat (%)"] = 100.0
                tot_row[f"{pname} full (%)"] = round(
                    float(psi_pct_this[:, pi].sum()), 1)
            rows_ind.append(tot_row)
            st.dataframe(pd.DataFrame(rows_ind),
                         use_container_width=True, hide_index=True)
            st.write("")



def analytics_leave_one_out():
    st.subheader("B. Leave-one-out category analysis")
    
    names = st.session_state.proc_names
    n_proc = len(names)
    l3_cats = ordered_l3_cats()
    cat_scores = st.session_state.cat_scores
    sel_weight_methods = st.session_state.sel_weight_methods
    sel_mcdm_methods = list(st.session_state.sel_mcdm_methods) or ALL_MCDM_KEYS
    method_ranks = st.session_state.get("last_method_ranks", {})
    multi = len(sel_mcdm_methods) > 1

    if len(l3_cats) < 2:
        st.warning("Need at least 2 categories selected for leave-one-out analysis.")
        return

    p_val = st.slider("p value (PSI)", min_value=0.0, max_value=1.0,
                       value=0.5, step=0.01, key="loo_p_slider",
                       help=FIELD_HELP["psi_p"])

    # ── Baseline PSI ranks ────────────────────────────────────────────────────
    if multi and method_ranks:
        base_psi = calc_psi(method_ranks, sel_mcdm_methods, p_val)
        base_psi_ranks = rank_with_ties(base_psi, ascending=False)
    else:
        st.warning("Need at least 2 MCDM methods for PSI-based leave-one-out.")
        return

    # ── Leave-one-out runs ────────────────────────────────────────────────────
    loo_results = {}  # ckey -> psi_ranks array

    for excl_ci, excl_ckey in enumerate(l3_cats):
        reduced_cats = [c for c in l3_cats if c != excl_ckey]
        mat_red = np.array([cat_scores[c] for c in reduced_cats])
        w_red = get_category_weights(mat_red, sel_weight_methods)
        wm_red = mat_red * w_red[:, None]
        ranks_red = run_mcdm_suite(wm_red, w_red, sel_mcdm_methods)
        psi_red = calc_psi(ranks_red, sel_mcdm_methods, p_val)
        loo_results[excl_ckey] = rank_with_ties(psi_red, ascending=False)

    # ── Rank change matrix ────────────────────────────────────────────────────
    st.markdown("**Rank change when each category is excluded**")

    matrix_rows = []
    any_change = False
    for excl_ckey in l3_cats:
        row = {"Category excluded": CATS[excl_ckey]["label"]}
        for pi, name in enumerate(names):
            base_r = int(base_psi_ranks[pi])
            new_r = int(loo_results[excl_ckey][pi])
            delta = new_r - base_r
            if delta != 0:
                any_change = True
            row[f"{name} (Δ rank)"] = (f"+{delta}" if delta > 0 else
                                        str(delta) if delta != 0 else "—")
        matrix_rows.append(row)

    df_matrix = pd.DataFrame(matrix_rows)
    st.dataframe(df_matrix, use_container_width=True, hide_index=True)

    # ── Heatmap ───────────────────────────────────────────────────────────────
    z_vals = []
    for excl_ckey in l3_cats:
        row_z = []
        for pi in range(n_proc):
            delta = int(loo_results[excl_ckey][pi]) - int(base_psi_ranks[pi])
            row_z.append(delta)
        z_vals.append(row_z)



    # ── Within-category ranking ───────────────────────────────────────────────
    st.divider()
    st.markdown("**Within-category ranking**")

    wc_rows = []
    for ckey in l3_cats:
        scores = cat_scores[ckey]
        ranks = rank_with_ties(scores, ascending=False)
        row = {"Category": CATS[ckey]["label"]}
        for pi, name in enumerate(names):
            row[f"{name} (score)"] = round(float(scores[pi]), 4)
            row[f"{name} (rank)"] = int(ranks[pi])
        wc_rows.append(row)
    st.dataframe(pd.DataFrame(wc_rows), use_container_width=True, hide_index=True)

    # ── Interpretation ────────────────────────────────────────────────────────
    st.divider()
    if not any_change:
        st.markdown('<div style="background:#E8F5E9;border-left:3px solid #16A34A;padding:8px 12px;border-radius:4px;margin:6px 0;font-family:Times New Roman,Tinos,Times,serif;color:#1B5E20;font-size:13px;">No rank changes detected across all leave-one-out runs. The PSI ranking is robust to the exclusion of any single category.</div>', unsafe_allow_html=True)
    else:
        critical = []
        for excl_ckey in l3_cats:
            for pi in range(n_proc):
                delta = int(loo_results[excl_ckey][pi]) - int(base_psi_ranks[pi])
                if delta != 0:
                    critical.append(CATS[excl_ckey]["label"])
                    break
        critical = list(dict.fromkeys(critical))
        st.warning(
            f"⚠️ Rank changes detected when the following "
            f"{'category is' if len(critical) == 1 else 'categories are'} excluded: "
            f"**{', '.join(critical)}**. "
            f"{'This category is' if len(critical) == 1 else 'These categories are'} "
            f"critical to the current ranking."
        )


def analytics_category_intro():
    pass  # replaced by analytics_combined_contribution



def analytics_indicator_contribution():
    """
    Indicator contribution — nested donut (sunburst) chart per process.

    Outer ring: indicators sized by Pct^PSI_j,i = n2 × w^MEREC × w^RCW / T × 100
    Inner ring: categories sized by Pct_c,i = S_c,i / T_i × 100
    Centre text: process name

    Also shows summary tables below the charts.
    """
    st.subheader("A. Indicator contribution to weighted MCDM input")

    names    = st.session_state.proc_names
    n_proc   = len(names)
    l3_cats  = ordered_l3_cats()
    n2_data  = st.session_state.n2_data
    merec_w  = st.session_state.merec_w
    final_w  = st.session_state.final_cat_weights
    method_ranks = st.session_state.get("last_method_ranks", {})
    sel_mcdm = list(st.session_state.sel_mcdm_methods) or ALL_MCDM_KEYS
    multi    = len(sel_mcdm) > 1

    p_val = st.slider("p value (PSI)", min_value=0.0, max_value=1.0,
                       value=0.5, step=0.01, key="ind_contrib_p",
                       help=FIELD_HELP["psi_p"])

    if multi and method_ranks:
        psi_scores = calc_psi(method_ranks, sel_mcdm, p_val)
    else:
        psi_scores = np.ones(n_proc)

    # ── Pre-compute S_c,i and T_i ─────────────────────────────────────────────
    S_cat = np.array([
        st.session_state.cat_scores[c] * final_w[ci]
        for ci, c in enumerate(l3_cats)
    ])  # (n_cats, n_proc)
    T = S_cat.sum(axis=0)  # (n_proc,)

    # ── Build flat indicator lists for outer ring ─────────────────────────────
    # For each indicator: category key, label, full-PSI pct per process
    ind_labels  = []   # display label
    ind_units_l = []
    ind_cat_key = []   # which category it belongs to
    ind_psi_pct = []   # (n_ind_total, n_proc) — outer ring sizes
    ind_within  = []   # (n_ind_total, n_proc) — within-category share
    ind_merec_w = []

    for ci, ckey in enumerate(l3_cats):
        ind_names, ind_units, _ = get_full_indicators(ckey)
        n_ind = len(ind_names)
        n2    = n2_data[ckey]
        w_ind = merec_w[ckey]
        u = n2 * w_ind[:, None]
        cat_total = u.sum(axis=0)
        shares = np.where(cat_total > 0, u / cat_total * 100, 0.0)
        psi_pct = np.where(T > 0, u * final_w[ci] / T * 100, 0.0)

        for j in range(n_ind):
            ind_labels.append(ind_names[j])
            ind_units_l.append(ind_units[j])
            ind_cat_key.append(ckey)
            ind_psi_pct.append(psi_pct[j])
            ind_within.append(shares[j])
            ind_merec_w.append(float(w_ind[j]))

    ind_psi_pct = np.array(ind_psi_pct)   # (n_ind_total, n_proc)
    ind_within  = np.array(ind_within)

    # Category pct for inner ring
    cat_pct = np.where(T > 0, S_cat / T * 100, 0.0)  # (n_cats, n_proc)

    # ── Draw one nested donut per process ─────────────────────────────────────
    n_cats  = len(l3_cats)
    n_total = len(ind_labels)
    from matplotlib.patches import Patch

    # Short process labels for centre
    shorts_c = proc_short_labels(names)

    # Pre-build outer colours (category colour + alpha shading per indicator)
    cat_alpha_counts = {}
    for j in range(n_total):
        ck = ind_cat_key[j]
        cat_alpha_counts[ck] = cat_alpha_counts.get(ck, 0) + 1

    def make_outer_colors():
        cat_alpha_idx = {}
        cols = []
        for j in range(n_total):
            ck = ind_cat_key[j]
            idx   = cat_alpha_idx.get(ck, 0)
            count = cat_alpha_counts[ck]
            alpha = 0.42 + 0.58 * (idx / max(count - 1, 1))
            hex_c = CATS[ck]["color"].lstrip("#")
            r,g,b = (int(hex_c[i:i+2],16)/255 for i in (0,2,4))
            cols.append((r, g, b, alpha))
            cat_alpha_idx[ck] = idx + 1
        return cols

    outer_colors_base = make_outer_colors()

    # One large figure per process — stacked vertically for readability
    for pi, name in enumerate(names):

        apply_mpl_style()
        # Large square figure — one per process
        fig, ax = plt.subplots(figsize=(7.5, 7.5))

        inner_sizes  = [float(cat_pct[ci, pi]) for ci in range(n_cats)]
        inner_colors = [CATS[ckey]["color"] for ckey in l3_cats]
        inner_labels = [CATS[ckey]["label"] for ckey in l3_cats]
        outer_sizes  = [max(float(ind_psi_pct[j, pi]), 0.001)
                        for j in range(n_total)]

        # ── Inner ring — categories ──────────────────────────────────────────
        wedges_inner, _ = ax.pie(
            inner_sizes,
            radius=0.52,
            colors=inner_colors,
            startangle=90,
            wedgeprops=dict(width=0.26, edgecolor="white", linewidth=1.5),
            labels=None,
        )

        # ── Outer ring — indicators ──────────────────────────────────────────
        wedges_outer, _ = ax.pie(
            outer_sizes,
            radius=1.0,
            colors=outer_colors_base,
            startangle=90,
            wedgeprops=dict(width=0.44, edgecolor="white", linewidth=1.0),
            labels=None,
        )

        # ── Centre text — short process name + PSI ────────────────────────────
        ax.text(0, 0.07, shorts_c[pi],
                ha="center", va="center", fontsize=13, fontweight="bold",
                fontfamily="Times New Roman", color="#0D2B5E")
        if multi and method_ranks:
            ax.text(0, -0.13, f"PSI = {psi_scores[pi]:.4f}",
                    ha="center", va="center", fontsize=9,
                    fontfamily="Times New Roman", color="#444444")

        # ── Category labels on inner ring ─────────────────────────────────────
        for ci, (wedge, lbl, sz) in enumerate(
                zip(wedges_inner, inner_labels, inner_sizes)):
            if sz < 4:
                continue
            ang = (wedge.theta1 + wedge.theta2) / 2
            x = 0.39 * np.cos(np.deg2rad(ang))
            y = 0.39 * np.sin(np.deg2rad(ang))
            ax.text(x, y, lbl[:3], ha="center", va="center",
                    fontsize=7.5, fontweight="bold",
                    fontfamily="Times New Roman", color="white")

        # ── Data labels on outer ring — indicator short name + % ─────────────
        for j, (wedge, sz) in enumerate(zip(wedges_outer, outer_sizes)):
            # Show ALL slices — use shorter label for tiny ones
            ang  = (wedge.theta1 + wedge.theta2) / 2
            xl   = 1.14 * np.cos(np.deg2rad(ang))
            yl   = 1.14 * np.sin(np.deg2rad(ang))
            ha   = "left" if np.cos(np.deg2rad(ang)) >= 0 else "right"
            if sz < 2.0:
                label_txt = f"{sz:.1f}%"   # % only for tiny slices
            else:
                label_txt = f"{short_name(ind_labels[j])}  {sz:.1f}%"
            ax.text(xl, yl, label_txt,
                    ha=ha, va="center",
                    fontsize=7.5, fontfamily="Times New Roman",
                    color="#222222",
                    bbox=dict(boxstyle="round,pad=0.15", fc="white",
                              ec="none", alpha=0.85))

        # ── Indicator legend — right side ─────────────────────────────────────
        legend_patches = []
        for j in range(n_total):
            lbl = f"{ind_labels[j]} ({ind_units_l[j]})  {float(ind_psi_pct[j,pi]):.1f}%"
            legend_patches.append(
                Patch(facecolor=outer_colors_base[j], label=lbl)
            )
        ax.legend(
            handles=legend_patches,
            loc="center left",
            bbox_to_anchor=(1.02, 0.5),
            fontsize=7.5,
            frameon=True,
            framealpha=0.9,
            edgecolor="#cccccc",
            title="Indicators",
            title_fontsize=8,
            prop={"family": "Times New Roman", "size": 7.5},
        )

        ax.set_title(
            f"{name}",
            fontsize=12, fontweight="bold",
            fontfamily="Times New Roman", color="#0D2B5E", pad=10,
        )
        ax.set_aspect("equal")
        plt.tight_layout()
        mpl_show(fig)

    # ── Summary tables below charts ───────────────────────────────────────────
    with st.expander("Detailed tables"):
        for ci, ckey in enumerate(l3_cats):
            cat = CATS[ckey]
            ind_names, ind_units, _ = get_full_indicators(ckey)
            n_ind = len(ind_names)
            n2    = n2_data[ckey]
            w_ind = merec_w[ckey]
            u = n2 * w_ind[:, None]
            cat_total = u.sum(axis=0)
            shares = np.where(cat_total > 0, u / cat_total * 100, 0.0)
            psi_pct_this = np.where(T > 0, u * final_w[ci] / T * 100, 0.0)

            st.markdown(
                f"<span style='background:{cat['bg']};color:{cat['color']};"
                f"padding:2px 10px;border-radius:12px;font-size:13px;"
                f"font-weight:600;font-family:Times New Roman,Tinos,serif;'>"
                f"{cat['label']}</span>", unsafe_allow_html=True,
            )
            rows_s = []
            for j in range(n_ind):
                row = {"Indicator": f"{ind_names[j]} ({ind_units[j]})",
                       "MEREC weight": round(float(w_ind[j]), 4)}
                for pi, pname in enumerate(names):
                    row[f"{pname} within-cat (%)"] = round(float(shares[j, pi]), 1)
                for pi, pname in enumerate(names):
                    row[f"{pname} full (%)"] = round(float(psi_pct_this[j, pi]), 1)
                rows_s.append(row)
            total_row = {"Indicator": "Total", "MEREC weight": ""}
            for pi, pname in enumerate(names):
                total_row[f"{pname} within-cat (%)"] = 100.0
                total_row[f"{pname} full (%)"] = round(float(psi_pct_this[:, pi].sum()), 1)
            rows_s.append(total_row)
            st.dataframe(pd.DataFrame(rows_s),
                         use_container_width=True, hide_index=True)
            st.write("")


def analytics_indicator_loo():
    st.subheader("B. Leave-one-out indicator analysis")
    
    names = st.session_state.proc_names
    n_proc = len(names)
    l3_cats = ordered_l3_cats()
    cat_scores = st.session_state.cat_scores
    sel_weight_methods = st.session_state.sel_weight_methods
    sel_mcdm_methods = list(st.session_state.sel_mcdm_methods) or ALL_MCDM_KEYS
    method_ranks = st.session_state.get("last_method_ranks", {})
    multi = len(sel_mcdm_methods) > 1

    if not multi:
        st.warning("Need at least 2 MCDM methods for PSI-based leave-one-out.")
        return

    p_val = st.slider("p value (PSI)", min_value=0.0, max_value=1.0,
                       value=0.5, step=0.01, key="ind_loo_p",
                       help=FIELD_HELP["psi_p"])

    # Baseline PSI ranks
    base_psi = calc_psi(method_ranks, sel_mcdm_methods, p_val)
    base_psi_ranks = rank_with_ties(base_psi, ascending=False)

    # Run leave-one-out per indicator per category
    def run_loo(excl_ckey, excl_j):
        """Rerun pipeline excluding indicator j from category excl_ckey."""
        new_cat_scores = {}
        for ckey in l3_cats:
            ind_names_, ind_units_, benefits_ = get_full_indicators(ckey)
            n_ind_ = len(ind_names_)
            raw, _ = get_raw_matrix(ckey)

            if ckey == excl_ckey and n_ind_ > 1:
                # Remove row j
                keep = [jj for jj in range(n_ind_) if jj != excl_j]
                raw_red = raw[keep, :]
                ben_red = [benefits_[jj] for jj in keep]
                nm = np.zeros_like(raw_red)
                n2 = np.zeros_like(raw_red)
                for jj in range(len(keep)):
                    nm[jj] = merec_norm(raw_red[jj], ben_red[jj])
                    n2[jj] = n2_norm(raw_red[jj], ben_red[jj])
                w = merec_weights(nm)
                new_cat_scores[ckey] = (n2 * w[:, None]).sum(axis=0)
            else:
                new_cat_scores[ckey] = cat_scores[ckey]

        mat = np.array([new_cat_scores[c] for c in l3_cats])
        w_cat = get_category_weights(mat, sel_weight_methods)
        wm = mat * w_cat[:, None]
        ranks_ = run_mcdm_suite(wm, w_cat, sel_mcdm_methods)
        psi_ = calc_psi(ranks_, sel_mcdm_methods, p_val)
        return rank_with_ties(psi_, ascending=False)

    any_change = False

    for ckey in l3_cats:
        cat = CATS[ckey]
        ind_names, ind_units, _ = get_full_indicators(ckey)
        n_ind = len(ind_names)

        if n_ind < 2:
            st.markdown(
                f"<span style='background:{cat['bg']};color:{cat['color']};"
                f"padding:2px 10px;border-radius:12px;font-size:13px;font-weight:600;"
                f"font-family:Times New Roman,Tinos,Times,serif;'>"
                f"{cat['label']}</span> — only one indicator, skip.", 
                unsafe_allow_html=True,
            )
            st.write("")
            continue

        st.markdown(
            f"<span style='background:{cat['bg']};color:{cat['color']};"
            f"padding:2px 10px;border-radius:12px;font-size:13px;font-weight:600;"
            f"font-family:Times New Roman,Tinos,Times,serif;'>"
            f"{cat['label']}</span>", unsafe_allow_html=True,
        )

        # Build rank change matrix for this category
        matrix_rows = []
        z_vals = []
        ind_labels = [f"{ind_names[j]} ({ind_units[j]})" for j in range(n_ind)]

        for j in range(n_ind):
            new_ranks = run_loo(ckey, j)
            row = {"Indicator excluded": ind_labels[j]}
            row_z = []
            for pi, name in enumerate(names):
                delta = int(new_ranks[pi]) - int(base_psi_ranks[pi])
                if delta != 0:
                    any_change = True
                row[f"{name} (Δ rank)"] = (
                    f"+{delta}" if delta > 0 else str(delta) if delta != 0 else "—"
                )
                row_z.append(delta)
            matrix_rows.append(row)
            z_vals.append(row_z)

        st.dataframe(pd.DataFrame(matrix_rows),
                     use_container_width=True, hide_index=True)

        # Heatmap per category

        st.write("")

    # Overall interpretation
    st.divider()
    if not any_change:
        st.markdown('<div style="background:#E8F5E9;border-left:3px solid #16A34A;padding:8px 12px;border-radius:4px;margin:6px 0;font-family:Times New Roman,Tinos,Times,serif;color:#1B5E20;font-size:13px;">No rank changes detected across all indicator leave-one-out runs. The PSI ranking is robust to the exclusion of any single indicator.</div>', unsafe_allow_html=True)
    else:
        st.warning(
            "⚠️ One or more rank changes detected. "
            "Indicators highlighted in red above are critical to the current ranking — "
            "their removal alters the final alternative order."
        )


def analytics_indicator_intro():
    pass  # replaced by analytics_combined_contribution



def analytics_stakeholder_preference():
    st.subheader("3. Stakeholder preference simulation")
    
    names = st.session_state.proc_names
    n_proc = len(names)
    l3_cats = ordered_l3_cats()
    cat_scores = st.session_state.cat_scores
    sel_mcdm_methods = list(st.session_state.sel_mcdm_methods) or ALL_MCDM_KEYS
    method_ranks = st.session_state.get("last_method_ranks", {})
    final_w = st.session_state.final_cat_weights
    multi = len(sel_mcdm_methods) > 1

    cat_labels = [CATS[c]["label"] for c in l3_cats]
    n_cats = len(l3_cats)

    stakeholder = "Stakeholder"

    # ── p value ───────────────────────────────────────────────────────────────
    p_val = st.slider("p value (PSI)", min_value=0.0, max_value=1.0,
                       value=0.5, step=0.01, key="sh_p_slider",
                       help=FIELD_HELP["psi_p"])

    # ── Category weight sliders ───────────────────────────────────────────────
    st.markdown("**Adjust category weights (%)**")

    custom_weights = {}
    for ci, ckey in enumerate(l3_cats):
        cat = CATS[ckey]
        c1, c2 = st.columns([1, 3])
        with c1:
            st.markdown(
                f"<span style='background:{cat['bg']};color:{cat['color']};"
                f"padding:2px 10px;border-radius:10px;font-size:13px;font-weight:600;"
                f"font-family:Times New Roman,Tinos,Times,serif;'>"
                f"{cat['label']}</span>", unsafe_allow_html=True,
            )
        with c2:
            default_val = int(round(float(final_w[ci]) * 100))
            val = st.slider(
                cat["label"], min_value=0, max_value=100,
                value=st.session_state.get(f"sh_w_{ckey}", default_val),
                step=1, key=f"sh_w_{ckey}",
                label_visibility="collapsed",
                help=FIELD_HELP["stakeholder_weight"],
            )
        custom_weights[ckey] = val

    total = sum(custom_weights.values())
    c1, c2 = st.columns([3, 1])
    with c1:
        st.progress(min(total / 100, 1.0))
    with c2:
        if total == 100:
            st.success(f"Total: {total}% ✅")
        else:
            st.error(f"Total: {total}%")

    if total != 100:
        st.error("Adjust weights to sum to exactly 100% before running.")
        return

    # ── Run button ────────────────────────────────────────────────────────────
    if not st.button("Run simulation →", type="primary", key="sh_run"):
            return

    # ── Original ranking ──────────────────────────────────────────────────────
    if multi and method_ranks:
        base_psi = calc_psi(method_ranks, sel_mcdm_methods, p_val)
        base_psi_ranks = rank_with_ties(base_psi, ascending=False)
    else:
        st.warning("Need at least 2 MCDM methods for PSI ranking.")
        return

    # ── Rerun pipeline with custom weights ────────────────────────────────────
    w_custom = np.array([custom_weights[c] / 100.0 for c in l3_cats])
    mat = np.array([cat_scores[c] for c in l3_cats])
    wm_custom = mat * w_custom[:, None]
    ranks_custom = run_mcdm_suite(wm_custom, w_custom, sel_mcdm_methods)
    psi_custom = calc_psi(ranks_custom, sel_mcdm_methods, p_val)
    psi_ranks_custom = rank_with_ties(psi_custom, ascending=False)

    # ── Comparison table ──────────────────────────────────────────────────────
    st.divider()
    st.markdown(f"**Ranking comparison — Original RCW vs {stakeholder}**")

    comp_rows = []
    any_change = False
    for pi, name in enumerate(names):
        orig_r = int(base_psi_ranks[pi])
        new_r = int(psi_ranks_custom[pi])
        delta = new_r - orig_r
        if delta != 0:
            any_change = True
        comp_rows.append({
            "Alternative": name,
            "Original PSI rank (RCW)": orig_r,
            f"New PSI rank ({stakeholder})": new_r,
            "Rank change": (f"+{delta}" if delta > 0
                            else str(delta) if delta != 0 else "—"),
        })

    df_sh = pd.DataFrame(comp_rows)
    st.session_state["export_stakeholder"] = df_sh
    st.dataframe(df_sh, use_container_width=True, hide_index=True)

    if any_change:
        st.warning(
            f"⚠️ The ranking changes under {stakeholder}'s weights. "
            "The current RCW-based ranking is sensitive to this stakeholder's priorities."
        )
    else:
        st.markdown(f'<div style="background:#E8F5E9;border-left:3px solid #16A34A;padding:8px 12px;border-radius:4px;margin:6px 0;font-family:Times New Roman,Tinos,Times,serif;color:#1B5E20;font-size:13px;">The ranking is identical under the custom weights. The result is robust to this stakeholder profile.</div>', unsafe_allow_html=True)

    # ── Method-by-method breakdown ────────────────────────────────────────────
    st.divider()
    st.markdown(f"**Method-by-method ranking under {stakeholder}'s weights**")

    method_rows = []
    for pi, name in enumerate(names):
        row = {"Alternative": name,
               f"PSI rank ({stakeholder})": int(psi_ranks_custom[pi]),
               f"PSI score ({stakeholder})": round(float(psi_custom[pi]), 4)}
        for m in sel_mcdm_methods:
            row[METHOD_LABELS[m]] = int(ranks_custom[m][pi])
        method_rows.append(row)

    st.dataframe(pd.DataFrame(method_rows),
                 use_container_width=True, hide_index=True)

    # ── Weight comparison chart ───────────────────────────────────────────────
    st.divider()
    st.markdown("**Category weight comparison — RCW vs stakeholder**")

    apply_mpl_style()
    x_w = np.arange(n_cats)
    w_bar = 0.35
    fig, ax = plt.subplots(figsize=(max(5, n_cats*1.2), 3.0))
    rcw_vals = [round(float(final_w[ci])*100, 1) for ci in range(n_cats)]
    sh_vals  = [custom_weights[c] for c in l3_cats]
    ax.bar(x_w - w_bar/2, rcw_vals, width=w_bar,
           color="#185FA5", label="RCW (original)", edgecolor="white")
    ax.bar(x_w + w_bar/2, sh_vals,  width=w_bar,
           color="#EA580C", label="Stakeholder",    edgecolor="white")
    for xi, (rv, sv) in enumerate(zip(rcw_vals, sh_vals)):
        ax.text(xi - w_bar/2, rv + 0.5, f"{rv}%", ha="center", fontsize=8, color="#0D2B5E")
        ax.text(xi + w_bar/2, sv + 0.5, f"{sv}%", ha="center", fontsize=8, color="#0D2B5E")
    ax.set_xticks(x_w)
    ax.set_xticklabels(cat_labels, fontsize=9)
    ax.set_ylabel("Weight (%)", fontsize=10)
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, 1.12),
              ncol=2, fontsize=9, frameon=False)
    ax.tick_params(labelsize=9)
    plt.tight_layout()
    mpl_show(fig)

    # ── PSI score comparison chart ────────────────────────────────────────────
    st.markdown("**PSI score comparison — Original vs stakeholder**")

    apply_mpl_style()
    shorts = proc_short_labels(names)
    x_p = np.arange(n_proc)
    w_bar_p = 0.35
    fig, ax = plt.subplots(figsize=(max(5, n_proc*1.2), 3.0))
    base_vals = [round(float(base_psi[pi]), 4) for pi in range(n_proc)]
    cust_vals  = [round(float(psi_custom[pi]), 4) for pi in range(n_proc)]
    ax.bar(x_p - w_bar_p/2, base_vals, width=w_bar_p,
           color="#185FA5", label="Original PSI (RCW)", edgecolor="white")
    ax.bar(x_p + w_bar_p/2, cust_vals, width=w_bar_p,
           color="#EA580C", label="Stakeholder PSI",    edgecolor="white")
    ax.set_xticks(x_p)
    ax.set_xticklabels(shorts, fontsize=9)
    ax.set_ylabel("PSI score", fontsize=10)
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, 1.12),
              ncol=2, fontsize=9, frameon=False)
    ax.tick_params(labelsize=9)
    plt.tight_layout()
    mpl_show(fig)

    # Values in table
    psi_table_rows = []
    for pi, name in enumerate(names):
        psi_table_rows.append({
            "Alternative": name,
            "Original PSI (RCW)": round(float(base_psi[pi]), 4),
            "Stakeholder PSI": round(float(psi_custom[pi]), 4),
            "Δ PSI": round(float(psi_custom[pi]) - float(base_psi[pi]), 4),
        })
    st.dataframe(pd.DataFrame(psi_table_rows), use_container_width=True, hide_index=True)



def auxiliary_intro():
    st.header("Step 14 - Analytics (optional)")

    choice = st.radio(
        "Select tool",
        [
            "None - skip",
            "1. Contribution analysis",
            "2. Leave-one-out: categories",
            "3. Leave-one-out: indicators",
            "4. Stakeholder preference simulation",
        ],
        index=0, key="auxiliary_radio",
        help=FIELD_HELP["analytics_tool"],
    )

    if choice.startswith("1."):
        analytics_combined_contribution()
    elif choice.startswith("2."):
        analytics_leave_one_out()
    elif choice.startswith("3."):
        analytics_indicator_loo()
    elif choice.startswith("4."):
        analytics_stakeholder_preference()

    st.divider()
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        if st.button("<- Back to validation"):
            st.session_state.step = 13
            st.rerun()
    with c2:
        if st.button("<- Back to results"):
            st.session_state.step = 12
            st.rerun()
    with c3:
        if st.button("Decision support ->", type="primary"):
            st.session_state.step = 15
            st.rerun()
    with c4:
        if st.button("Reset all"):
            reset_all()
            st.rerun()



def step15():
    st.header("Step 15 - Decision support")

    method_ranks = st.session_state.get("last_method_ranks")
    if not method_ranks:
        st.warning("Complete Step 12 first to generate rankings and decision support.")
        st.divider()
        c1, c2 = st.columns(2)
        with c1:
            if st.button("<- Back to results", key="ds_back_results_empty"):
                st.session_state.step = 12
                st.rerun()
        with c2:
            if st.button("Reset all", key="ds_reset_empty"):
                reset_all()
                st.rerun()
        return

    names = st.session_state.proc_names
    methods = list(st.session_state.sel_mcdm_methods) or list(method_ranks.keys())
    winner_info = _resolve_overall_winner(names, method_ranks, methods)
    summary, strengths, weaknesses, profile = _build_decision_support_tables(winner_info)

    st.subheader("Overall recommendation")
    st.dataframe(pd.DataFrame(summary), use_container_width=True, hide_index=True)

    st.subheader("Strengths — why the winner wins")
    st.dataframe(pd.DataFrame(strengths), use_container_width=True, hide_index=True)

    st.subheader("Weaknesses — where the winner fails")
    st.dataframe(pd.DataFrame(weaknesses), use_container_width=True, hide_index=True)

    st.subheader("Category score profile")
    st.dataframe(pd.DataFrame(profile), use_container_width=True, hide_index=True)

    st.divider()
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        if st.button("<- Back to results", key="ds_back_results"):
            st.session_state.step = 12
            st.rerun()
    with c2:
        if st.button("<- Back to validation", key="ds_back_validation"):
            st.session_state.step = 13
            st.rerun()
    with c3:
        if st.button("<- Back to analytics", key="ds_back_analytics"):
            st.session_state.step = 14
            st.rerun()
    with c4:
        if st.button("Reset all", key="ds_reset"):
            reset_all()
            st.rerun()


STEPS = {
    -3: compare_assessments_page,
    -2: method_descriptions_page,
    -1: how_to_use_page,
    0: landing_page, 1: step1, 2: step2, 3: step3, 4: step4, 5: step5, 6: step6,
    7: step7, 8: step8, 9: step9, 10: step10, 11: step11, 12: step12,
    13: validation_intro, 14: auxiliary_intro, 15: step15,
}

STEPS[st.session_state.step]()
